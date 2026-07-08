from flask import (
    Flask,
    render_template,
    request,
    redirect,
    session,
    send_file,
    jsonify,
    g
)
import shutil
from werkzeug.utils import secure_filename
import os
from openpyxl import Workbook
import sqlite3
import psycopg2
from dotenv import load_dotenv
import cv2
import qrcode
from datetime import datetime
from cloudinary_helper import upload_photo, upload_file
from email_helper import send_welcome_email, send_payment_confirmation_email

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Image,
    Table
)

from reportlab.lib.styles import (
    getSampleStyleSheet
)

from flask import send_file
from werkzeug.security import check_password_hash, generate_password_hash

load_dotenv()

app = Flask(__name__)

app.secret_key = "FCCI_SECRET_KEY"

# ── GLOBAL ERROR LOGGING ────────────────────────────────────
import logging
import traceback

logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# ── DEV PANEL LOG BUFFER ────────────────────────────────────
# In-memory na nag-iipon ng huling 200 log entries para
# makita sa /dev panel nang hindi na kailangang buksan
# ang Render logs
from collections import deque as _deque

LOG_BUFFER = _deque(maxlen=200)

class _BufferLogHandler(logging.Handler):
    def emit(self, record):
        try:
            LOG_BUFFER.append({
                "time": datetime.fromtimestamp(record.created).strftime("%H:%M:%S"),
                "level": record.levelname,
                "msg": record.getMessage()[:300]
            })
        except Exception:
            pass

logging.getLogger().addHandler(_BufferLogHandler())

# ── ERROR TRACKER ───────────────────────────────────────────
# Nag-iipon ng huling 50 errors na may BUONG traceback —
# para hindi na kailangang buksan ang Render logs
ERROR_BUFFER = _deque(maxlen=50)

# ── REQUEST TRACKER ─────────────────────────────────────────
# Huling 100 requests: route, status, bilis (ms)
REQUEST_BUFFER = _deque(maxlen=100)

import time as _req_time


@app.before_request
def _track_request_start():
    g._req_start = _req_time.time()


@app.after_request
def _track_request_end(response):
    try:
        path = request.path
        if not path.startswith("/static") and path != "/favicon.ico":
            duration = round(
                (_req_time.time() - getattr(g, "_req_start", _req_time.time())) * 1000, 1
            )
            REQUEST_BUFFER.append({
                "time": datetime.now().strftime("%H:%M:%S"),
                "method": request.method,
                "path": path[:80],
                "status": response.status_code,
                "ms": duration
            })
    except Exception:
        pass
    return response


@app.errorhandler(500)
def internal_error(error):
    # I-record ang buong traceback sa ERROR_BUFFER para
    # makita sa /dev panel nang may kumpletong detalye
    tb = traceback.format_exc()
    try:
        ERROR_BUFFER.append({
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "route": f"{request.method} {request.path}",
            "error": str(error)[:200],
            "traceback": tb[-3000:]
        })
    except Exception:
        pass
    logger.error(f"[500 ERROR] {request.path} — {error}\n{tb}")
    return """
    <div style='font-family:Arial;padding:40px;text-align:center;'>
      <h2>⚠️ May nangyaring error</h2>
      <p>Nagreport na kami ng problema. Subukan mong mag-reload.</p>
      <a href='/dashboard' style='color:#00562a;'>← Bumalik sa Dashboard</a>
    </div>
    """, 500


@app.errorhandler(404)
def not_found(error):
    return """
    <div style='font-family:Arial;padding:40px;text-align:center;'>
      <h2>🔍 Page Not Found</h2>
      <p>Hindi mahanap ang page na hinahanap mo.</p>
      <a href='/dashboard' style='color:#00562a;'>← Bumalik sa Dashboard</a>
    </div>
    """, 404

UPLOAD_FOLDER = "static/uploads"

app.config[
    "UPLOAD_FOLDER"
] = UPLOAD_FOLDER

DATABASE_URL = os.environ.get("DATABASE_URL")

# ── CONNECTION POOL ──────────────────────────────────────────
# Gumagamit ng connection pool para hindi na kailangang gumawa
# ng bagong koneksyon sa Supabase sa bawat request — mas mabilis!
from psycopg2 import pool as pg_pool

_db_pool = None

def get_pool():
    global _db_pool
    if _db_pool is None or _db_pool.closed:
        _db_pool = pg_pool.ThreadedConnectionPool(
            minconn=1,
            maxconn=5,
            dsn=DATABASE_URL
        )
    return _db_pool


def get_db():
    """
    Kumuha ng koneksyon mula sa connection pool.
    Mas mabilis kaysa gumawa ng bagong koneksyon sa bawat request.
    """
    try:
        return get_pool().getconn()
    except Exception:
        # Fallback sa direct connect kung may pool error
        import psycopg2
        return psycopg2.connect(DATABASE_URL)


def return_db(conn):
    """
    I-return ang koneksyon sa pool pagkatapos gamitin.
    Tawagan ito imbes na return_db(conn) para ma-reuse ang koneksyon.
    """
    try:
        get_pool().putconn(conn)
    except Exception:
        try:
            return_db(conn)
        except Exception:
            pass


def fetch_member_with_photo(cursor, member_id):
    """
    Kinukuha ang lahat ng columns ng isang member (SELECT *) PLUS
    ang photo_path mula sa member_photos table, idinadagdag bilang
    HULING ELEMENT ng tuple. Ginagawa itong list para magamit ang
    .append(), tapos ibinabalik bilang tuple para gumana pa rin ang
    member[index] access pattern sa templates.

    Kung walang member na nahanap, nagbabalik ng None.
    """
    cursor.execute("SELECT id, member_id, full_name, contact, address, registration_fee, member_since, email, birthday, date_registered, status, proof_of_payment FROM members WHERE member_id = %s", (member_id,))
    row = cursor.fetchone()

    if row is None:
        return None

    cursor.execute("""
    SELECT photo_path FROM member_photos
    WHERE member_id = %s
    ORDER BY id DESC LIMIT 1
    """, (member_id,))

    photo_row = cursor.fetchone()
    photo_path = photo_row[0] if photo_row else None

    return tuple(list(row) + [photo_path])


def fetch_all_members_with_photo(cursor, where_clause="", params=()):
    """
    Kinukuha ang lahat ng members (o filtered gamit ang where_clause)
    PLUS photo_path bilang huling column ng bawat row. Ginagamit ito
    sa mga listahan tulad ng registration_approval at members list.
    """
    query = "SELECT id, member_id, full_name, contact, address, registration_fee, member_since, email, birthday, date_registered, status, proof_of_payment FROM members"
    if where_clause:
        query += f" WHERE {where_clause}"

    cursor.execute(query, params)
    rows = cursor.fetchall()

    results = []

    for row in rows:
        cursor.execute("""
        SELECT photo_path FROM member_photos
        WHERE member_id = %s
        ORDER BY id DESC LIMIT 1
        """, (row[1],))  # row[1] = member_id column

        photo_row = cursor.fetchone()
        photo_path = photo_row[0] if photo_row else None

        results.append(tuple(list(row) + [photo_path]))

    return results


def download_photo_for_pdf(member_id):
    """
    Kunin ang Cloudinary photo URL ng member mula sa member_photos
    table, i-download sa temporary file, at ibalik ang local path
    para magamit ng ReportLab Image(). Nagbabalik ng None kung
    walang photo o nag-error ang download.
    """
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
        SELECT photo_path FROM member_photos
        WHERE member_id = %s
        ORDER BY id DESC LIMIT 1
        """, (member_id,))
        photo_row = cursor.fetchone()
        return_db(conn)

        if photo_row and photo_row[0]:
            photo_url = photo_row[0]
            if photo_url.startswith("http"):
                import urllib.request
                import tempfile
                tmp_photo = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg")
                urllib.request.urlretrieve(photo_url, tmp_photo.name)
                return tmp_photo.name
    except Exception as e:
        print(f"[PDF] Photo download error: {e}")

    return None


# ════════════════════════════════════════════════════════════
#  SHARED PDF HELPERS — modern design, logo, watermark, Korean
# ════════════════════════════════════════════════════════════

_KOREAN_FONT_READY = False

def _ensure_korean_font():
    """I-register ang Korean-compatible font (isang beses lang).
    Ginagamit para tama ang render ng Hangul (한글) sa PDF."""
    global _KOREAN_FONT_READY
    if _KOREAN_FONT_READY:
        return "HYGothic-Medium"
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
        pdfmetrics.registerFont(UnicodeCIDFont("HYSMyeongJo-Medium"))
        _KOREAN_FONT_READY = True
        return "HYGothic-Medium"
    except Exception as e:
        print(f"[PDF] Korean font error: {e}")
        return "Helvetica"


def _pdf_logo_path():
    """Hanapin ang FCCI logo file para sa PDF."""
    for p in ["static/fcci_logo.jpeg", "logo/fcci_logo.jpeg"]:
        if os.path.exists(p):
            return p
    return None


def _modern_pdf_header(doc_type_label):
    """Gumawa ng modern na HUD-style header table na may logo.
    Ibinabalik ang isang Table flowable."""
    from reportlab.platypus import Table, TableStyle, Paragraph, Image
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT

    logo_path = _pdf_logo_path()
    title_st = ParagraphStyle("phTitle", fontSize=21, textColor=colors.white,
                              fontName="Helvetica-Bold", alignment=TA_LEFT, leading=23)
    sub_st = ParagraphStyle("phSub", fontSize=7, textColor=colors.HexColor("#8fb8c9"),
                            alignment=TA_LEFT, leading=10)
    type_st = ParagraphStyle("phType", fontSize=8, textColor=colors.HexColor("#00d4c8"),
                             alignment=TA_RIGHT, fontName="Helvetica-Bold")

    text_cell = [Paragraph("FCCI", title_st),
                 Paragraph("FILIPINO COMMUNITY CENTER INTERNATIONAL", sub_st)]

    if logo_path:
        logo = Image(logo_path, width=16*mm, height=16*mm)
        row = [[logo, text_cell, Paragraph(f"◈ {doc_type_label}", type_st)]]
        col_widths = [20*mm, 116*mm, 34*mm]
    else:
        row = [[text_cell, Paragraph(f"◈ {doc_type_label}", type_st)]]
        col_widths = [136*mm, 34*mm]

    header = Table(row, colWidths=col_widths)
    header.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#0b2236")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 12),
        ("BOTTOMPADDING", (0,0), (-1,-1), 12),
        ("LEFTPADDING", (0,0), (0,0), 14),
        ("RIGHTPADDING", (-1,0), (-1,0), 14),
    ]))
    return header


def _watermark_canvas(canvas, doc):
    """Callback na nagdo-draw ng faded logo sa gitna ng bawat page
    (background watermark). Tinatawag ng SimpleDocTemplate.build()."""
    logo_path = _pdf_logo_path()
    if not logo_path:
        return
    try:
        from reportlab.lib.units import mm
        canvas.saveState()
        canvas.setFillAlpha(0.055)      # napaka-faint
        w, h = doc.pagesize
        size = 110 * mm
        canvas.drawImage(logo_path,
                         (w - size) / 2, (h - size) / 2,
                         width=size, height=size,
                         preserveAspectRatio=True, mask="auto")
        canvas.restoreState()
    except Exception as e:
        print(f"[PDF] Watermark error: {e}")


def _cert_star(c, cx, cy, r_out, r_in, fill):
    import math
    c.saveState()
    p = c.beginPath()
    for i in range(10):
        ang = math.pi/2 + i*math.pi/5
        r = r_out if i % 2 == 0 else r_in
        x = cx + r*math.cos(ang); y = cy + r*math.sin(ang)
        (p.moveTo if i == 0 else p.lineTo)(x, y)
    p.close()
    c.setFillColor(fill)
    c.drawPath(p, fill=1, stroke=0)
    c.restoreState()


def _cert_diamond(c, cx, cy, r, fill):
    c.saveState()
    p = c.beginPath()
    p.moveTo(cx, cy+r); p.lineTo(cx+r, cy); p.lineTo(cx, cy-r); p.lineTo(cx-r, cy)
    p.close()
    c.setFillColor(fill)
    c.drawPath(p, fill=1, stroke=0)
    c.restoreState()


def _cert_seal(c, cx, cy, r, gold, gold_light, logo_path):
    import math
    from reportlab.lib import colors
    from reportlab.lib.utils import ImageReader
    c.saveState()
    n = 22
    p = c.beginPath()
    for i in range(n*2):
        ang = 2*math.pi*i/(n*2)
        rad = r if i % 2 == 0 else r*0.9
        x = cx + rad*math.cos(ang); y = cy + rad*math.sin(ang)
        (p.moveTo if i == 0 else p.lineTo)(x, y)
    p.close()
    c.setFillColor(gold)
    c.drawPath(p, fill=1, stroke=0)
    c.setFillColor(gold_light)
    c.circle(cx, cy, r*0.80, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.circle(cx, cy, r*0.63, fill=1, stroke=0)
    if logo_path and os.path.exists(logo_path):
        img = ImageReader(logo_path)
        d = r*1.18
        c.saveState()
        clip = c.beginPath(); clip.circle(cx, cy, r*0.61)
        c.clipPath(clip, stroke=0, fill=0)
        c.drawImage(img, cx-d/2, cy-d/2, width=d, height=d,
                    preserveAspectRatio=True, mask="auto")
        c.restoreState()
    c.setFillColor(gold)
    tail_w = r*0.32
    for dx in (-1, 1):
        p2 = c.beginPath()
        x0 = cx + dx*tail_w*0.55
        p2.moveTo(x0-tail_w/2, cy-r*0.85)
        p2.lineTo(x0+tail_w/2, cy-r*0.85)
        p2.lineTo(x0+tail_w*0.35, cy-r*1.45)
        p2.lineTo(x0-tail_w*0.35, cy-r*1.45)
        p2.close()
        c.drawPath(p2, fill=1, stroke=0)
    c.restoreState()


def _cert_laurel(c, cx, cy, r, lines, navy, gold):
    import math
    from reportlab.lib.units import mm
    c.saveState()
    n = 8
    for side in (-1, 1):
        for i in range(n):
            t = i/(n-1)
            ang = math.radians(95 + t*170)
            x = cx - side*r*math.cos(ang)
            y = cy - r*math.sin(ang) + r*0.62
            taper = 0.55 + 0.55*math.sin(math.radians(t*180))
            c.saveState()
            c.translate(x, y)
            tangent_deg = math.degrees(ang) + 90
            c.rotate(tangent_deg * side)
            c.setFillColor(gold)
            lw_ = 2.3*mm*taper
            c.ellipse(-lw_, -0.75*mm*taper, lw_, 0.75*mm*taper, fill=1, stroke=0)
            c.restoreState()
    c.setFillColor(navy)
    c.setFont("Times-Bold", 9.3)
    for i, line in enumerate(lines):
        c.drawCentredString(cx, cy + (len(lines)-1)*3.6 - i*7.2, line)
    c.restoreState()


def _draw_elegant_certificate(buf, data):
    """Gumagawa ng elegant, landscape, gold/navy certificate
    (parang formal na donation certificate) at isinusulat papunta sa buf (BytesIO)."""
    import math
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdfcanvas

    NAVY = colors.HexColor("#0b1d3a")
    GOLD = colors.HexColor("#b8964f")
    GOLD_LIGHT = colors.HexColor("#e3cd97")
    CREAM = colors.HexColor("#fbf8f1")
    TEXT = colors.HexColor("#1a2a44")
    MUTED = colors.HexColor("#6b7a94")

    W, H = landscape(A4)
    logo_path = _pdf_logo_path()

    c = pdfcanvas.Canvas(buf, pagesize=landscape(A4))
    c.setFillColor(CREAM); c.rect(0, 0, W, H, fill=1, stroke=0)

    if logo_path and os.path.exists(logo_path):
        from reportlab.lib.utils import ImageReader
        c.saveState()
        c.setFillAlpha(0.07)
        wm_size = 150*mm
        c.drawImage(ImageReader(logo_path), W/2-wm_size/2, H/2-wm_size/2,
                    width=wm_size, height=wm_size,
                    preserveAspectRatio=True, mask="auto")
        c.restoreState()


    band = 8*mm
    c.setFillColor(NAVY)
    c.rect(0, 0, W, band, fill=1, stroke=0)
    c.rect(0, H-band, W, band, fill=1, stroke=0)
    c.setFillColor(GOLD)
    c.rect(0, band-0.9*mm, W, 0.9*mm, fill=1, stroke=0)
    c.rect(0, H-band, W, 0.9*mm, fill=1, stroke=0)

    m1 = band + 3*mm
    m2 = m1 + 3.2*mm
    c.setStrokeColor(GOLD); c.setLineWidth(1.4)
    c.rect(m1, m1, W-2*m1, H-2*m1, fill=0, stroke=1)
    c.setLineWidth(0.6)
    c.rect(m2, m2, W-2*m2, H-2*m2, fill=0, stroke=1)

    for sx, sy in ((1, 1), (-1, 1), (1, -1), (-1, -1)):
        cx0 = m2 + 3*mm if sx > 0 else W-m2-3*mm
        cy0 = m2 + 3*mm if sy > 0 else H-m2-3*mm
        _cert_diamond(c, cx0, cy0, 2.2*mm, GOLD)

    cx = W/2
    y = H - 31*mm
    logo_r = 14*mm
    c.setStrokeColor(GOLD); c.setLineWidth(1)
    c.circle(cx, y, logo_r+1.1*mm, fill=0, stroke=1)
    if logo_path and os.path.exists(logo_path):
        from reportlab.lib.utils import ImageReader
        img = ImageReader(logo_path)
        d = logo_r*2
        c.saveState()
        clip = c.beginPath(); clip.circle(cx, y, logo_r)
        c.clipPath(clip, stroke=0, fill=0)
        c.drawImage(img, cx-d/2, y-d/2, width=d, height=d,
                    preserveAspectRatio=True, mask="auto")
        c.restoreState()

    y -= (logo_r + 12*mm)
    c.setFillColor(NAVY); c.setFont("Times-Bold", 32)
    c.drawCentredString(cx, y, data.get("cert_title", "DONATION CERTIFICATE"))

    y -= 8*mm
    c.setStrokeColor(GOLD); c.setLineWidth(0.8)
    lw = 48*mm
    c.line(cx-lw, y, cx-6*mm, y); c.line(cx+6*mm, y, cx+lw, y)
    _cert_star(c, cx, y, 2.3*mm, 1.0*mm, GOLD)

    y -= 10*mm
    c.setFillColor(NAVY); c.setFont("Times-Italic", 13.5)
    c.drawCentredString(cx, y, "This certificate is proudly presented to")

    y -= 13*mm
    c.setFont("Times-Bold", 30)
    c.drawCentredString(cx, y, str(data.get("donor_name") or "-"))

    y -= 6*mm
    c.setStrokeColor(GOLD); c.setLineWidth(0.6)
    c.line(cx-56*mm, y, cx+56*mm, y)

    y -= 10*mm
    c.setFillColor(TEXT); c.setFont("Times-Italic", 12.5)
    c.drawCentredString(cx, y, "In sincere appreciation and gratitude for your generous donation")
    y -= 6.5*mm
    c.drawCentredString(cx, y, "and valuable support to the programs and activities of the")
    y -= 7.5*mm
    c.setFont("Times-Bold", 13); c.setFillColor(NAVY)
    c.drawCentredString(cx, y, "FILIPINO COMMUNITY CENTER INTERNATIONAL (FCCI)")

    y -= 9.5*mm
    c.setFont("Times-Italic", 12.5); c.setFillColor(TEXT)
    amt_line = f"in the amount of \u20a9{data['amount']:,}" if data.get("amount") else ""
    if data.get("purpose"):
        amt_line += (" for " if amt_line else "for ") + str(data["purpose"])
    if amt_line:
        c.drawCentredString(cx, y, amt_line)
        y -= 7*mm
    c.setFont("Times-Italic", 11.5); c.setFillColor(MUTED)
    c.drawCentredString(cx, y, "Your kindness and commitment help empower our community")
    y -= 5.5*mm
    c.drawCentredString(cx, y, "and make a lasting difference.")

    base_y = 35*mm
    _cert_seal(c, m2 + 23*mm, base_y+1*mm, 15.5*mm, GOLD, GOLD_LIGHT, logo_path)

    sig_x = cx - 52*mm
    c.setStrokeColor(NAVY); c.setLineWidth(0.7)
    c.line(sig_x-26*mm, base_y+3*mm, sig_x+26*mm, base_y+3*mm)
    c.setFont("Times-Bold", 11); c.setFillColor(NAVY)
    c.drawCentredString(sig_x, base_y-2.2*mm, data.get("signee", "Authorized Signatory"))
    c.setFont("Times-Roman", 9); c.setFillColor(MUTED)
    c.drawCentredString(sig_x, base_y-6.5*mm, data.get("signee_title", "Treasurer"))

    _cert_laurel(c, cx, base_y+7*mm, 12.5*mm, ["THANK", "YOU"], NAVY, GOLD)

    date_x = cx + 52*mm
    c.setStrokeColor(NAVY)
    c.line(date_x-26*mm, base_y+3*mm, date_x+26*mm, base_y+3*mm)
    c.setFont("Times-Bold", 11); c.setFillColor(NAVY)
    c.drawCentredString(date_x, base_y-2.2*mm, data.get("date_str", "-"))
    c.setFont("Times-Roman", 9); c.setFillColor(MUTED)
    c.drawCentredString(date_x, base_y-6.5*mm, "Date")

    if data.get("receipt_no"):
        c.setFont("Helvetica", 7); c.setFillColor(MUTED)
        footer = f"Receipt No. {data['receipt_no']}"
        if data.get("contact"):
            footer += f"   |   Contact: {data['contact']}"
        c.drawCentredString(cx, m2+3.2*mm, footer)

    c.showPage()
    c.save()


def _cert_arc_text(c, cx, cy, radius, text, font, size, color, start_deg, end_deg):
    """Ipinapatong ang text sa isang arc, letra-por-letra, pantay ang spacing sa anggulo."""
    import math
    c.saveState()
    c.setFillColor(color)
    c.setFont(font, size)
    n = len(text)
    if n == 0:
        c.restoreState(); return
    span = end_deg - start_deg
    step = span / max(n - 1, 1)
    for i, ch in enumerate(text):
        ang = math.radians(start_deg + i*step)
        x = cx + radius*math.cos(ang)
        y = cy + radius*math.sin(ang)
        c.saveState()
        c.translate(x, y)
        c.rotate(math.degrees(ang) - 90)
        c.drawCentredString(0, 0, ch)
        c.restoreState()
    c.restoreState()


def _cert_official_seal(c, cx, cy, r, blue):
    """Asul na 'Official Seal' stamp, parang notary ink stamp."""
    from reportlab.lib.units import mm
    c.saveState()
    c.setStrokeColor(blue); c.setLineWidth(1.3)
    c.circle(cx, cy, r, fill=0, stroke=1)
    c.setLineWidth(0.6)
    c.circle(cx, cy, r*0.82, fill=0, stroke=1)

    _cert_arc_text(c, cx, cy, r*0.91, "FILIPINO COMMUNITY CENTER",
                   "Helvetica-Bold", 4.6, blue, 200, 340)
    _cert_arc_text(c, cx, cy, r*0.91, "INTERNATIONAL",
                   "Helvetica-Bold", 4.6, blue, 20, 160)

    _cert_star(c, cx-r*0.68, cy+r*0.12, 1.4*mm, 0.6*mm, blue)
    _cert_star(c, cx+r*0.68, cy+r*0.12, 1.4*mm, 0.6*mm, blue)

    c.setFillColor(blue)
    c.setFont("Helvetica-Bold", 9.5)
    c.drawCentredString(cx, cy+2.3*mm, "OFFICIAL")
    c.drawCentredString(cx, cy-4.5*mm, "SEAL")
    c.setFont("Helvetica-Bold", 6.5)
    c.drawCentredString(cx, cy-1.6*mm, "FCCI")
    c.restoreState()


def _draw_withdrawal_certificate(buf, data):
    """Gumagawa ng elegant, landscape, gold/navy withdrawal certificate
    na may computation ng refund/community share, isinusulat papunta sa buf (BytesIO)."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as pdfcanvas

    NAVY = colors.HexColor("#0b1d3a")
    GOLD = colors.HexColor("#b8964f")
    GOLD_LIGHT = colors.HexColor("#e3cd97")
    CREAM = colors.HexColor("#fbf8f1")
    TEXT = colors.HexColor("#1a2a44")
    MUTED = colors.HexColor("#6b7a94")
    BLUE = colors.HexColor("#1f3f8f")
    BOX_BG = colors.HexColor("#f4efe2")
    REFUND_GREEN = colors.HexColor("#0d7a5f")

    W, H = landscape(A4)
    logo_path = _pdf_logo_path()

    c = pdfcanvas.Canvas(buf, pagesize=landscape(A4))
    c.setFillColor(CREAM); c.rect(0, 0, W, H, fill=1, stroke=0)

    if logo_path and os.path.exists(logo_path):
        from reportlab.lib.utils import ImageReader
        c.saveState()
        c.setFillAlpha(0.07)
        wm_size = 150*mm
        c.drawImage(ImageReader(logo_path), W/2-wm_size/2, H/2-wm_size/2,
                    width=wm_size, height=wm_size,
                    preserveAspectRatio=True, mask="auto")
        c.restoreState()

    band = 8*mm
    c.setFillColor(NAVY)
    c.rect(0, 0, W, band, fill=1, stroke=0)
    c.rect(0, H-band, W, band, fill=1, stroke=0)
    c.setFillColor(GOLD)
    c.rect(0, band-0.9*mm, W, 0.9*mm, fill=1, stroke=0)
    c.rect(0, H-band, W, 0.9*mm, fill=1, stroke=0)

    m1 = band + 3*mm
    m2 = m1 + 3.2*mm
    c.setStrokeColor(GOLD); c.setLineWidth(1.4)
    c.rect(m1, m1, W-2*m1, H-2*m1, fill=0, stroke=1)
    c.setLineWidth(0.6)
    c.rect(m2, m2, W-2*m2, H-2*m2, fill=0, stroke=1)

    for sx, sy in ((1, 1), (-1, 1), (1, -1), (-1, -1)):
        cx0 = m2 + 3*mm if sx > 0 else W-m2-3*mm
        cy0 = m2 + 3*mm if sy > 0 else H-m2-3*mm
        _cert_diamond(c, cx0, cy0, 2.2*mm, GOLD)

    cx = W/2
    y = H - 31*mm
    logo_r = 14*mm
    c.setStrokeColor(GOLD); c.setLineWidth(1)
    c.circle(cx, y, logo_r+1.1*mm, fill=0, stroke=1)
    if logo_path and os.path.exists(logo_path):
        from reportlab.lib.utils import ImageReader
        img = ImageReader(logo_path)
        d = logo_r*2
        c.saveState()
        clip = c.beginPath(); clip.circle(cx, y, logo_r)
        c.clipPath(clip, stroke=0, fill=0)
        c.drawImage(img, cx-d/2, y-d/2, width=d, height=d,
                    preserveAspectRatio=True, mask="auto")
        c.restoreState()

    y -= (logo_r + 12*mm)
    c.setFillColor(NAVY); c.setFont("Times-Bold", 30)
    c.drawCentredString(cx, y, "WITHDRAWAL CERTIFICATE")

    y -= 7.5*mm
    c.setStrokeColor(GOLD); c.setLineWidth(0.8)
    lw = 48*mm
    c.line(cx-lw, y, cx-6*mm, y); c.line(cx+6*mm, y, cx+lw, y)
    _cert_star(c, cx, y, 2.2*mm, 1.0*mm, GOLD)

    y -= 9.5*mm
    c.setFillColor(NAVY); c.setFont("Times-Italic", 13)
    c.drawCentredString(cx, y, "This is to certify that")

    y -= 12*mm
    c.setFont("Times-Bold", 27)
    c.drawCentredString(cx, y, str(data.get("member_name") or "-"))

    y -= 5.5*mm
    c.setStrokeColor(GOLD); c.setLineWidth(0.6)
    c.line(cx-56*mm, y, cx+56*mm, y)

    y -= 9*mm
    c.setFont("Times-Roman", 11.5); c.setFillColor(TEXT)
    c.drawCentredString(cx, y, "is a registered member of the FILIPINO COMMUNITY CENTER INTERNATIONAL (FCCI)")
    y -= 6*mm
    c.drawCentredString(cx, y, f"with Membership ID No. {data.get('member_id', '-')}.")
    y -= 7.5*mm
    c.setFont("Times-Italic", 11.5)
    c.drawCentredString(cx, y, "has formally submitted a request for withdrawal from membership")
    y -= 6*mm
    c.drawCentredString(cx, y, f"effective this {data.get('withdrawal_date', '-')}.")

    y -= 12*mm
    box_w = 172*mm
    box_h = 24*mm
    box_x = cx - box_w/2
    box_y = y - box_h
    c.setFillColor(BOX_BG)
    c.setStrokeColor(GOLD); c.setLineWidth(0.9)
    c.roundRect(box_x, box_y, box_w, box_h, 2.5*mm, fill=1, stroke=1)

    col_w = box_w/3
    for i in (1, 2):
        xline = box_x + col_w*i
        c.setStrokeColor(GOLD); c.setLineWidth(0.4)
        c.line(xline, box_y+3*mm, xline, box_y+box_h-3*mm)

    cols = [
        ("TOTAL CONTRIBUTIONS", f"\u20a9{data.get('total_contributions', 0):,}", NAVY),
        ("REFUND (75%)", f"\u20a9{data.get('refund_amount', 0):,}", REFUND_GREEN),
        ("COMMUNITY SHARE (25%)", f"\u20a9{data.get('community_share', 0):,}", NAVY),
    ]
    for i, (label, val, vcolor) in enumerate(cols):
        colcx = box_x + col_w*i + col_w/2
        c.setFillColor(MUTED)
        c.setFont("Helvetica-Bold", 7)
        c.drawCentredString(colcx, box_y+box_h-8*mm, label)
        c.setFillColor(vcolor)
        c.setFont("Times-Bold", 15)
        c.drawCentredString(colcx, box_y+6.5*mm, val)

    y = box_y - 7*mm
    c.setFont("Times-Italic", 9.6); c.setFillColor(MUTED)
    c.drawCentredString(cx, y, "This certification is issued upon the request of the above-named individual")
    y -= 4.6*mm
    c.drawCentredString(cx, y, "for whatever legal purpose it may serve.")

    base_y = 33*mm
    _cert_seal(c, m2 + 22*mm, base_y+1*mm, 15*mm, GOLD, GOLD_LIGHT, logo_path)
    _cert_official_seal(c, W - m2 - 22*mm, base_y+2*mm, 15*mm, BLUE)

    sig_x = cx - 33*mm
    c.setStrokeColor(NAVY); c.setLineWidth(0.7)
    c.line(sig_x-26*mm, base_y+3*mm, sig_x+26*mm, base_y+3*mm)
    c.setFont("Times-Bold", 10.6); c.setFillColor(NAVY)
    c.drawCentredString(sig_x, base_y-2.2*mm, data.get("signee", "Maria Santos"))
    c.setFont("Times-Roman", 8.6); c.setFillColor(MUTED)
    c.drawCentredString(sig_x, base_y-6.3*mm, data.get("signee_title", "President"))
    c.drawCentredString(sig_x, base_y-10*mm, "Filipino Community Center International (FCCI)")

    date_x = cx + 33*mm
    c.setStrokeColor(NAVY)
    c.line(date_x-22*mm, base_y+3*mm, date_x+22*mm, base_y+3*mm)
    c.setFont("Times-Bold", 10.6); c.setFillColor(NAVY)
    c.drawCentredString(date_x, base_y-2.2*mm, data.get("withdrawal_date", "-"))
    c.setFont("Times-Roman", 8.6); c.setFillColor(MUTED)
    c.drawCentredString(date_x, base_y-6.3*mm, "Date")

    c.setFont("Helvetica", 7); c.setFillColor(MUTED)
    c.drawCentredString(cx, m2+3.2*mm,
        f"Member ID: {data.get('member_id','-')}   |   Generated: {data.get('generated','-')}")

    c.showPage()
    c.save()


@app.route("/")
def home():
    return redirect("/login")


@app.route(
    "/login",
    methods=["GET", "POST"]
)
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        # ── HIDDEN DEVELOPER ACCOUNT ─────────────────────────
        # Code-level account — WALA sa database, hindi makikita
        # sa Settings user list, hindi ma-e-edit o ma-de-delete.
        # SHA-256 hash ang naka-store, hindi plaintext.
        import hashlib as _hl
        if (username == "paulo20" and
                _hl.sha256(password.encode()).hexdigest() ==
                "346dedb24bec0911ed3fe4b9a6e03543754e10e0d3e2e955e323bb21a3809eb1"):
            session["username"] = "paulo20"
            session["role"] = "admin"
            session["is_developer"] = True
            return redirect("/dashboard")

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT role, password
        FROM users
        WHERE username = %s
        """, (username,))

        user = cursor.fetchone()
        return_db(conn)

        # Supports both hashed and plain passwords (para hindi masisira ang existing users)
        if user:
            stored_password = user[1]
            role = user[0]

            # Check if password is already hashed (starts with 'pbkdf2:' or 'scrypt:')
            if stored_password.startswith("pbkdf2:") or stored_password.startswith("scrypt:"):
                password_ok = check_password_hash(stored_password, password)
            else:
                # Plain text fallback (old accounts)
                password_ok = (stored_password == password)

            if password_ok:
                session["username"] = username
                session["role"] = role
                return redirect("/dashboard")

        return render_template(
            "login.html",
            error="Invalid username or password."
        )

    return render_template("login.html")

@app.route(
    "/member_registration",
    methods=["GET","POST"]
)
def member_registration():

    if request.method == "POST":

        full_name          = request.form["full_name"]
        contact            = request.form["contact"]
        birthday           = request.form["birthday"]
        email              = request.form["email"]
        address            = request.form["address"]
        member_since_input = request.form.get("member_since", "").strip()
        photo              = request.files.get("photo")

        # ── DUPLICATE CHECK BAGO MAG-UPLOAD ─────────────────────
        # I-check muna kung may existing na member na may parehong
        # contact o full_name — bago pa mag-upload sa Cloudinary
        # para hindi masayang ang bandwidth at oras
        conn   = get_db()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT member_id, full_name, status
        FROM members
        WHERE contact = %s OR lower(full_name) = lower(%s)
        LIMIT 1
        """, (contact, full_name))

        existing = cursor.fetchone()
        return_db(conn)

        if existing:
            existing_member_id = existing[0]
            existing_name      = existing[1]
            existing_status    = existing[2]

            # Kung Applicant pa — i-redirect sa kanyang
            # registration confirmation page
            if existing_status == "Applicant":
                return redirect(
                    f"/registration_confirmation/{existing_member_id}"
                )

            # Kung Active na — mag-show ng friendly message
            return render_template(
                "member_registration.html",
                error=f"Mukhang naka-register ka na! "
                      f"Ang '{existing_name}' ay {existing_status} na. "
                      f"Makipag-ugnayan sa FCCI admin kung may katanungan ka."
            )

        # Wala pang existing — i-upload ang photo at i-register
        photo_filename = ""
        if photo and photo.filename:
            photo_filename = upload_photo(photo, folder="fcci_member_photos") or ""

        # Gamitin ang Supabase advisory lock para maiwasan ang
        # race condition kapag sabay-sabay na nag-reregister.
        # Ang SELECT + INSERT ay ginagawa sa loob ng single transaction
        # na may FOR UPDATE SKIP LOCKED para atomic ito.
        max_retries = 5
        member_id = None

        for attempt in range(max_retries):
            conn = get_db()
            cursor = conn.cursor()
            try:
                # Advisory lock — nagsi-serialize ng sabay-sabay na
                # registrations. (Bawal ang FOR UPDATE kasama ng MAX()
                # sa PostgreSQL, kaya ito ang tamang paraan.)
                # Awtomatikong nare-release sa commit/rollback.
                cursor.execute("SELECT pg_advisory_xact_lock(202600)")

                cursor.execute("""
                SELECT COALESCE(
                    MAX(CAST(SPLIT_PART(member_id, '-', 3) AS INTEGER)),
                    0
                )
                FROM members
                WHERE member_id LIKE 'APP-%%'
                """)
                highest_num = cursor.fetchone()[0]
                count       = highest_num + 1
                member_id   = f"APP-{datetime.now().year}-{count:06d}"

                cursor.execute("""
                INSERT INTO members
                (
                    member_id, full_name, contact, address,
                    registration_fee, member_since, email,
                    birthday, date_registered, status
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    member_id, full_name, contact, address,
                    0, member_since_input, email, birthday,
                    datetime.now().strftime("%Y-%m-%d"), "Applicant"
                ))

                if photo_filename:
                    cursor.execute("""
                    INSERT INTO member_photos (member_id, photo_path)
                    VALUES (%s, %s)
                    """, (member_id, photo_filename))

                conn.commit()
                return_db(conn)
                break  # Matagumpay — lumabas sa retry loop

            except Exception as e:
                conn.rollback()
                return_db(conn)
                if "unique" in str(e).lower() and attempt < max_retries - 1:
                    # Duplicate member_id — subukan ulit
                    import time
                    time.sleep(0.1)
                    member_id = None
                    continue
                else:
                    raise

        if not member_id:
            return "Registration failed. Please try again.", 500

        return redirect(f"/registration_confirmation/{member_id}")

    return render_template(
        "member_registration.html"
    )

@app.route("/dashboard")
def dashboard():

    if "username" not in session:
        return redirect("/login")

    try:
        conn = get_db()
        cursor = conn.cursor()

        # ── OPTIMIZED: Lahat ng stats sa ISANG query ────────────
        # Imbes na 5 hiwalay na queries, isang query na lang
        cursor.execute("""
        SELECT
            (SELECT COUNT(*) FROM members)                              AS total_members,
            (SELECT COALESCE(SUM(amount),0) FROM payments)             AS collections,
            (SELECT COALESCE(SUM(amount),0) FROM donations)            AS donations,
            (SELECT COALESCE(SUM(amount),0) FROM expenses)             AS expenses,
            (SELECT COUNT(*) FROM members WHERE status = 'Applicant')  AS applicants,
            (SELECT COUNT(*) FROM members WHERE status = 'Active')     AS active_count
        """)
        stats = cursor.fetchone()
        total_members = stats[0] or 0
        collections   = stats[1] or 0
        donations     = stats[2] or 0
        expenses      = stats[3] or 0
        applicants    = stats[4] or 0
        balance       = collections + donations - expenses

        # ── OPTIMIZED: Outstanding — isang query para sa lahat ──
        # Imbes na mag-loop at mag-query sa bawat member/buwan,
        # kunin lahat ng payments nang sabay at i-process sa Python
        cursor.execute("""
        SELECT member_id, member_since FROM members WHERE status = 'Active'
        """)
        active_members_list = cursor.fetchall()

        # Kunin lahat ng Monthly Contribution payments nang sabay
        cursor.execute("""
        SELECT member_id, payment_month, payment_year
        FROM payments
        WHERE payment_type = 'Monthly Contribution'
        """)
        all_mc_payments = cursor.fetchall()

        # Gumawa ng set para sa mabilis na lookup
        paid_set = set()
        for row in all_mc_payments:
            paid_set.add((row[0], row[1], str(row[2])))

        current_date = datetime.now()
        month_map = {
            "January":1,"February":2,"March":3,"April":4,
            "May":5,"June":6,"July":7,"August":8,
            "September":9,"October":10,"November":11,"December":12
        }

        active_members  = 0
        inactive_members = 0
        total_outstanding = 0

        for member in active_members_list:
            member_id    = member[0]
            member_since = member[1]
            if not member_since:
                continue
            try:
                parts = member_since.split()
                month = month_map[parts[0]]
                year  = int(parts[1])
            except:
                continue

            missing_count = 0
            y, m = year, month
            while (y < current_date.year or
                   (y == current_date.year and m <= current_date.month)):
                month_name = datetime(y, m, 1).strftime("%B")
                if (member_id, month_name, str(y)) not in paid_set:
                    missing_count += 1
                m += 1
                if m > 12:
                    m = 1
                    y += 1

            if missing_count < 5:
                active_members += 1
            else:
                inactive_members += 1
            total_outstanding += (missing_count * 10000)

        # ── RECENT PAYMENTS ──────────────────────────────────────
        cursor.execute("""
        SELECT p.member_id, p.amount, m.full_name,
               p.payment_type, p.payment_month, p.payment_year
        FROM payments p
        LEFT JOIN members m ON p.member_id = m.member_id
        ORDER BY p.id DESC LIMIT 10
        """)
        recent_payments = cursor.fetchall()

        # ── BIRTHDAY THIS MONTH ──────────────────────────────────
        current_month = current_date.month
        current_day   = current_date.day

        cursor.execute("""
        SELECT m.full_name, m.birthday, mp.photo_path
        FROM members m
        LEFT JOIN member_photos mp ON m.member_id = mp.member_id
        WHERE m.status = 'Active'
        AND m.birthday IS NOT NULL AND m.birthday != ''
        ORDER BY m.birthday
        """)
        all_members  = cursor.fetchall()
        birthday_list = []

        for row in all_members:
            full_name  = row[0]
            birthday   = row[1]
            photo_path = row[2]
            if not birthday:
                continue
            try:
                bday_obj = datetime.strptime(birthday, "%Y-%m-%d")
            except ValueError:
                try:
                    bday_obj = datetime.strptime(birthday, "%m/%d/%Y")
                except ValueError:
                    continue
            if bday_obj.month == current_month:
                is_today = (bday_obj.day == current_day)
                birthday_list.append({
                    "full_name": full_name,
                    "birthday":  bday_obj.strftime("%B %d"),
                    "photo":     photo_path or "",
                    "is_today":  is_today,
                    "day":       bday_obj.day
                })

        birthday_list.sort(key=lambda x: (not x["is_today"], x["day"]))

        return_db(conn)

    except Exception as e:
        import traceback
        print(f"[DASHBOARD ERROR] {e}")
        print(traceback.format_exc())
        total_members = collections = donations = expenses = 0
        balance = applicants = active_members = 0
        inactive_members = total_outstanding = 0
        recent_payments = []
        birthday_list = []

    return render_template(
        "dashboard.html",
        username=session["username"],
        total_members=total_members,
        active_members=active_members,
        inactive_members=inactive_members,
        applicants=applicants,
        collections=collections,
        donations=donations,
        expenses=expenses,
        balance=balance,
        total_outstanding=total_outstanding,
        birthday_list=birthday_list,
        recent_payments=recent_payments
    )

@app.route("/members")
def members():

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    search = request.args.get(
        "search",
        ""
    )

    if search:

        cursor.execute("""
        SELECT
            member_id,
            full_name,
            contact,
            email,
            member_since
        FROM members
        WHERE
            member_id ILIKE %s
            OR full_name ILIKE %s
            OR contact ILIKE %s
        ORDER BY full_name
        """, (
            f"%{search}%",
            f"%{search}%",
            f"%{search}%"
        ))

    else:

        cursor.execute("""
        SELECT
            member_id,
            full_name,
            contact,
            email,
            member_since
        FROM members
        ORDER BY full_name
        """)

    members = cursor.fetchall()

    updated_members = []

    month_map = {
        "January":1,
        "February":2,
        "March":3,
        "April":4,
        "May":5,
        "June":6,
        "July":7,
        "August":8,
        "September":9,
        "October":10,
        "November":11,
        "December":12
    }

    current_date = datetime.now()

    for member in members:

        member_id = member[0]
        full_name = member[1]
        contact = member[2]
        email = member[3]
        member_since = member[4]

        status = "Applicant"

        if member_since:

            try:

                month_name, year = member_since.split()

                year = int(year)

                month = month_map[month_name]

                missing_count = 0

                temp_year = year
                temp_month = month

                while (
                    temp_year < current_date.year
                    or
                    (
                        temp_year == current_date.year
                        and temp_month <= current_date.month
                    )
                ):

                    current_month = datetime(
                        temp_year,
                        temp_month,
                        1
                   ).strftime("%B")

                    cursor.execute("""
                    SELECT COUNT(*)
                    FROM payments
                    WHERE member_id=%s
                    AND payment_type='Monthly Contribution'
                    AND payment_month=%s
                    AND payment_year=%s
                    """,(
                       member_id,
                       current_month,
                       str(temp_year)
                    ))

                    paid = cursor.fetchone()[0]

                    if paid == 0:
                        missing_count += 1

                    temp_month += 1

                    if temp_month > 12:
                        temp_month = 1
                        temp_year += 1

                if missing_count >= 5:
                    status = "Inactive"
                else:
                    status = "Active"

            except:
                status = "Active"

        updated_members.append(
            (
                member_id,
                full_name,
                contact,
                email,
                status
            )
        )

    members = updated_members

    return_db(conn)

    return render_template(
        "members.html",
        members=members,
        search=search,
        username=session["username"]
    )

@app.route(
    "/add_member",
    methods=["GET", "POST"]
)
def add_member():

    if "username" not in session:
        return redirect("/login")
    
    conn = get_db()
    cursor = conn.cursor()


    if request.method == "POST":

        # Hanapin ang pinakamataas na existing APP- number
        # (hindi FCCI-, dahil "Applicant" pa ang status nito —
        # dapat mag-bayad muna ng Registration Fee bago maging
        # Active member, gaya ng public registration flow)
        cursor.execute("""
        SELECT COALESCE(
            MAX(CAST(SPLIT_PART(member_id, '-', 3) AS INTEGER)),
            0
        )
        FROM members
        WHERE member_id LIKE 'APP-%%'
        """)
        highest_num = cursor.fetchone()[0]
        next_no     = highest_num + 1
        member_id   = f"APP-{datetime.now().year}-{next_no:06d}"
        full_name = request.form["full_name"]
        contact = request.form["contact"]
        birthday = request.form["birthday"]
        email = request.form["email"]
        address = request.form["address"]
        member_since_input = request.form.get("member_since", "").strip()
        photo = request.files["photo"]
        

        photo_filename = ""

        if photo and photo.filename:
            photo_filename = upload_photo(photo, folder="fcci_member_photos") or ""

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO members
        (
            member_id,
            full_name,
            contact,
            address,
            registration_fee,
            member_since,
            email,
            birthday,
            date_registered,
            status
        )
        VALUES
        (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            member_id,
            full_name,
            contact,
            address,
            0,
            member_since_input,
            email,
            birthday,
            datetime.now().strftime("%Y-%m-%d"),
            "Applicant"
        ))

        # I-upload ang photo sa Cloudinary at i-save sa member_photos
        if photo_filename:
            cursor.execute("""
            INSERT INTO member_photos (member_id, photo_path)
            VALUES (%s, %s)
            """, (member_id, photo_filename))

        conn.commit()
        return_db(conn)

        return redirect("/members")

    return render_template(
        "add_member.html",
        username=session["username"]
    )

@app.route(
    "/edit_member/<member_id>",
    methods=["GET", "POST"]
)
def edit_member(member_id):

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    if request.method == "POST":

        photo = request.files.get("photo")

        if photo and photo.filename:

            photo_url = upload_photo(photo, folder="fcci_member_photos")

            if photo_url:
                cursor.execute("""
                INSERT INTO member_photos (member_id, photo_path)
                VALUES (%s, %s)
                """, (
                    member_id,
                    photo_url
                ))

        cursor.execute("""
        UPDATE members
        SET
            full_name    = %s,
            contact      = %s,
            birthday     = %s,
            email        = %s,
            address      = %s,
            member_since = %s
        WHERE member_id = %s
        """, (
            request.form["full_name"],
            request.form["contact"],
            request.form["birthday"],
            request.form["email"],
            request.form["address"],
            request.form.get("member_since", "").strip(),
            member_id
        ))

        conn.commit()
        return_db(conn)

        return redirect(
            f"/view_member/{member_id}"
        )

    member = fetch_member_with_photo(cursor, member_id)

    return_db(conn)

    return render_template(
        "edit_member.html",
        member=member,
        username=session["username"]
    )


@app.route("/delete_member/<member_id>")
def delete_member(member_id):

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    # Payments
    cursor.execute("""
    DELETE FROM payments
    WHERE member_id = %s
    """, (member_id,))

    # Attendance
    cursor.execute("""
    DELETE FROM attendance
    WHERE member_id = %s
    """, (member_id,))

    # Photos
    cursor.execute("""
    DELETE FROM member_photos
    WHERE member_id = %s
    """, (member_id,))

    # Withdrawals
    cursor.execute("""
    DELETE FROM withdrawals
    WHERE member_id = %s
    """, (member_id,))

    # ID Cards
    cursor.execute("""
    DELETE FROM id_cards
    WHERE member_id = %s
    """, (member_id,))

    # Main Member Record
    cursor.execute("""
    DELETE FROM members
    WHERE member_id = %s
    """, (member_id,))

    conn.commit()
    return_db(conn)

    return redirect("/members")


@app.route("/view_member/<member_id>")
def view_member(member_id):

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    member = fetch_member_with_photo(cursor, member_id)

    return_db(conn)

    return render_template(
        "view_member.html",
        member=member,
        username=session["username"]
    )


@app.route("/check_duplicate_payment")
def check_duplicate_payment():

    if "username" not in session:
        return jsonify({"error": "Not logged in"}), 401

    member_id = request.args.get("member_id", "").strip()
    payment_type = request.args.get("payment_type", "")
    payment_month = request.args.get("payment_month", "")
    payment_year = request.args.get("payment_year", "")

    conn = get_db()
    cursor = conn.cursor()

    if payment_type == "Registration Fee":

        # I-check kung may existing na Registration Fee payment record
        # sa payments table — hindi yung status ng member.
        # Kaya kahit Active ang member, kung na-delete ang payment niya,
        # pwede pa rin siyang mag-bayad ulit.
        cursor.execute("""
        SELECT COUNT(*) FROM payments
        WHERE member_id = %s
        AND payment_type = 'Registration Fee'
        """, (member_id,))

        count = cursor.fetchone()[0]
        return_db(conn)

        if count > 0:
            return jsonify({
                "duplicate": True,
                "message": f"Ang member na ito ({member_id}) ay nakabayad na ng Registration Fee. Kung kailangan ng pagbabago, i-delete muna ang existing na payment record."
            })

        return jsonify({"duplicate": False})

    else:
        cursor.execute("""
        SELECT COUNT(*) FROM payments
        WHERE member_id = %s
        AND payment_type = 'Monthly Contribution'
        AND payment_month = %s
        AND payment_year = %s
        """, (member_id, payment_month, payment_year))

        count = cursor.fetchone()[0]
        return_db(conn)

        if count > 0:
            return jsonify({
                "duplicate": True,
                "message": f"Nakabayad na ang member na ito ng Monthly Contribution para sa {payment_month} {payment_year}. Hindi na ito dapat bayaran ulit."
            })

        return jsonify({"duplicate": False})


@app.route("/search_member_payments")
def search_member_payments():

    if "username" not in session:
        return jsonify({"error": "Not logged in"}), 401

    search_term = request.args.get("name", "").strip()

    if not search_term:
        return jsonify({"results": []})

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        p.id,
        p.receipt_no,
        p.member_id,
        p.payment_type,
        p.amount,
        p.payment_date,
        m.full_name,
        p.payment_month,
        p.payment_year
    FROM payments p
    LEFT JOIN members m ON p.member_id = m.member_id
    WHERE LOWER(m.full_name) LIKE LOWER(%s)
    ORDER BY p.id DESC
    """, (f"%{search_term}%",))

    rows = cursor.fetchall()
    return_db(conn)

    results = []
    for r in rows:
        results.append({
            "id": r[0],
            "receipt_no": r[1],
            "member_id": r[2],
            "payment_type": r[3],
            "amount": r[4],
            "payment_date": r[5],
            "full_name": r[6] or r[2],
            "payment_month": r[7],
            "payment_year": r[8]
        })

    return jsonify({"results": results})


@app.route(
    "/payments",
    methods=["GET", "POST"]
)
def payments():

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    if request.method == "POST":

        member_id = request.form["member_id"]
        payment_type = request.form["payment_type"]

        payment_month = request.form["payment_month"]
        payment_year = request.form["payment_year"]

        if payment_type == "Registration Fee":
            amount = 20000
        else:
            amount = 10000

        payment_date = datetime.now().strftime(
            "%Y-%m-%d"
        )

        cursor.execute("""
        SELECT COALESCE(MAX(id), 0)
        FROM payments
        """)

        max_id = cursor.fetchone()[0] + 1

        receipt_no = (
            f"RCPT-{datetime.now().year}-{max_id:06d}"
        )

        cursor.execute("""
        SELECT COUNT(*)
        FROM members
        WHERE member_id = %s
        """, (member_id,))

        exists = cursor.fetchone()[0]

        if exists == 0:

            return_db(conn)

            return "Member ID Not Found"

        # Kunin ang full_name at email ng member BAGO mag-INSERT,
        # para magamit sa pagpapadala ng confirmation email mamaya
        cursor.execute("""
        SELECT full_name, email FROM members WHERE member_id = %s
        """, (member_id,))
        payer_info = cursor.fetchone()

        cursor.execute("""
        INSERT INTO payments
        (
            receipt_no,
            member_id,
            payment_type,
            amount,
            payment_date,
            payment_year,
            payment_month
            
        )
        VALUES
        (%s, %s, %s, %s, %s, %s, %s)
        """, (
            receipt_no,
            member_id,
            payment_type,
            amount,
            payment_date,
            payment_year,
            payment_month
            
        ))

        # ── Kung Registration Fee ang nabayad at Applicant pa ang member,
        # automatic na i-convert sa FCCI (Option B) ─────────────
        # Payments page at ang member ay Applicant pa (APP-),
        # automatic na i-convert sa FCCI — parehong logic ng
        # approve_applicant route
        if payment_type == "Registration Fee" and member_id.startswith("APP-"):

            # SQL MAX para sa FCCI ID — mas mabilis kaysa Python loop
            cursor.execute("""
            SELECT COALESCE(
                MAX(CAST(SPLIT_PART(member_id, '-', 3) AS INTEGER)),
                0
            )
            FROM members
            WHERE member_id LIKE 'FCCI-%%'
            """)
            highest_num   = cursor.fetchone()[0]
            new_member_id = f"FCCI-2026-{highest_num + 1:06d}"

            # Gamitin ang payment month/year bilang member_since
            # — ito ang "totoo" na petsa ng pagpasok, hindi yung
            # posibleng maling nalagay sa registration form.
            # Hal. kung nag-bayad siya ng "May 2026", "May 2026"
            # din ang magiging member_since niya.
            member_since_from_payment = f"{payment_month} {payment_year}"

            # I-update ang members table
            cursor.execute("""
            UPDATE members
            SET member_id = %s,
                status = 'Active',
                registration_fee = 20000,
                member_since = %s
            WHERE member_id = %s
            """, (new_member_id, member_since_from_payment, member_id))

            # I-update ang member_photos
            cursor.execute("""
            UPDATE member_photos
            SET member_id = %s
            WHERE member_id = %s
            """, (new_member_id, member_id))

            # I-update ang payment record na kakaka-INSERT lang
            cursor.execute("""
            UPDATE payments
            SET member_id = %s
            WHERE receipt_no = %s
            """, (new_member_id, receipt_no))

        elif payment_type == "Registration Fee" and not member_id.startswith("APP-"):
            # FCCI member na pero nag-bayad ulit ng Registration Fee
            # (hal. na-delete ang dating payment) — i-update lang
            # ang member_since at registration_fee, huwag baguhin
            # ang member_id at status
            member_since_from_payment = f"{payment_month} {payment_year}"
            cursor.execute("""
            UPDATE members
            SET registration_fee = 20000,
                member_since = %s
            WHERE member_id = %s
            """, (member_since_from_payment, member_id))

        conn.commit()

        # I-send ang payment confirmation email (hindi mag-crash
        # ang app kung walang email o nag-error ang pagpapadala)
        if payer_info and payer_info[1]:
            try:
                send_payment_confirmation_email(
                    payer_info[1],
                    payer_info[0],
                    payment_type,
                    amount,
                    receipt_no,
                    payment_date
                )
            except Exception as e:
                print(f"[EMAIL] Failed to send payment confirmation: {e}")

    cursor.execute("""
    SELECT
        member_id,
        full_name,
        status
    FROM members
    ORDER BY full_name
    """)

    members = cursor.fetchall()

    cursor.execute("""
    SELECT
        p.id,
        p.receipt_no,
        p.member_id,
        p.payment_type,
        p.amount,
        p.payment_date,
        m.full_name,
        p.payment_month,
        p.payment_year
    FROM payments p
    LEFT JOIN members m ON p.member_id = m.member_id
    ORDER BY p.id DESC
    LIMIT 50
    """)

    payment_history = cursor.fetchall()

    return_db(conn)

    return render_template(
        "payments.html",
        members=members,
        payment_history=payment_history,
        username=session["username"]
    )


@app.route("/approval_certificate/<member_id>")
def approval_certificate(member_id):
    if "username" not in session:
        return redirect("/login")

    from reportlab.lib.pagesizes import A5
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, Image)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import io, tempfile, os as _os

    # ── Kunin ang member + registration fee payment ──
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT member_id, full_name, contact, member_since, status, email
    FROM members WHERE member_id = %s
    """, (member_id,))
    m = cursor.fetchone()

    if not m:
        return_db(conn)
        return "Member not found", 404

    cursor.execute("""
    SELECT receipt_no, amount, payment_date, payment_month, payment_year
    FROM payments
    WHERE member_id = %s AND payment_type = 'Registration Fee'
    ORDER BY id DESC LIMIT 1
    """, (member_id,))
    pay = cursor.fetchone()
    return_db(conn)

    m_id, full_name, contact, member_since, status, email = m

    # ── Generate QR code (naglalaman ng member ID) ──
    qr_path = None
    try:
        qr_img = qrcode.make(m_id)
        qr_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        qr_img.save(qr_tmp.name)
        qr_path = qr_tmp.name
    except Exception as e:
        print(f"[CERT] QR error: {e}")

    # ── Kunin ang member photo ──
    photo_path = download_photo_for_pdf(m_id)

    # ── Buuin ang PDF ──
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A5,
                            topMargin=12*mm, bottomMargin=10*mm,
                            leftMargin=12*mm, rightMargin=12*mm)
    styles = getSampleStyleSheet()
    content = []

    # Modern header na may logo (shared helper)
    content.append(_modern_pdf_header("OFFICIAL"))
    content.append(Spacer(1, 12))

    # Approved title
    content.append(Paragraph("Membership Approved",
        ParagraphStyle("at", fontSize=15, textColor=colors.HexColor("#0b1d2e"),
                       fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=2)))
    content.append(Paragraph("Ito ang iyong temporary membership card habang hinihintay ang opisyal na ID.",
        ParagraphStyle("ats", fontSize=8.5, textColor=colors.HexColor("#5c8270"),
                       alignment=TA_CENTER, spaceAfter=12)))

    # ── TEMPORARY ID CARD (photo + details + QR) ──
    lbl = ParagraphStyle("lbl", fontSize=7.5, textColor=colors.HexColor("#8a6000"),
                         fontName="Helvetica-Bold")
    name_st = ParagraphStyle("nm", fontSize=15, textColor=colors.HexColor("#0b1d2e"),
                             fontName="Helvetica-Bold", leading=17)
    idnum_st = ParagraphStyle("idn", fontSize=13, textColor=colors.HexColor("#00a89e"),
                              fontName="Helvetica-Bold", leading=18)
    meta_st = ParagraphStyle("mt", fontSize=8.5, textColor=colors.HexColor("#5c8270"), leading=13)

    # Photo cell
    if photo_path and _os.path.exists(photo_path):
        photo_cell = Image(photo_path, width=26*mm, height=26*mm)
    else:
        photo_cell = Paragraph("NO<br/>PHOTO", ParagraphStyle("np", fontSize=8,
            textColor=colors.HexColor("#9ab5a8"), alignment=TA_CENTER))

    # QR cell
    if qr_path and _os.path.exists(qr_path):
        qr_cell = Image(qr_path, width=22*mm, height=22*mm)
    else:
        qr_cell = Paragraph("", styles["Normal"])

    details = [
        Paragraph("⏳ TEMPORARY ID", lbl),
        Paragraph(full_name or "—", name_st),
        Paragraph(m_id, idnum_st),
        Paragraph(f"<b>Status:</b> {status}<br/><b>Member Since:</b> {member_since or '—'}<br/><b>Contact:</b> {contact or '—'}", meta_st),
    ]
    details_tbl = Table([[d] for d in details], colWidths=[62*mm])
    details_tbl.setStyle(TableStyle([
        ("TOPPADDING", (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ("LEFTPADDING", (0,0), (-1,-1), 0),
    ]))

    id_card = Table([[photo_cell, details_tbl, qr_cell]],
                    colWidths=[28*mm, 64*mm, 24*mm])
    id_card.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#f4fbf7")),
        ("BOX", (0,0), (-1,-1), 1, colors.HexColor("#c5e8dc")),
        ("LINEABOVE", (0,0), (-1,0), 4, colors.HexColor("#00d4c8")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 14),
        ("BOTTOMPADDING", (0,0), (-1,-1), 14),
        ("LEFTPADDING", (0,0), (-1,-1), 12),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
    ]))
    content.append(id_card)
    content.append(Spacer(1, 6))
    content.append(Paragraph("↑ I-scan ang QR code para sa mabilis na verification",
        ParagraphStyle("qn", fontSize=7, textColor=colors.HexColor("#9ab5a8"),
                       alignment=TA_CENTER, spaceAfter=14)))

    # ── RECEIPT SECTION ──
    content.append(Paragraph("REGISTRATION FEE RECEIPT",
        ParagraphStyle("rl", fontSize=8, textColor=colors.HexColor("#00a89e"),
                       fontName="Helvetica-Bold", spaceAfter=8)))

    if pay:
        rcpt_no, amount, pdate, pmonth, pyear = pay
        receipt_data = [
            ["Receipt No.", rcpt_no or "—"],
            ["Payment Type", "Registration Fee"],
            ["Date Paid", str(pdate) if pdate else "—"],
            ["Amount Paid", f"₩{amount:,}" if amount else "₩0"],
        ]
    else:
        receipt_data = [["Receipt", "Walang registration fee record"]]

    receipt_tbl = Table(receipt_data, colWidths=[50*mm, 66*mm])
    receipt_tbl.setStyle(TableStyle([
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("TEXTCOLOR", (0,0), (0,-1), colors.HexColor("#5c8270")),
        ("TEXTCOLOR", (1,0), (1,-1), colors.HexColor("#0c2418")),
        ("FONTNAME", (1,0), (1,-1), "Helvetica-Bold"),
        ("FONTSIZE", (1,-1), (1,-1), 14),
        ("TEXTCOLOR", (1,-1), (1,-1), colors.HexColor("#00a89e")),
        ("TOPPADDING", (0,0), (-1,-1), 7),
        ("BOTTOMPADDING", (0,0), (-1,-1), 7),
        ("LINEBELOW", (0,0), (-1,-2), 0.5, colors.HexColor("#e0f0e5")),
    ]))
    content.append(receipt_tbl)
    content.append(Spacer(1, 16))

    # Footer
    foot = ParagraphStyle("ft", fontSize=8, textColor=colors.HexColor("#9ab5a8"),
                          alignment=TA_CENTER, leading=12)
    content.append(Paragraph("<b>United in Faith, Serving with Love</b>", foot))
    content.append(Paragraph("Ipakita ang temporary ID na ito sa opisina para makuha ang opisyal na membership card.", foot))
    content.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} · Valid as proof of membership",
        ParagraphStyle("fn", fontSize=6.5, textColor=colors.HexColor("#b5c9be"), alignment=TA_CENTER)))

    doc.build(content, onFirstPage=_watermark_canvas, onLaterPages=_watermark_canvas)
    buf.seek(0)

    # Cleanup temp files
    for tmp in [qr_path, photo_path]:
        if tmp and _os.path.exists(tmp):
            try:
                _os.remove(tmp)
            except Exception:
                pass

    return send_file(buf, mimetype="application/pdf",
                     download_name=f"Approval_{m_id}.pdf")


@app.route("/print_receipt/<int:payment_id>")
def print_receipt(payment_id):
    if "username" not in session:
        return redirect("/login")

    from reportlab.lib.pagesizes import A5
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import io

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT p.receipt_no, p.member_id, p.payment_type, p.amount,
           p.payment_month, p.payment_year, p.payment_date, m.full_name
    FROM payments p
    LEFT JOIN members m ON p.member_id = m.member_id
    WHERE p.id = %s
    """, (payment_id,))
    row = cursor.fetchone()
    return_db(conn)

    if not row:
        return "Payment not found", 404

    receipt_no, member_id, ptype, amount, pmonth, pyear, pdate, full_name = row

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A5,
                            topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("t", parent=styles["Title"],
                                 fontSize=18, textColor=colors.HexColor("#00562a"),
                                 alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle("s", parent=styles["Normal"],
                               fontSize=9, textColor=colors.HexColor("#5c8270"),
                               alignment=TA_CENTER, spaceAfter=2)
    rcpt_style = ParagraphStyle("r", parent=styles["Normal"],
                                fontSize=10, textColor=colors.HexColor("#0c2418"),
                                alignment=TA_CENTER, spaceAfter=14)

    content = []

    # Logo sa taas (centered)
    _logo = _pdf_logo_path()
    if _logo:
        from reportlab.platypus import Image as _Img
        content.append(Table([[_Img(_logo, width=18*mm, height=18*mm)]],
                       colWidths=[120*mm],
                       style=TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER")])))
        content.append(Spacer(1, 6))

    content.append(Paragraph("FCCI", title_style))
    content.append(Paragraph("Filipino Community Center International", sub_style))
    content.append(Paragraph("Official Payment Receipt", sub_style))
    content.append(Spacer(1, 10))
    content.append(Paragraph(f"Receipt No: <b>{receipt_no}</b>", rcpt_style))

    data = [
        ["Member ID", member_id or "—"],
        ["Name", full_name or "—"],
        ["Payment Type", ptype or "—"],
        ["Period", f"{pmonth} {pyear}" if pmonth else "—"],
        ["Amount", f"₩{amount:,}" if amount else "₩0"],
        ["Date Paid", str(pdate) if pdate else "—"],
    ]
    table = Table(data, colWidths=[45*mm, 75*mm])
    table.setStyle(TableStyle([
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("TEXTCOLOR", (0,0), (0,-1), colors.HexColor("#5c8270")),
        ("TEXTCOLOR", (1,0), (1,-1), colors.HexColor("#0c2418")),
        ("FONTNAME", (1,0), (1,-1), "Helvetica-Bold"),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LINEBELOW", (0,0), (-1,-2), 0.5, colors.HexColor("#e0f0e5")),
    ]))
    content.append(table)
    content.append(Spacer(1, 20))

    foot_style = ParagraphStyle("f", parent=styles["Normal"],
                                fontSize=8, textColor=colors.HexColor("#9ab5a8"),
                                alignment=TA_CENTER)
    content.append(Paragraph("United in Faith, Serving with Love", foot_style))
    content.append(Paragraph("Salamat sa iyong kontribusyon!", foot_style))

    doc.build(content, onFirstPage=_watermark_canvas, onLaterPages=_watermark_canvas)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf",
                     download_name=f"Receipt_{receipt_no}.pdf")


@app.route("/delete_payment/<int:payment_id>")
def delete_payment(payment_id):

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    # I-check muna kung Registration Fee ito bago i-delete
    cursor.execute("""
    SELECT member_id, payment_type FROM payments WHERE id = %s
    """, (payment_id,))
    payment = cursor.fetchone()

    if payment:
        paid_member_id = payment[0]
        payment_type   = payment[1]

        # I-delete ang payment
        cursor.execute(
            "DELETE FROM payments WHERE id = %s",
            (payment_id,)
        )

        # Kung Registration Fee ang na-delete at FCCI member siya,
        # i-revert siya pabalik sa APP- at Applicant status
        if payment_type == "Registration Fee" and paid_member_id.startswith("FCCI-"):

            # SQL MAX para sa APP ID — mas mabilis at mas tama
            cursor.execute("""
            SELECT COALESCE(
                MAX(CAST(SPLIT_PART(member_id, '-', 3) AS INTEGER)),
                0
            )
            FROM members
            WHERE member_id LIKE 'APP-%%'
            """)
            highest_num = cursor.fetchone()[0]
            new_app_id  = f"APP-{datetime.now().year}-{highest_num + 1:06d}"

            # I-revert ang member sa APP- at Applicant status
            cursor.execute("""
            UPDATE members
            SET member_id        = %s,
                status           = 'Applicant',
                registration_fee = 0,
                member_since     = ''
            WHERE member_id = %s
            """, (new_app_id, paid_member_id))

            # I-update din ang member_photos at payments tables
            cursor.execute("""
            UPDATE member_photos
            SET member_id = %s
            WHERE member_id = %s
            """, (new_app_id, paid_member_id))

            cursor.execute("""
            UPDATE payments
            SET member_id = %s
            WHERE member_id = %s
            """, (new_app_id, paid_member_id))

            print(f"[REVERT] {paid_member_id} → {new_app_id} (Applicant)")

    conn.commit()
    return_db(conn)

    return redirect("/payments")


@app.route(
    "/member_id_card/<member_id>"
)
def member_id_card(member_id):

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    member = fetch_member_with_photo(cursor, member_id)

    return_db(conn)

    if not member:
        return "Member Not Found"

    qr = qrcode.make(member_id)

    qr_path = (
        f"static/qr/{member_id}.png"
    )

    qr.save(qr_path)

    return render_template(
        "member_id_card.html",
        member=member,
        qr_file=f"{member_id}.png"
    )


@app.route("/bulk_id_cards", methods=["POST"])
def bulk_id_cards():

    if "username" not in session:
        return redirect("/login")

    member_ids = request.form.getlist("selected_members")

    if not member_ids:
        return redirect("/members")

    conn = get_db()
    cursor = conn.cursor()

    members_list = []

    for mid in member_ids:

        member = fetch_member_with_photo(cursor, mid)

        if member:

            qr = qrcode.make(mid)
            qr_path = f"static/qr/{mid}.png"
            qr.save(qr_path)

            members_list.append({
                "member": member,
                "qr_file": f"{mid}.png"
            })

    return_db(conn)

    return render_template(
        "bulk_id_cards.html",
        members_list=members_list
    )


@app.route(
"/expenses",
methods=["GET","POST"]
)
def expenses():

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    if request.method == "POST":

        expense_type = request.form["expense_type"]
        description = request.form["description"]

        try:
            amount = int(
                request.form["amount"]
            )
        except:
            amount = 0

        receipt_file = request.files.get(
            "receipt"
        )

        receipt_filename = ""

        if (
            receipt_file
            and
            receipt_file.filename
        ):

            receipt_filename = upload_file(receipt_file, folder="fcci_receipts") or ""

        cursor.execute("""
        INSERT INTO expenses
        (
            expense_type,
            description,
            amount,
            receipt_path,
            expense_date
        )
        VALUES
        (%s, %s, %s, %s, %s)
        """, (
            expense_type,
            description,
            amount,
            receipt_filename,
            datetime.now().strftime(
               "%Y-%m-%d"
           )
       ))
    
        conn.commit()

    cursor.execute("""
    SELECT
        id,
        expense_type,
        description,
        amount,
        expense_date,
        receipt_path
    FROM expenses
    ORDER BY id DESC
    """)

    expense_history = cursor.fetchall()

    return_db(conn)

    return render_template(
        "expenses.html",
        expense_history=expense_history,
        username=session["username"]
    )

@app.route(
"/donations",
methods=["GET","POST"]
)
def donations():


    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    if request.method == "POST":

        donor_name = request.form["donor_name"]
        contact = request.form["contact"]
        amount = int(request.form["amount"])
        purpose = request.form["purpose"]

        cursor.execute("""
        INSERT INTO donations
        (
            donor_name,
            contact,
            amount,
            purpose,
            donation_date
        )
        VALUES
        (%s, %s, %s, %s, %s)
        """, (
            donor_name,
            contact,
            amount,
            purpose,
            datetime.now().strftime(
                "%Y-%m-%d"
            )
        ))

        conn.commit()

    cursor.execute("""
    SELECT
        id,
        donor_name,
        contact,
        amount,
        purpose,
        donation_date
    FROM donations
    ORDER BY id DESC
    """)

    donation_history = cursor.fetchall()

    return_db(conn)

    return render_template(
        "donations.html",
        donation_history=donation_history,
        username=session["username"]
    )

@app.route("/exports")
def exports():


    if "username" not in session:
        return redirect("/login")
    
    return render_template(
        "exports.html",
        username=session["username"]
    )


@app.route("/export_members")
def export_members():

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        member_id,
        full_name,
        contact,
        email,
        address,
        member_since,
        status
    FROM members
    ORDER BY full_name
    """)

    rows = cursor.fetchall()

    return_db(conn)

    wb = Workbook()

    ws = wb.active

    ws.title = "Members"

    ws.append([
        "Member ID",
        "Full Name",
        "Contact",
        "Email",
        "Address",
        "Member Since",
        "Status"
    ])

    for row in rows:
        ws.append(row)

    export_dir = "exports"

    os.makedirs(
        export_dir,
        exist_ok=True
    )

    filename = os.path.join(
        export_dir,
        "FCCI_Members.xlsx"
    )

    wb.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )

@app.route("/export_payments")
def export_payments():


    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        receipt_no,
        member_id,
        payment_type,
        amount,
        payment_month,
        payment_year,
        payment_date
    FROM payments
    ORDER BY id DESC
    """)

    rows = cursor.fetchall()

    return_db(conn)

    wb = Workbook()

    ws = wb.active

    ws.title = "Payments"

    ws.append([
        "Receipt No",
        "Member ID",
        "Payment Type",
        "Amount",
        "Month",
        "Year",
        "Payment Date"
    ])

    for row in rows:
        ws.append(row)

    export_dir = "exports"

    os.makedirs(
        export_dir,
        exist_ok=True
    )

    filename = os.path.join(
        export_dir,
        "FCCI_Payments.xlsx"
    )

    wb.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )

@app.route("/export_donations")
def export_donations():


    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        donor_name,
        contact,
        amount,
        purpose,
        donation_date
    FROM donations
    ORDER BY id DESC
    """)

    rows = cursor.fetchall()

    return_db(conn)

    wb = Workbook()

    ws = wb.active

    ws.title = "Donations"

    ws.append([
        "Donor Name",
        "Contact",
        "Amount",
        "Purpose",
        "Donation Date"
    ])

    for row in rows:
        ws.append(row)

    export_dir = "exports"

    os.makedirs(
        export_dir,
        exist_ok=True
    )

    filename = os.path.join(
        export_dir,
        "FCCI_Donations.xlsx"
    )

    wb.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )

@app.route(
"/donation_certificate/<int:donation_id>"
)
def donation_certificate(donation_id):


    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        donor_name,
        contact,
        amount,
        purpose,
        donation_date
    FROM donations
    WHERE id = %s
    """, (
        donation_id,
    ))

    donation = cursor.fetchone()

    return_db(conn)

    if not donation:
        return "Donation Not Found"

    import io as _io
    donor_name, d_contact, d_amount, d_purpose, d_date = donation

    buf = _io.BytesIO()
    _draw_elegant_certificate(buf, dict(
        cert_title="DONATION CERTIFICATE",
        donor_name=donor_name,
        amount=d_amount,
        purpose=d_purpose,
        date_str=str(d_date) if d_date else datetime.now().strftime("%B %d, %Y"),
        receipt_no=f"DON-2026-{donation_id:06d}",
        contact=d_contact,
        signee="Authorized Signatory",
        signee_title="Treasurer",
    ))
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"Donation_Certificate_{donation_id}.pdf")


@app.route("/export_expenses")
def export_expenses():


    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        expense_type,
        description,
        amount,
        expense_date
    FROM expenses
    ORDER BY id DESC
    """)

    rows = cursor.fetchall()

    return_db(conn)

    wb = Workbook()

    ws = wb.active

    ws.title = "Expenses"

    ws.append([
        "Expense Type",
        "Description",
        "Amount",
        "Expense Date"
    ])

    for row in rows:
        ws.append(row)

    export_dir = "exports"

    os.makedirs(
        export_dir,
        exist_ok=True
    )

    filename = os.path.join(
        export_dir,
        "FCCI_Expenses.xlsx"
    )

    wb.save(filename)
    
    return send_file(
        filename,
        as_attachment=True
    )

@app.route("/export_attendance")
def export_attendance():

    if "username" not in session:
        return redirect("/login")

    export_date = request.args.get(
        "export_date"
    )

    if not export_date:
        return "Please select a date."

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        a.member_id,
        m.full_name,
        a.attendance_date,
        a.attendance_time,
        a.time_out
    FROM attendance a
    LEFT JOIN members m
    ON a.member_id = m.member_id
    WHERE a.attendance_date = %s
    ORDER BY a.id DESC
    """, (
        export_date,
    ))

    rows = cursor.fetchall()

    return_db(conn)

    wb = Workbook()

    ws = wb.active

    ws.title = "Attendance"

    ws.append([
        "Member ID",
        "Full Name",
        "Date",
        "Time In",
        "Time Out"
    ])

    for row in rows:
        ws.append(row)

    export_dir = "exports"

    os.makedirs(
        export_dir,
        exist_ok=True
    )

    filename = os.path.join(
        export_dir,
        f"Attendance_{export_date}.xlsx"
    )

    wb.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )

@app.route("/export_monitoring")
def export_monitoring():


    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        member_id,
        full_name
    FROM members
    ORDER BY full_name
    """)

    members = cursor.fetchall()

    wb = Workbook()

    ws = wb.active

    ws.title = "Monthly Monitoring"

    ws.append([
        "Member ID",
        "Full Name",
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec"
    ])

    month_map = {
        "January":"Jan",
        "February":"Feb",
        "March":"Mar",
        "April":"Apr",
        "May":"May",
        "June":"Jun",
        "July":"Jul",
        "August":"Aug",
        "September":"Sep",
        "October":"Oct",
        "November":"Nov",
        "December":"Dec"
    }

    for member in members:

        member_id = member[0]
        full_name = member[1]

        row = {
            "Jan":"-",
            "Feb":"-",
            "Mar":"-",
            "Apr":"-",
            "May":"-",
            "Jun":"-",
            "Jul":"-",
            "Aug":"-",
            "Sep":"-",
            "Oct":"-",
            "Nov":"-",
            "Dec":"-"
        }

        cursor.execute("""
        SELECT payment_month
        FROM payments
        WHERE member_id = %s
        """, (member_id,))

        payments = cursor.fetchall()

        for p in payments:

            month = p[0]

            if month in month_map:

                row[
                    month_map[month]
                ] = "✔"

        ws.append([
            member_id,
            full_name,
            row["Jan"],
            row["Feb"],
            row["Mar"],
            row["Apr"],
            row["May"],
            row["Jun"],
            row["Jul"],
            row["Aug"],
            row["Sep"],
            row["Oct"],
            row["Nov"],
            row["Dec"]
        ])

    return_db(conn)

    export_dir = "exports"

    os.makedirs(
        export_dir,
        exist_ok=True
    )

    filename = os.path.join(
        export_dir,
        "FCCI_Monthly_Monitoring.xlsx"
    )

    wb.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )

@app.route(
"/attendance",
methods=["GET","POST"]
)
def attendance():


    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    message = ""

    if request.method == "POST":

        member_id = request.form[
            "member_id"
        ].strip()

        cursor.execute("""
        SELECT full_name
        FROM members
        WHERE member_id = %s
        """, (
            member_id,
        ))

        member = cursor.fetchone()

        if member:

            today = datetime.now().strftime(
                "%Y-%m-%d"
            )

            current_time = datetime.now().strftime(
                "%H:%M:%S"
            )

            cursor.execute("""
            SELECT
                id,
                time_out
            FROM attendance
            WHERE member_id = %s
            AND attendance_date = %s
            """, (
                member_id,
                today
            ))

            record = cursor.fetchone()

            if record:

                if not record[1]:

                    cursor.execute("""
                    UPDATE attendance
                    SET time_out = %s
                    WHERE id = %s
                    """, (
                        current_time,
                        record[0]
                    ))

                    conn.commit()

                    message = (
                        f"{member[0]}"
                        f" TIME OUT "
                        f"{current_time}"
                    )

                else:

                    message = (
                        "Already Timed Out Today"
                    )

            else:

                cursor.execute("""
                INSERT INTO attendance
                (
                    member_id,
                    attendance_date,
                    attendance_time,
                    time_out
                )
                VALUES (%s, %s, %s, %s)
                """, (
                    member_id,
                    today,
                    current_time,
                    ""
                ))


                conn.commit()

                message = (
                    f"{member[0]}"
                    f" TIME IN "
                    f"{current_time}"
                )

        else:

            message = "Member Not Found"

    today = datetime.now().strftime(
        "%Y-%m-%d"
    )

    cursor.execute("""
    SELECT
        a.member_id,
        m.full_name,
        a.attendance_date,
        a.attendance_time,
        a.time_out
    FROM attendance a
    LEFT JOIN members m
    ON a.member_id = m.member_id
    WHERE attendance_date = %s
    ORDER BY a.id DESC
    """, (
        today,
    ))

    attendance_history = (
        cursor.fetchall()
    )

    return_db(conn)

    return render_template(
        "attendance.html",
        attendance_history=attendance_history,
        message=message,
        username=session["username"]
    )

@app.route("/qr_attendance")
def qr_attendance():


    if "username" not in session:
        return redirect("/login")

    cap = cv2.VideoCapture(0)

    detector = cv2.QRCodeDetector()

    while True:

        success, frame = cap.read()

        if not success:
            break

        data, bbox, _ = detector.detectAndDecode(
            frame
        )

        if data:

            member_id = data.strip()

            conn = get_db()
            cursor = conn.cursor()

            cursor.execute("""
            SELECT full_name
            FROM members
            WHERE member_id = %s
            """, (
                member_id,
            ))

            member = cursor.fetchone()

            if member:

                today = datetime.now().strftime(
                    "%Y-%m-%d"
                )

                current_time = datetime.now().strftime(
                    "%H:%M:%S"
                )

                cursor.execute("""
                SELECT
                    id,
                    time_out
                FROM attendance
                WHERE member_id = %s
                AND attendance_date = %s
                """, (
                    member_id,
                    today
                ))

                record = cursor.fetchone()

                if record:

                    if not record[1]:

                        cursor.execute("""
                        UPDATE attendance
                        SET time_out = %s
                        WHERE id = %s
                        """, (
                            current_time,
                            record[0]
                        ))

                    conn.commit()

                else:

                    cursor.execute("""
                    INSERT INTO attendance
                    (
                        member_id,
                        attendance_date,
                        attendance_time,
                        time_out
                    )
                    VALUES (%s, %s, %s, %s)
                    """, (
                        member_id,
                        today,
                        current_time,
                        ""
                    ))

                    conn.commit()

            return_db(conn)

            break

        cv2.imshow(
            "FCCI QR Attendance",
            frame
        )

        if cv2.waitKey(1) == 27:
            break

        import time

        current_scan_time = time.time()

        if data:

            member_id = data.strip()

            if (
                member_id == last_scan
                and
                current_scan_time - last_scan_time < 5
            ):
                pass

            else:

                last_scan = member_id
                last_scan_time = current_scan_time

                # attendance logic dito

    cap.release()

    cv2.destroyAllWindows()

    return redirect("/attendance")

@app.route(
    "/qr_attendance_scan",
    methods=["POST"]
)
def qr_attendance_scan():

    if "username" not in session:
        return {"success":False}

    member_id = request.form["member_id"]

    

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT full_name
    FROM members
    WHERE member_id = %s
    """, (member_id,))

    member = cursor.fetchone()

    if not member:

        return_db(conn)

        return {
            "success":False,
            "message":"Member Not Found"
        }

    today = datetime.now().strftime(
        "%Y-%m-%d"
    )

    current_time = datetime.now().strftime(
        "%H:%M:%S"
    )

    cursor.execute("""
    SELECT id,time_out
    FROM attendance
    WHERE member_id = %s
    AND attendance_date = %s
    """, (
        member_id,
        today
    ))

    record = cursor.fetchone()

    if record:

        if not record[1]:

            cursor.execute("""
            UPDATE attendance
            SET time_out = %s
            WHERE id = %s
            """, (
                current_time,
                record[0]
            ))

            action = "TIME OUT"
            print("TIME OUT SAVED:", member_id)

        else:

            return_db(conn)

            return {
                "success":True,
                "message":"Already Timed Out"
            }

    else:

        cursor.execute("""
        INSERT INTO attendance
        (
            member_id,
            attendance_date,
            attendance_time,
            time_out
        )
        VALUES (%s, %s, %s, %s)
        """, (
            member_id,
            today,
            current_time,
            ""
        ))

        action = "TIME IN"
        print("TIME IN SAVED:", member_id)

    conn.commit()

    cursor.execute("""
    SELECT COUNT(*)
    FROM attendance
    WHERE member_id = %s
    AND attendance_date = %s
    """, (
        member_id,
        today
    ))

    count = cursor.fetchone()[0]

    return_db(conn)

    return {
        "success":True,
        "message":
        f"{member[0]} {action} {current_time}",
        "count": count
    }

@app.route("/member_profile/<member_id>")
def member_profile(member_id):

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    member = fetch_member_with_photo(cursor, member_id)

    if not member:

        return_db(conn)
        return "Member Not Found"

    cursor.execute("""
    SELECT
        payment_date,
        payment_type,
        amount,
        payment_month,
        payment_year
    FROM payments
    WHERE member_id = %s
    ORDER BY id DESC
    """, (member_id,))

    payments = cursor.fetchall()

    cursor.execute("""
    SELECT
        COALESCE(SUM(amount),0)
    FROM payments
    WHERE member_id = %s
    """, (member_id,))

    total_payment = cursor.fetchone()[0]
    
    cursor.execute("""
    SELECT attendance_date
    FROM attendance
    WHERE member_id = %s
    ORDER BY id DESC
    LIMIT 1
    """, (member_id,))

    attendance = cursor.fetchone()

    return_db(conn)

    return render_template(
        "member_profile.html",
        member=member,
        payments=payments,
        attendance=attendance,
        total_payment=total_payment
    )

@app.route(
    "/withdrawals",
    methods=["GET", "POST"]
)
def withdrawals():

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    member = None
    total_contributions = 0
    refund_amount = 0
    community_share = 0

    if request.method == "POST":

        action = request.form.get("action")

        member_id = request.form["member_id"]

        member = fetch_member_with_photo(cursor, member_id)

        if member:

            cursor.execute("""
            SELECT
                COALESCE(SUM(amount),0)
            FROM payments
            WHERE member_id = %s
            AND payment_type = 'Monthly Contribution'
            """, (member_id,))

            total_contributions = (
                cursor.fetchone()[0]
            )

            refund_amount = int(
                total_contributions * 0.75
            )

            community_share = (
                total_contributions
                - refund_amount
            )

            

            if action == "finalize":

                cursor.execute("""
                INSERT INTO withdrawals
                (
                    member_id,
                    full_name,
                    total_contributions,
                    refund_amount,
                    community_share,
                    withdrawal_date
                )
                VALUES
                (%s, %s, %s, %s, %s, %s)
                """, (
                    member[1],
                    member[2],
                    total_contributions,
                    refund_amount,
                    community_share,
                    datetime.now().strftime(
                        "%Y-%m-%d"
                    )
                ))

                cursor.execute("""
                DELETE FROM payments
                WHERE member_id = %s
                """, (member_id,))

                cursor.execute("""
                DELETE FROM attendance
                WHERE member_id = %s
                """, (member_id,))

                cursor.execute("""
                DELETE FROM members
                WHERE member_id = %s
                """, (member_id,))
                

                conn.commit()

                return_db(conn)

                return redirect(
                    "/withdrawals"
                )

    cursor.execute("""
    SELECT *
    FROM withdrawals
    ORDER BY id DESC
    """)

    withdrawal_history = (
        cursor.fetchall()
    )

    return_db(conn)

    return render_template(
        "withdrawals.html",
        member=member,
        total_contributions=
        total_contributions,
        refund_amount=
        refund_amount,
        community_share=
        community_share,
        withdrawal_history=
        withdrawal_history,
        username=session["username"]
    )

@app.route(
"/withdrawal_certificate/<member_id>"
)
def withdrawal_certificate(member_id):


    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
    id, member_id, full_name, contact, address, registration_fee, member_since, email, birthday, date_registered, status, proof_of_payment
    FROM members
    WHERE member_id = %s
    """, (member_id,))

    member = cursor.fetchone()

    if not member:
        return_db(conn)
        return "Member Not Found"

    cursor.execute("""
    SELECT
        COALESCE(SUM(amount),0)
    FROM payments
    WHERE member_id = %s
    AND payment_type =
    'Monthly Contribution'
    """, (member_id,))

    total_contributions = (
        cursor.fetchone()[0]
    )

    refund_amount = int(
        total_contributions * 0.75
    )

    community_share = (
        total_contributions
        - refund_amount
    )

    return_db(conn)

    import io as _io
    buf = _io.BytesIO()
    _draw_withdrawal_certificate(buf, dict(
        member_name=member[2],
        member_id=member[1],
        withdrawal_date=datetime.now().strftime("%B %d, %Y"),
        total_contributions=total_contributions,
        refund_amount=refund_amount,
        community_share=community_share,
        signee="Maria Santos",
        signee_title="President",
        generated=datetime.now().strftime("%Y-%m-%d %H:%M"),
    ))
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"Withdrawal_{member_id}.pdf")

@app.route(
"/member_profile_pdf/<member_id>"
)
def member_profile_pdf(member_id):


    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
    id, member_id, full_name, contact, address, registration_fee, member_since, email, birthday, date_registered, status, proof_of_payment
    FROM members
    WHERE member_id = %s
    """, (member_id,))

    member = cursor.fetchone()

    if not member:

        return_db(conn)

        return "Member Not Found"

    cursor.execute("""
    SELECT
        COALESCE(SUM(amount),0)
    FROM payments
    WHERE member_id = %s
    """, (member_id,))

    total_payment = cursor.fetchone()[0]

    cursor.execute("""
    SELECT attendance_date
    FROM attendance
    WHERE member_id = %s
    ORDER BY id DESC
    LIMIT 1
    """, (member_id,))
    
    attendance = cursor.fetchone()

    cursor.execute("""
    SELECT
        payment_date,
        payment_type,
        amount
    FROM payments
    WHERE member_id = %s
    ORDER BY id DESC
    """, (member_id,))

    payments = cursor.fetchall()

    return_db(conn)

    import io as _io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, Image)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    kfont = _ensure_korean_font()
    buf = _io.BytesIO()
    pdf = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=14*mm, bottomMargin=14*mm,
                            leftMargin=16*mm, rightMargin=16*mm)
    content = []

    # Modern header na may logo
    content.append(_modern_pdf_header("MEMBER PROFILE"))
    content.append(Spacer(1, 16))

    # Member photo (centered)
    photo_path = download_photo_for_pdf(member[1])
    if photo_path:
        content.append(Table([[Image(photo_path, width=32*mm, height=32*mm)]],
                             colWidths=[178*mm],
                             style=TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER")])))
        content.append(Spacer(1, 10))

    content.append(Paragraph("Member Profile",
        ParagraphStyle("t", fontSize=17, textColor=colors.HexColor("#0b1d2e"),
                       fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=2)))
    content.append(Paragraph("Official Member Record",
        ParagraphStyle("s", fontSize=9, textColor=colors.HexColor("#5c8270"),
                       alignment=TA_CENTER, spaceAfter=16)))

    last_attendance = attendance[0] if attendance else "No Record"

    # Field rows — gumagamit ng Korean font para sa address (baka may Hangul)
    kr_style = ParagraphStyle("kr", fontSize=10, fontName=kfont,
                              textColor=colors.HexColor("#0c2418"))
    def field(lbl, val, mono=False, korean=False, color="#0c2418"):
        if korean:
            val_cell = Paragraph(str(val), kr_style)
        else:
            fn = "Courier-Bold" if mono else "Helvetica-Bold"
            val_cell = Paragraph(f'<font name="{fn}" color="{color}">{val}</font>',
                                 ParagraphStyle("v", fontSize=10.5))
        return [Paragraph(f'<font color="#5c8270">{lbl}</font>',
                          ParagraphStyle("l", fontSize=10.5)), val_cell]

    info_data = [
        field("Member ID", member[1], mono=True, color="#00a89e"),
        field("Full Name", member[2] or "-"),
        field("Contact", member[3] or "-", mono=True),
        field("Address", member[4] or "-", korean=True),
        field("Email", member[7] or "-"),
        field("Birthday", member[6] or "-"),
        field("Member Since", member[6] if False else (member[6] or "-")),
        field("Total Payments", f"\u20a9{total_payment:,}", color="#00a89e"),
        field("Last Attendance", str(last_attendance)),
        field("Status", member[10] or "-", color="#00a89e"),
    ]
    info_tbl = Table(info_data, colWidths=[50*mm, 128*mm])
    info_tbl.setStyle(TableStyle([
        ("TOPPADDING",(0,0),(-1,-1),8), ("BOTTOMPADDING",(0,0),(-1,-1),8),
        ("LINEBELOW",(0,0),(-1,-1),0.5,colors.HexColor("#eef3f2")),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    content.append(info_tbl)
    content.append(Spacer(1, 18))

    # PAYMENT HISTORY table
    content.append(Paragraph("PAYMENT HISTORY",
        ParagraphStyle("ph", fontSize=9, textColor=colors.HexColor("#00a89e"),
                       fontName="Helvetica-Bold", spaceAfter=8)))

    if payments:
        ph_data = [["Date", "Type", "Amount"]]
        for p in payments:
            ph_data.append([str(p[0]), str(p[1]), f"\u20a9{p[2]:,}"])
        ph_tbl = Table(ph_data, colWidths=[50*mm, 78*mm, 50*mm])
        ph_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0b2236")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),9.5),
            ("TEXTCOLOR",(2,1),(2,-1),colors.HexColor("#00a89e")),
            ("FONTNAME",(2,1),(2,-1),"Courier-Bold"),
            ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
            ("LEFTPADDING",(0,0),(-1,-1),12),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f6fbf9")]),
            ("LINEBELOW",(0,0),(-1,-1),0.5,colors.HexColor("#e0f0e5")),
        ]))
        content.append(ph_tbl)
    else:
        content.append(Paragraph("Walang payment records pa.",
            ParagraphStyle("np", fontSize=10, textColor=colors.HexColor("#9ab5a8"))))

    content.append(Spacer(1, 20))
    content.append(Paragraph("United in Faith, Serving with Love",
        ParagraphStyle("f", fontSize=8, textColor=colors.HexColor("#9ab5a8"),
                       alignment=TA_CENTER)))
    content.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        ParagraphStyle("g", fontSize=7, textColor=colors.HexColor("#b5c9be"),
                       alignment=TA_CENTER)))

    pdf.build(content, onFirstPage=_watermark_canvas, onLaterPages=_watermark_canvas)
    buf.seek(0)
    if photo_path and os.path.exists(photo_path):
        try: os.remove(photo_path)
        except Exception: pass
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"{member_id}_Profile.pdf")

@app.route("/dashboard_outstanding")
def dashboard_outstanding():

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    data = []

    cursor.execute("""
    SELECT
        member_id,
        full_name,
        member_since
    FROM members
    WHERE status='Active'
    ORDER BY full_name
    """)

    members = cursor.fetchall()

    current_date = datetime.now()

    for member in members:

        member_id = member[0]
        full_name = member[1]
        member_since = member[2]

        if not member_since:
            continue

        try:

            month_name, year = member_since.split()

            year = int(year)

            month_map = {
                "January":1,
                "February":2,
                "March":3,
                "April":4,
                "May":5,
                "June":6,
                "July":7,
                "August":8,
                "September":9,
                "October":10,
                "November":11,
                "December":12
            }

            month = month_map[month_name]

        except:
            continue

        missing_months = []

        while (
            year < current_date.year
            or
            (
                year == current_date.year
                and month <= current_date.month
            )
        ):

            current_month = datetime(
                year,
                month,
                1
            ).strftime("%B")

            cursor.execute("""
            SELECT COUNT(*)
            FROM payments
            WHERE member_id=%s
            AND payment_type='Monthly Contribution'
            AND payment_month=%s
            AND payment_year=%s
            """, (
                member_id,
                current_month,
                str(year)
            ))

            paid = cursor.fetchone()[0]

            if paid == 0:

                missing_months.append(
                    f"{current_month} {year}"
                )

            month += 1

            if month > 12:
                month = 1
                year += 1

        if missing_months:

            data.append({
                "member_id": member_id,
                "full_name": full_name,
                "months": missing_months,
                "amount": len(missing_months)*10000
            })

    return_db(conn)

    return render_template(
        "dashboard_outstanding.html",
        data=data,
        username=session["username"]
    )


@app.route("/dashboard_active")
def dashboard_active():


    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    active_members = []

    cursor.execute("""
    SELECT
        member_id,
        full_name,
        contact,
        member_since
    FROM members
    ORDER BY full_name
    """)

    members = cursor.fetchall()

    current_date = datetime.now()

    month_map = {
        "January":1,
        "February":2,
        "March":3,
        "April":4,
        "May":5,
        "June":6,
        "July":7,
        "August":8,
        "September":9,
        "October":10,
        "November":11,
        "December":12
    }

    for member in members:

        member_id = member[0]
        full_name = member[1]
        contact = member[2]
        member_since = member[3]

        if not member_since:
            continue

        try:

            month_name, year = member_since.split()

            year = int(year)

            month = month_map[month_name]

        except:
            continue

        missing_count = 0

        temp_year = year
        temp_month = month

        while (
            temp_year < current_date.year
            or
            (
                temp_year == current_date.year
                and temp_month <= current_date.month
            )
        ):

            current_month = datetime(
                temp_year,
                temp_month,
                1
            ).strftime("%B")

            cursor.execute("""
            SELECT COUNT(*)
            FROM payments
            WHERE member_id = %s
            AND payment_type='Monthly Contribution'
            AND payment_month = %s
            AND payment_year = %s
            """, (
                member_id,
                current_month,
                str(temp_year)
            ))

            paid = cursor.fetchone()[0]

            if paid == 0:
                missing_count += 1

            temp_month += 1

            if temp_month > 12:
                temp_month = 1
                temp_year += 1

        if missing_count < 5:

            active_members.append(
                (
                    member_id,
                    full_name,
                    contact,
                    missing_count
                )
            )

    return_db(conn)

    return render_template(
        "dashboard_active.html",
        members=active_members,
        username=session["username"]
    )




@app.route("/dashboard_inactive")
def dashboard_inactive():

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    inactive_members = []

    cursor.execute("""
    SELECT
        member_id,
        full_name,
        contact,
        member_since
    FROM members
    ORDER BY full_name
    """)

    members = cursor.fetchall()

    current_date = datetime.now()

    month_map = {
        "January":1,
        "February":2,
        "March":3,
        "April":4,
        "May":5,
        "June":6,
        "July":7,
        "August":8,
        "September":9,
        "October":10,
        "November":11,
        "December":12
    }

    for member in members:

        member_id = member[0]
        full_name = member[1]
        contact = member[2]
        member_since = member[3]

        if not member_since:
            continue

        try:

            month_name, year = member_since.split()

            year = int(year)

            month = month_map[month_name]

        except:
            continue

        missing_count = 0

        while (
            year < current_date.year
            or
            (
                year == current_date.year
                and month <= current_date.month
            )
        ):

            current_month = datetime(
                year,
                month,
                1
            ).strftime("%B")

            cursor.execute("""
            SELECT COUNT(*)
            FROM payments
            WHERE member_id = %s
            AND payment_type='Monthly Contribution'
            AND payment_month = %s
            AND payment_year = %s
            """, (
                member_id,
                current_month,
                str(year)
            ))

            paid = cursor.fetchone()[0]

            if paid == 0:
                missing_count += 1

            month += 1

            if month > 12:
                month = 1
                year += 1

        if missing_count >= 5:

            inactive_members.append(
                (
                    member_id,
                    full_name,
                    contact,
                    missing_count
                )
            )

    return_db(conn)

    return render_template(
        "dashboard_inactive.html",
        members=inactive_members,
        username=session["username"]
    )

@app.route("/dashboard_applicants")
def dashboard_applicants():


    if "username" not in session:
        return redirect("/login")
    
    conn = get_db()
    cursor = conn.cursor()

    # AUTO DELETE APPLICANTS AFTER 3 DAYS

    cursor.execute("""
    SELECT
        member_id,
        date_registered
    FROM members
    WHERE status='Applicant'
    """)

    applicant_list = cursor.fetchall()

    today = datetime.now()

    for applicant in applicant_list:

        try:

            registered_date = datetime.strptime(
                applicant[1],
                "%Y-%m-%d"
            )

            days_pending = (
                today - registered_date
            ).days
    
            if days_pending >= 3:

                cursor.execute("""
                SELECT COUNT(*)
                FROM payments
                WHERE member_id = %s
                AND payment_type='Registration Fee'
                """, (
                    applicant[0],
                ))

                has_payment = cursor.fetchone()[0]

                if has_payment == 0:

                    cursor.execute("""
                    DELETE FROM members
                    WHERE member_id = %s
                    """, (
                        applicant[0],
                    ))

        except:
            pass

    conn.commit()
    
    cursor.execute("""
    SELECT
        member_id,
        full_name,
        contact,
        date_registered
    FROM members
    WHERE status='Applicant'
    ORDER BY id DESC
    """)

    applicants = cursor.fetchall()

    return_db(conn)

    return render_template(
        "dashboard_applicants.html",
        applicants=applicants,
        username=session["username"]
    )

@app.route("/export_pdf_report")
def export_pdf_report():


    if "username" not in session:
        return redirect("/login")

    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer
    )

    from reportlab.lib.styles import (
        getSampleStyleSheet
    )

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT COUNT(*) FROM members"
    )
    total_members = cursor.fetchone()[0]

    cursor.execute("""
    SELECT COUNT(*)
    FROM members
    WHERE status='Active'
    """)
    active_members = cursor.fetchone()[0]

    cursor.execute("""
    SELECT COUNT(*)
    FROM members
    WHERE status='Applicant'
    """)
    applicants = cursor.fetchone()[0]

    cursor.execute("""
    SELECT COALESCE(SUM(amount),0)
    FROM payments
    """)
    collections = cursor.fetchone()[0]

    cursor.execute("""
    SELECT COALESCE(SUM(amount),0)
    FROM donations
    """)
    donations = cursor.fetchone()[0]

    cursor.execute("""
    SELECT COALESCE(SUM(amount),0)
    FROM expenses
    """)
    expenses = cursor.fetchone()[0]

    balance = (
        collections +
        donations -
        expenses
    )

    return_db(conn)

    os.makedirs(
        "exports",
        exist_ok=True
    )

    filename = (
        f"exports/FCCI_Report_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    )

    doc = SimpleDocTemplate(
        filename
    )

    styles = getSampleStyleSheet()

    content = []

    content.append(
        Paragraph(
            "FCCI FINANCIAL REPORT",
            styles["Title"]
        )
    )

    content.append(
        Spacer(1,20)
    )

    content.append(
        Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles["Normal"]
        )
    )

    content.append(
        Spacer(1,20)
    )

    content.append(
        Paragraph(
            f"Total Members: {total_members}",
            styles["Normal"]
        )
    )

    content.append(
        Paragraph(
            f"Active Members: {active_members}",
            styles["Normal"]
        )
    )

    content.append(
        Paragraph(
            f"Applicants: {applicants}",
            styles["Normal"]
        )
    )

    content.append(
        Paragraph(
            f"Collections: ₩{collections:,}",
            styles["Normal"]
        )
    )
    
    content.append(
        Paragraph(
            f"Donations: ₩{donations:,}",
            styles["Normal"]
        )
    )

    content.append(
        Paragraph(
            f"Expenses: ₩{expenses:,}",
            styles["Normal"]
        )
    )

    content.append(
        Paragraph(
            f"Current Balance: ₩{balance:,}",
            styles["Normal"]
        )
    )

    doc.build(content)

    return send_file(
        filename,
        as_attachment=True
    )



@app.route(
"/approve_applicant/<member_id>"
)
def approve_applicant(member_id):

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    # ── STEP 1: Lock ang member row para walang race condition ──
    # Kapag dalawang admin nag-approve sabay, ang FOR UPDATE
    # ay magse-serialize — isa lang ang makakapag-proceed
    cursor.execute("""
    SELECT full_name, email, member_since, status
    FROM members
    WHERE member_id = %s
    FOR UPDATE
    """, (member_id,))
    applicant_info = cursor.fetchone()

    if not applicant_info:
        return_db(conn)
        return redirect("/registration_approval")

    full_name           = applicant_info[0]
    email               = applicant_info[1]
    stored_member_since = applicant_info[2]
    current_status      = applicant_info[3]

    # Kung Active na (na-approve na ng ibang admin), tumigil na
    if current_status != "Applicant":
        return_db(conn)
        return redirect("/registration_approval")

    # ── STEP 2: I-check kung may existing na Registration Fee ──
    cursor.execute("""
    SELECT 1 FROM payments
    WHERE member_id = %s
    AND payment_type = 'Registration Fee'
    LIMIT 1
    """, (member_id,))

    if cursor.fetchone():
        print(f"[APPROVE] Duplicate registration fee blocked for {member_id}")
        return_db(conn)
        return redirect("/registration_approval")

    # ── STEP 3: Generate bagong FCCI- ID gamit ang SQL MAX ────
    # Mas mabilis at mas tama kaysa Python loop
    cursor.execute("""
    SELECT COALESCE(
        MAX(CAST(SPLIT_PART(member_id, '-', 3) AS INTEGER)),
        0
    )
    FROM members
    WHERE member_id LIKE 'FCCI-%%'
    """)
    highest_num   = cursor.fetchone()[0]
    new_member_id = f"FCCI-2026-{highest_num + 1:06d}"

    now          = datetime.now()
    member_since = stored_member_since if stored_member_since else now.strftime("%B %Y")
    payment_date = now.strftime("%Y-%m-%d")

    # ── STEP 4: I-update ang member record ────────────────────
    cursor.execute("""
    UPDATE members
    SET member_id        = %s,
        status           = 'Active',
        registration_fee = 20000,
        member_since     = %s
    WHERE member_id = %s
    """, (new_member_id, member_since, member_id))

    # ── STEP 5: I-update ang member_photos table ──────────────
    cursor.execute("""
    UPDATE member_photos
    SET member_id = %s
    WHERE member_id = %s
    """, (new_member_id, member_id))

    # ── STEP 6: Gumawa ng payment record (Registration Fee) ───
    try:
        since_parts = member_since.split()
        pay_month   = since_parts[0]
        pay_year    = since_parts[1]
    except:
        pay_month = now.strftime("%B")
        pay_year  = str(now.year)

    cursor.execute("""
    SELECT COALESCE(MAX(id), 0) FROM payments
    """)
    pay_count  = cursor.fetchone()[0] + 1
    receipt_no = f"RCPT-{datetime.now().year}-{pay_count:06d}"

    cursor.execute("""
    INSERT INTO payments
    (receipt_no, member_id, payment_type, amount, payment_date,
     payment_year, payment_month)
    VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (
        receipt_no, new_member_id, "Registration Fee",
        20000, payment_date, pay_year, pay_month
    ))

    try:
        conn.commit()
    except Exception as e:
        # Kung mag-disconnect ang Supabase sa gitna,
        # i-rollback lahat para walang partial data
        conn.rollback()
        return_db(conn)
        print(f"[APPROVE ERROR] Transaction failed for {member_id}: {e}")
        return redirect("/registration_approval")

    return_db(conn)

    # ── STEP 7: Magpadala ng welcome email (background) ───────
    if email:
        try:
            send_welcome_email(email, full_name, new_member_id)
            print(f"[EMAIL] Welcome email sent to {email} for {new_member_id}")
        except Exception as e:
            print(f"[EMAIL] Failed to send welcome email to {email}: {e}")

    # I-redirect sa success page na may auto-download ng
    # approval certificate (temp ID + receipt + QR)
    return redirect(f"/approval_success/{new_member_id}")


@app.route("/approval_success/<member_id>")
def approval_success(member_id):
    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT full_name FROM members WHERE member_id = %s
    """, (member_id,))
    row = cursor.fetchone()
    return_db(conn)

    full_name = row[0] if row else ""
    return render_template("approval_success.html",
                           member_id=member_id, full_name=full_name)

@app.route(
"/reject_applicant/<member_id>"
)
def reject_applicant(member_id):


    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    DELETE FROM members
    WHERE member_id=%s
    """, (
        member_id,
    ))

    conn.commit()
    return_db(conn)

    return redirect(
        "/registration_approval"
    )

@app.route("/registration_confirmation/<member_id>")
def registration_confirmation(member_id):

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT member_id, full_name, proof_of_payment
    FROM members
    WHERE member_id = %s
    """, (member_id,))

    row = cursor.fetchone()
    return_db(conn)

    if not row:
        return "Applicant Not Found"

    member = {
        "member_id": row[0],
        "full_name": row[1],
        "proof_uploaded": bool(row[2])
    }

    return render_template(
        "registration_confirmation.html",
        member=member
    )


@app.route("/upload_proof_of_payment", methods=["POST"])
def upload_proof_of_payment():

    member_id  = request.form["member_id"]
    proof_file = request.files.get("proof_of_payment")

    if proof_file and proof_file.filename:

        allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf'}
        import os as _os
        ext = _os.path.splitext(proof_file.filename)[1].lower()

        if ext not in allowed_extensions:
            return redirect(f"/registration_confirmation/{member_id}")

        # ── FILE SIZE LIMIT: max 10MB ────────────────────────
        # I-read ang file content at i-check ang size
        # bago mag-upload sa Cloudinary
        proof_file.seek(0, 2)  # Pumunta sa dulo ng file
        file_size = proof_file.tell()
        proof_file.seek(0)     # Bumalik sa simula

        max_size = 10 * 1024 * 1024  # 10MB
        if file_size > max_size:
            return redirect(f"/registration_confirmation/{member_id}")

        conn = get_db()
        cursor = conn.cursor()

        # I-check muna kung may existing na proof — kung meron,
        # hindi na nag-uupload ng bago para maiwasan ang duplicates
        # sa Cloudinary at sa database
        cursor.execute("""
        SELECT proof_of_payment FROM members WHERE member_id = %s
        """, (member_id,))
        row = cursor.fetchone()
        existing_proof = row[0] if row else None

        if existing_proof:
            # May proof na — hindi na mag-uupload ng bago
            return_db(conn)
            return redirect(f"/registration_confirmation/{member_id}")

        # Wala pang proof — i-upload ang bago
        proof_filename = upload_photo(proof_file, folder="fcci_proof_of_payment") or ""

        if proof_filename:
            cursor.execute("""
            UPDATE members
            SET proof_of_payment = %s
            WHERE member_id = %s
            """, (proof_filename, member_id))

            conn.commit()

        return_db(conn)

    return redirect(f"/registration_confirmation/{member_id}")


@app.route("/applicant_slip/<member_id>")
def applicant_slip(member_id):

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
    id, member_id, full_name, contact, address, registration_fee, member_since, email, birthday, date_registered, status, proof_of_payment
    FROM members
    WHERE member_id = %s
    """, (
        member_id,
    ))

    member = cursor.fetchone()

    return_db(conn)

    if not member:
        return "Applicant Not Found"

    import os
    import qrcode

    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Image,
        Table
    )

    from reportlab.lib.styles import (
        getSampleStyleSheet
    )

    os.makedirs(
        "exports",
        exist_ok=True
    )

    os.makedirs(
        "static/qr",
        exist_ok=True
    )

    qr_file = f"static/qr/{member_id}.png"

    qr = qrcode.make(member_id)
    qr.save(qr_file)

    filename = (
        f"exports/Applicant_{member_id}.pdf"
    )

    pdf = SimpleDocTemplate(filename)

    styles = getSampleStyleSheet()

    content = []

    logo_path = "static/fcci_logo.jpeg"

    if os.path.exists(logo_path):

        content.append(
            Image(
                logo_path,
                width=80,
                height=80
            )
        )

    content.append(
        Paragraph(
            "FILIPINO COMMUNITY CENTER INTERNATIONAL",
            styles["Title"]
        )
    )

    content.append(
        Paragraph(
            "APPLICANT REGISTRATION SLIP",
            styles["Heading1"]
        )
    )

    content.append(
        Spacer(1,20)
    )

    # Kunin ang applicant photo mula sa member_photos table (Cloudinary)
    slip_photo_path = download_photo_for_pdf(member_id)

    if slip_photo_path:
        content.append(Image(slip_photo_path, width=120, height=120))
        content.append(Spacer(1, 10))

    content.append(
        Image(
            qr_file,
            width=120,
            height=120
        )
    )

    content.append(
        Spacer(1,20)
    )

    data = [

        ["Applicant ID", member[1]],
        ["Full Name", member[2]],
        ["Contact", member[3]],
        ["Email", member[7]],
        ["Birthday", member[8]],
        ["Date Registered", member[9]],
        ["Status", member[10]]

    ]

    table = Table(
        data,
        colWidths=[180,300]
    )

    content.append(table)

    content.append(
        Spacer(1,30)
    )

    content.append(
        Paragraph(
            "Please pay the Registration Fee of ₩20,000 to become an Official FCCI Member.",
            styles["Normal"]
        )
    )

    content.append(
        Paragraph(
            "Present this Applicant Slip during payment.",
            styles["Normal"]
        )
    )

    pdf.build(content)

    return send_file(
        filename,
        as_attachment=True
    )

@app.route("/registration_approval")
def registration_approval():

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    applicants = fetch_all_members_with_photo(
        cursor,
        where_clause="status='Applicant'"
    )

    return_db(conn)

    return render_template(
        "registration_approval.html",
        applicants=applicants
    )

@app.route("/get_member_info/<member_id>")
def get_member_info(member_id):
 
    if "username" not in session:
        return jsonify({"found": False, "error": "Not logged in"}), 401
 
    conn = get_db()
    cursor = conn.cursor()
 
    cursor.execute("""
    SELECT member_id, full_name, status
    FROM members
    WHERE member_id = %s
    """, (member_id,))
 
    member = cursor.fetchone()
    return_db(conn)
 
    if not member:
        return jsonify({"found": False})
 
    return jsonify({
        "found": True,
        "member_id": member[0],
        "full_name": member[1],
        "status": member[2]
    })


# ============================================================
# WIRELESS PHONE SCANNER PAIRING SYSTEM
# Ginagamit para makapag-scan ng QR gamit ang isang phone, at
# automatic na lalabas ang resulta sa ibang device (hal. iPad
# o PC) na naka-display sa Payments/Attendance/Members page.
# ============================================================

import random
import string


def generate_pair_code():
    return "".join(random.choices(string.digits, k=6))


@app.route("/create_pair_session", methods=["POST"])
def create_pair_session():

    if "username" not in session:
        return jsonify({"error": "Not logged in"}), 401

    target_page = request.form.get("target_page", "payments")

    conn = get_db()
    cursor = conn.cursor()

    # Gumawa ng unique 6-digit code (subukan ulit kung sakaling
    # may existing na pareho)
    pair_code = generate_pair_code()

    for _ in range(5):
        cursor.execute(
            "SELECT id FROM pairing_sessions WHERE pair_code = %s",
            (pair_code,)
        )
        if not cursor.fetchone():
            break
        pair_code = generate_pair_code()

    cursor.execute("""
    INSERT INTO pairing_sessions (pair_code, target_page, scanned_value)
    VALUES (%s, %s, NULL)
    """, (pair_code, target_page))

    conn.commit()
    return_db(conn)

    return jsonify({"pair_code": pair_code})


@app.route("/mobile_scan/<pair_code>")
def mobile_scan(pair_code):

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT id, target_page FROM pairing_sessions WHERE pair_code = %s",
        (pair_code,)
    )
    session_row = cursor.fetchone()
    return_db(conn)

    if not session_row:
        return render_template(
            "mobile_scan.html",
            valid_session=False,
            pair_code=pair_code
        )

    return render_template(
        "mobile_scan.html",
        valid_session=True,
        pair_code=pair_code,
        target_page=session_row[1]
    )


@app.route("/submit_mobile_scan", methods=["POST"])
def submit_mobile_scan():

    pair_code = request.form.get("pair_code", "")
    scanned_value = request.form.get("scanned_value", "")

    if not pair_code or not scanned_value:
        return jsonify({"success": False, "error": "Missing data"})

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    UPDATE pairing_sessions
    SET scanned_value = %s, updated_at = NOW()
    WHERE pair_code = %s
    """, (scanned_value, pair_code))

    conn.commit()
    return_db(conn)

    return jsonify({"success": True})


@app.route("/check_pair_scan/<pair_code>")
def check_pair_scan(pair_code):

    if "username" not in session:
        return jsonify({"error": "Not logged in"}), 401

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT scanned_value FROM pairing_sessions WHERE pair_code = %s
    """, (pair_code,))

    row = cursor.fetchone()

    if row and row[0]:
        # I-clear agad pagkatapos mabasa, para hindi paulit-ulit
        # mag-trigger ang parehong scan
        cursor.execute("""
        UPDATE pairing_sessions SET scanned_value = NULL WHERE pair_code = %s
        """, (pair_code,))
        conn.commit()
        return_db(conn)
        return jsonify({"has_scan": True, "value": row[0]})

    return_db(conn)
    return jsonify({"has_scan": False})


@app.route("/settings")
def settings():
 
    if "username" not in session:
        return redirect("/login")
 
    conn = get_db()
    cursor = conn.cursor()
 
    cursor.execute("SELECT username, role FROM users ORDER BY username")
    all_users = cursor.fetchall()
 
    return_db(conn)
 
    return render_template(
        "settings.html",
        username=session["username"],
        all_users=all_users
    )
 
 
@app.route("/settings/change_username", methods=["POST"])
def change_username():
 
    if "username" not in session:
        return redirect("/login")
 
    new_username = request.form["new_username"].strip()
    current_password = request.form["current_password"]
 
    conn = get_db()
    cursor = conn.cursor()
 
    # Kunin ang current user's stored password
    cursor.execute(
        "SELECT password FROM users WHERE username = %s",
        (session["username"],)
    )
    row = cursor.fetchone()
 
    if not row:
        return_db(conn)
        return redirect("/login")
 
    stored_password = row[0]
 
    # Verify current password (supports hashed or plain)
    if stored_password.startswith("pbkdf2:") or stored_password.startswith("scrypt:"):
        password_ok = check_password_hash(stored_password, current_password)
    else:
        password_ok = (stored_password == current_password)
 
    if not password_ok:
        return_db(conn)
        return render_template(
            "settings.html",
            username=session["username"],
            all_users=[],
            error="Maling current password. Subukan ulit."
        )
 
    # Check kung taken na ang bagong username
    # (kasama ang reserved developer account name)
    cursor.execute(
        "SELECT COUNT(*) FROM users WHERE username = %s",
        (new_username,)
    )
    taken = cursor.fetchone()[0]

    if (taken > 0 and new_username != session["username"]) or new_username == "paulo20":
        return_db(conn)
        return render_template(
            "settings.html",
            username=session["username"],
            all_users=[],
            error=f'Ang username "{new_username}" ay ginagamit na.'
        )
 
    # I-update ang username
    cursor.execute(
        "UPDATE users SET username = %s WHERE username = %s",
        (new_username, session["username"])
    )
    conn.commit()
    return_db(conn)
 
    session["username"] = new_username
 
    cursor = get_db().cursor()
 
    return redirect("/settings")
 
 
@app.route("/settings/change_password", methods=["POST"])
def change_password():
 
    if "username" not in session:
        return redirect("/login")
 
    current_password = request.form["current_password"]
    new_password = request.form["new_password"]
    confirm_password = request.form["confirm_password"]
 
    conn = get_db()
    cursor = conn.cursor()
 
    cursor.execute(
        "SELECT password FROM users WHERE username = %s",
        (session["username"],)
    )
    row = cursor.fetchone()
 
    if not row:
        return_db(conn)
        return redirect("/login")
 
    stored_password = row[0]
 
    if stored_password.startswith("pbkdf2:") or stored_password.startswith("scrypt:"):
        password_ok = check_password_hash(stored_password, current_password)
    else:
        password_ok = (stored_password == current_password)
 
    if not password_ok:
        return_db(conn)
        cursor2 = get_db().cursor()
        cursor2.execute("SELECT username, role FROM users ORDER BY username")
        all_users = cursor2.fetchall()
        return render_template(
            "settings.html",
            username=session["username"],
            all_users=all_users,
            error="Maling current password."
        )
 
    if new_password != confirm_password:
        return_db(conn)
        cursor2 = get_db().cursor()
        cursor2.execute("SELECT username, role FROM users ORDER BY username")
        all_users = cursor2.fetchall()
        return render_template(
            "settings.html",
            username=session["username"],
            all_users=all_users,
            error="Hindi tugma ang New Password at Confirm Password."
        )
 
    hashed_password = generate_password_hash(new_password)
 
    cursor.execute(
        "UPDATE users SET password = %s WHERE username = %s",
        (hashed_password, session["username"])
    )
    conn.commit()
    return_db(conn)
 
    cursor2 = get_db().cursor()
    cursor2.execute("SELECT username, role FROM users ORDER BY username")
    all_users = cursor2.fetchall()
 
    return render_template(
        "settings.html",
        username=session["username"],
        all_users=all_users,
        message="Matagumpay na na-update ang password!"
    )
 
 
@app.route("/settings/add_user", methods=["POST"])
def add_user():
 
    if "username" not in session:
        return redirect("/login")
 
    new_username = request.form["username"].strip()
    new_password = request.form["password"]
    role = request.form["role"]
 
    conn = get_db()
    cursor = conn.cursor()
 
    cursor.execute(
        "SELECT COUNT(*) FROM users WHERE username = %s",
        (new_username,)
    )
    taken = cursor.fetchone()[0]
 
    if taken > 0 or new_username == "paulo20":
        return_db(conn)
        cursor2 = get_db().cursor()
        cursor2.execute("SELECT username, role FROM users ORDER BY username")
        all_users = cursor2.fetchall()
        return render_template(
            "settings.html",
            username=session["username"],
            all_users=all_users,
            error=f'Ang username "{new_username}" ay ginagamit na.'
        )
 
    hashed_password = generate_password_hash(new_password)
 
    cursor.execute(
        "INSERT INTO users (username, password, role) VALUES (%s, %s, %s)",
        (new_username, hashed_password, role)
    )
    conn.commit()
    return_db(conn)
 
    cursor2 = get_db().cursor()
    cursor2.execute("SELECT username, role FROM users ORDER BY username")
    all_users = cursor2.fetchall()
 
    return render_template(
        "settings.html",
        username=session["username"],
        all_users=all_users,
        message=f'Matagumpay na nagawa ang user "{new_username}"!'
    )

@app.route("/delete_donation/<int:donation_id>")
def delete_donation(donation_id):
 
    if "username" not in session:
        return redirect("/login")
 
    conn = get_db()
    cursor = conn.cursor()
 
    cursor.execute(
        "DELETE FROM donations WHERE id = %s",
        (donation_id,)
    )
 
    conn.commit()
    return_db(conn)
 
    return redirect("/donations")

@app.route("/settings/backup_database")
def backup_database():

    if "username" not in session:
        return redirect("/login")

    return render_template(
        "settings.html",
        username=session["username"],
        all_users=get_all_users(),
        message="Ang database ay naka-Supabase na (cloud). Para mag-backup, pumunta sa iyong Supabase dashboard → Project → Backups."
    )

@app.route("/settings/restore_database", methods=["POST"])
def restore_database():

    if "username" not in session:
        return redirect("/login")

    return render_template(
        "settings.html",
        username=session["username"],
        all_users=get_all_users(),
        message="Ang database ay naka-Supabase na (cloud). Ang Restore from .db file ay hindi na available. Para mag-restore, gamitin ang Supabase dashboard → Project → Backups."
    )

@app.route("/settings/merge_database", methods=["POST"])
def merge_database():

    if "username" not in session:
        return redirect("/login")

    return render_template(
        "settings.html",
        username=session["username"],
        all_users=get_all_users(),
        message="Ang database ay naka-Supabase na (cloud). Ang Merge from .db file ay hindi na available."
    )

def merge_old_database(old_db_path):
    """
    Kinukuha ang lahat ng members, payments, donations, at expenses
    mula sa lumang offline .db file at idinadagdag (INSERT) sila sa
    kasalukuyang database, hindi pinapalitan ang existing records.

    SAFETY: Gumagamit lang ng mga column na PARESHAS sa dalawang
    database (matching column names). Kung may column sa luma na
    wala sa bago (o vice versa), hindi ito sasama sa insert at
    hindi magiging error — para hindi masira ang merge kung may
    konting pagkaiba sa schema.

    Member duplicates ay kinakheck gamit ang member_id (skip kung
    existing na). Payments/donations/expenses ay direktang idinadagdag
    dahil walang natural duplicate key dito.
    """

    old_conn = sqlite3.connect(old_db_path)
    old_cursor = old_conn.cursor()

    new_conn = get_db()
    new_cursor = new_conn.cursor()

    summary_parts = []

    def get_columns(cursor, table_name):
        # Whitelist para maiwasan ang SQL injection
        allowed = ["members","member_photos","payments","donations","expenses","attendance","feed_posts"]
        if table_name not in allowed:
            return []
        cursor.execute(f"PRAGMA table_info({table_name})")
        return [col[1] for col in cursor.fetchall()]

    def table_exists(cursor, table_name):
        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=%s",
            (table_name,)
        )
        return cursor.fetchone() is not None

    # ── MEMBERS (may duplicate check gamit ang member_id) ──
    if table_exists(old_cursor, "members") and table_exists(new_cursor, "members"):

        old_columns = get_columns(old_cursor, "members")
        new_columns = get_columns(new_cursor, "members")

        # Gamitin lang ang columns na PARESHAS sa dalawa, laktawan ang 'id'
        common_columns = [c for c in old_columns if c in new_columns and c != "id"]

        if "member_id" not in common_columns:
            summary_parts.append("Members: hindi na-merge — walang 'member_id' column na pareho sa dalawang database")
        else:
            # common_columns comes from schema inspection, not user input
            old_cursor.execute(f"SELECT {', '.join(common_columns)} FROM members")
            old_members = old_cursor.fetchall()

            member_id_index = common_columns.index("member_id")

            added_members = 0
            skipped_members = 0

            for row in old_members:
                member_id = row[member_id_index]

                new_cursor.execute(
                    "SELECT COUNT(*) FROM members WHERE member_id = %s",
                    (member_id,)
                )
                exists = new_cursor.fetchone()[0]

                if exists:
                    skipped_members += 1
                    continue

                placeholders = ", ".join(["%s"] * len(common_columns))
                columns_str = ", ".join(common_columns)

                new_cursor.execute(
                    f"INSERT INTO members ({columns_str}) VALUES ({placeholders})",
                    row
                )

                added_members += 1

            summary_parts.append(f"Members: {added_members} idinagdag, {skipped_members} na-skip (duplicate)")
    else:
        summary_parts.append("Members: walang nahanap na table")

    # ── MEMBER_PHOTOS (hiwalay na table, gamit ng offline version) ──
    if table_exists(old_cursor, "member_photos") and table_exists(new_cursor, "member_photos"):

        old_columns = get_columns(old_cursor, "member_photos")
        new_columns = get_columns(new_cursor, "member_photos")
        common_columns = [c for c in old_columns if c in new_columns and c != "id"]

        if common_columns and "member_id" in common_columns:
            old_cursor.execute(f"SELECT {', '.join(common_columns)} FROM member_photos")
            old_photos = old_cursor.fetchall()

            member_id_index = common_columns.index("member_id")

            placeholders = ", ".join(["%s"] * len(common_columns))
            columns_str = ", ".join(common_columns)

            added_photos = 0

            for row in old_photos:
                member_id = row[member_id_index]

                # Skip kung may existing na photo record na ang member na ito
                new_cursor.execute(
                    "SELECT COUNT(*) FROM member_photos WHERE member_id = %s",
                    (member_id,)
                )
                exists = new_cursor.fetchone()[0]

                if exists:
                    continue

                new_cursor.execute(
                    f"INSERT INTO member_photos ({columns_str}) VALUES ({placeholders})",
                    row
                )
                added_photos += 1

            summary_parts.append(f"Member Photos: {added_photos} idinagdag")
        else:
            summary_parts.append("Member Photos: walang pareho na columns")
    else:
        summary_parts.append("Member Photos: walang nahanap na table (okay lang, optional)")

    # ── PAYMENTS (walang duplicate check, direktang idagdag) ──
    if table_exists(old_cursor, "payments") and table_exists(new_cursor, "payments"):

        old_columns = get_columns(old_cursor, "payments")
        new_columns = get_columns(new_cursor, "payments")
        common_columns = [c for c in old_columns if c in new_columns and c != "id"]

        if common_columns:
            old_cursor.execute(f"SELECT {', '.join(common_columns)} FROM payments")
            old_payments = old_cursor.fetchall()

            placeholders = ", ".join(["%s"] * len(common_columns))
            columns_str = ", ".join(common_columns)

            for row in old_payments:
                new_cursor.execute(
                    f"INSERT INTO payments ({columns_str}) VALUES ({placeholders})",
                    row
                )

            summary_parts.append(f"Payments: {len(old_payments)} idinagdag")
        else:
            summary_parts.append("Payments: walang pareho na columns")
    else:
        summary_parts.append("Payments: walang nahanap na table")

    # ── DONATIONS (walang duplicate check, direktang idagdag) ──
    if table_exists(old_cursor, "donations") and table_exists(new_cursor, "donations"):

        old_columns = get_columns(old_cursor, "donations")
        new_columns = get_columns(new_cursor, "donations")
        common_columns = [c for c in old_columns if c in new_columns and c != "id"]

        if common_columns:
            old_cursor.execute(f"SELECT {', '.join(common_columns)} FROM donations")
            old_donations = old_cursor.fetchall()

            placeholders = ", ".join(["%s"] * len(common_columns))
            columns_str = ", ".join(common_columns)

            for row in old_donations:
                new_cursor.execute(
                    f"INSERT INTO donations ({columns_str}) VALUES ({placeholders})",
                    row
                )

            summary_parts.append(f"Donations: {len(old_donations)} idinagdag")
        else:
            summary_parts.append("Donations: walang pareho na columns")
    else:
        summary_parts.append("Donations: walang nahanap na table")

    # ── EXPENSES (walang duplicate check, direktang idagdag) ──
    if table_exists(old_cursor, "expenses") and table_exists(new_cursor, "expenses"):

        old_columns = get_columns(old_cursor, "expenses")
        new_columns = get_columns(new_cursor, "expenses")
        common_columns = [c for c in old_columns if c in new_columns and c != "id"]

        if common_columns:
            old_cursor.execute(f"SELECT {', '.join(common_columns)} FROM expenses")
            old_expenses = old_cursor.fetchall()

            placeholders = ", ".join(["%s"] * len(common_columns))
            columns_str = ", ".join(common_columns)

            for row in old_expenses:
                new_cursor.execute(
                    f"INSERT INTO expenses ({columns_str}) VALUES ({placeholders})",
                    row
                )

            summary_parts.append(f"Expenses: {len(old_expenses)} idinagdag")
        else:
            summary_parts.append("Expenses: walang pareho na columns")
    else:
        summary_parts.append("Expenses: walang nahanap na table")

    # ── ATTENDANCE (walang duplicate check, direktang idagdag) ──
    if table_exists(old_cursor, "attendance") and table_exists(new_cursor, "attendance"):

        old_columns = get_columns(old_cursor, "attendance")
        new_columns = get_columns(new_cursor, "attendance")
        common_columns = [c for c in old_columns if c in new_columns and c != "id"]

        if common_columns:
            old_cursor.execute(f"SELECT {', '.join(common_columns)} FROM attendance")
            old_attendance = old_cursor.fetchall()

            placeholders = ", ".join(["%s"] * len(common_columns))
            columns_str = ", ".join(common_columns)

            for row in old_attendance:
                new_cursor.execute(
                    f"INSERT INTO attendance ({columns_str}) VALUES ({placeholders})",
                    row
                )

            summary_parts.append(f"Attendance: {len(old_attendance)} idinagdag")
    else:
        summary_parts.append("Attendance: walang nahanap na table (okay lang, optional)")

    new_conn.commit()
    new_return_db(conn)
    old_return_db(conn)

    return "Matagumpay na na-merge ang backup! " + " | ".join(summary_parts)


def get_all_users():
    """Helper function para makuha ang listahan ng users (ginagamit sa settings page)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT username, role FROM users ORDER BY username")
    users = cursor.fetchall()
    return_db(conn)
    return users

@app.route("/member_login", methods=["GET", "POST"])
def member_login():

    # ── RATE LIMITING — max 5 attempts per IP per 5 minutes ──
    # Ginagamit ang Flask session para i-track ang failed attempts
    # Simple pero epektibo para sa FCCI's scale
    if "login_attempts" not in session:
        session["login_attempts"] = 0
        session["login_lockout_until"] = 0

    import time
    now_ts = time.time()

    # I-check kung naka-lockout
    if session.get("login_lockout_until", 0) > now_ts:
        remaining = int(session["login_lockout_until"] - now_ts)
        return render_template(
            "member_login.html",
            error=f"Napakaraming maling pagsubok. Subukan ulit pagkatapos ng {remaining} segundo.",
            open_events=_get_open_events()
        )

    if request.method == "POST":

        member_id      = request.form["member_id"].strip()
        birthday_input = request.form["birthday"]

        conn   = get_db()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT member_id, full_name, birthday, status
        FROM members
        WHERE member_id = %s
        """, (member_id,))

        member = cursor.fetchone()
        return_db(conn)

        if not member:
            session["login_attempts"] = session.get("login_attempts", 0) + 1
            if session["login_attempts"] >= 5:
                session["login_lockout_until"] = now_ts + 300  # 5 minuto
                session["login_attempts"] = 0
            return render_template(
                "member_login.html",
                error="Member ID na hindi natagpuan. Pakitiyak na tama ang inilagay mo.",
                open_events=_get_open_events()
            )

        stored_birthday = member[2]
        status          = member[3]

        normalized_stored = normalize_birthday(stored_birthday)

        if normalized_stored != birthday_input:
            session["login_attempts"] = session.get("login_attempts", 0) + 1
            if session["login_attempts"] >= 5:
                session["login_lockout_until"] = now_ts + 300  # 5 minuto
                session["login_attempts"] = 0
                return render_template(
                    "member_login.html",
                    error="Napakaraming maling pagsubok. Sandali lang at subukan ulit.",
                    open_events=_get_open_events()
                )
            return render_template(
                "member_login.html",
                error="Maling Member ID o Birthday. Subukan ulit.",
                open_events=_get_open_events()
            )

        if status != "Active":
            return render_template(
                "member_login.html",
                error=f"Hindi mo pa magagamit ang portal na ito. Ang status mo ay '{status}'. Pakibayaran muna ang Registration Fee sa FCCI office para ma-activate ang account mo.",
                open_events=_get_open_events()
            )

        # Successful login
        session["member_logged_in"] = True
        session["member_id"] = member[0]
        session["member_name"] = member[1]

        return redirect("/feed")

    return render_template("member_login.html", open_events=_get_open_events())


def normalize_birthday(raw_value):
    """
    Sinusubukan i-convert ang anumang stored birthday format papunta sa
    'YYYY-MM-DD' para ma-compare sa value mula sa HTML date input.
    """
    if not raw_value:
        return ""

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            parsed = datetime.strptime(raw_value, fmt)
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue

    return raw_value  # fallback: ibalik na lang as-is


# ── MEMBER LOGOUT ──────────────────────────────────────────
@app.route("/member_logout")
def member_logout():
    session.pop("member_logged_in", None)
    session.pop("member_id", None)
    session.pop("member_name", None)
    return redirect("/member_login")


# ── DECORATOR/HELPER: I-CHECK KUNG NAKA-LOGIN ANG MEMBER ───
def require_member_login():
    """Tawagin ito sa simula ng bawat member-portal route."""
    if not session.get("member_logged_in"):
        return False
    return True

EVENT_TYPES = [
    "Basketball",
    "Volleyball",
    "Feeding Program",
    "General Assembly",
    "Fun Run",
    "Fiesta / Celebration",
    "Other",
]

# Event types na "team-based" (may Team Name + list ng members).
# Ang lahat ng iba pang types ay "RSVP-based" (simpleng sign-up lang).
EVENT_TEAM_TYPES = {"Basketball", "Volleyball"}

EVENT_TYPE_ICONS = {
    "Basketball": "🏀",
    "Volleyball": "🏐",
    "Feeding Program": "🍲",
    "General Assembly": "📋",
    "Fun Run": "🏃",
    "Fiesta / Celebration": "🎉",
    "Other": "📌",
}


def _event_is_team_mode(event_type):
    return event_type in EVENT_TEAM_TYPES


def _get_feed_event_registrations(cursor, post_id):
    """Ibinabalik ang lahat ng registrations ng isang event, kasama
    ang team members (kung team-based) bilang listahan ng pangalan."""
    cursor.execute("""
        SELECT id, member_id, is_guest, team_name, captain_name,
               contact, companions, notes, status, registered_date, registered_time
        FROM feed_event_registrations
        WHERE post_id = %s
        ORDER BY id ASC
    """, (post_id,))
    rows = cursor.fetchall()

    regs = []
    for r in rows:
        reg_id = r[0]
        cursor.execute("""
            SELECT member_name FROM feed_event_registration_members
            WHERE registration_id = %s ORDER BY id ASC
        """, (reg_id,))
        members = [m[0] for m in cursor.fetchall()]

        regs.append({
            "id": reg_id,
            "member_id": r[1],
            "is_guest": r[2],
            "team_name": r[3],
            "captain_name": r[4],
            "contact": r[5],
            "companions": r[6],
            "notes": r[7],
            "status": r[8],
            "registered_date": r[9],
            "registered_time": r[10],
            "members": members,
        })
    return regs


def _get_open_events():
    """Ibinabalik ang listahan ng mga event na bukas pa ang registration —
    ginagamit ito sa public widget sa member_login page (kahit hindi naka-login)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, event_type, event_title, content, event_date, event_location, max_slots
        FROM feed_posts
        WHERE is_event = TRUE AND registration_closed = FALSE
        ORDER BY id DESC
        LIMIT 5
    """)
    rows = cursor.fetchall()

    events = []
    for r in rows:
        post_id = r[0]
        cursor.execute("""
            SELECT COUNT(*) FROM feed_event_registrations
            WHERE post_id = %s AND status = 'Confirmed'
        """, (post_id,))
        join_count = cursor.fetchone()[0]
        max_slots = r[6]

        events.append({
            "id": post_id,
            "type": r[1],
            "icon": EVENT_TYPE_ICONS.get(r[1], "📌"),
            "title": r[2] or r[1],
            "description": r[3],
            "date": r[4],
            "location": r[5],
            "max_slots": max_slots,
            "slots_remaining": (max_slots - join_count) if max_slots else None,
            "join_count": join_count,
            "team_mode": _event_is_team_mode(r[1]),
        })

    return_db(conn)
    return events


@app.route("/feed", methods=["GET", "POST"])
def feed():

    if not session.get("member_logged_in"):
        return redirect("/member_login")

    conn = get_db()
    cursor = conn.cursor()

    if request.method == "POST":

        content = request.form.get("content", "").strip()
        photo_file = request.files.get("photo")

        photo_filename = None

        if photo_file and photo_file.filename != "":
            photo_filename = upload_photo(photo_file, folder="fcci_feed") or None

        if content or photo_filename:
            now = datetime.now()

            cursor.execute("""
            INSERT INTO feed_posts
            (member_id, full_name, content, photo_path, is_pinned, post_date, post_time)
            VALUES (%s, %s, %s, %s, FALSE, %s, %s)
            """, (
                session["member_id"],
                session["member_name"],
                content,
                photo_filename,
                now.strftime("%B %d, %Y"),
                now.strftime("%I:%M %p")
            ))

            conn.commit()

        return_db(conn)
        return redirect("/feed")

    # GET request: kunin lahat ng posts, pinned muna, tapos pinaka-bago
    cursor.execute("""
    SELECT id, member_id, full_name, content, photo_path, is_pinned, post_date, post_time,
           is_event, event_type, event_title, event_date, event_location,
           max_slots, registration_closed, event_time
    FROM feed_posts
    ORDER BY is_pinned DESC, id DESC
    """)
    posts = cursor.fetchall()

    # Bilangin ang likes at comments bawat post
    posts_with_meta = []

    for post in posts:
        post_id = post[0]

        cursor.execute("SELECT COUNT(*) FROM feed_likes WHERE post_id = %s", (post_id,))
        like_count = cursor.fetchone()[0]

        cursor.execute("""
        SELECT COUNT(*) FROM feed_likes WHERE post_id = %s AND member_id = %s
        """, (post_id, session["member_id"]))
        liked_by_me = cursor.fetchone()[0] > 0

        cursor.execute("""
        SELECT full_name, comment_text, comment_date, comment_time
        FROM feed_comments WHERE post_id = %s ORDER BY id ASC
        """, (post_id,))
        comments = cursor.fetchall()

        is_event = post[8]
        event_info = None

        if is_event:
            event_type = post[9]
            team_mode = _event_is_team_mode(event_type)

            cursor.execute("""
                SELECT COUNT(*) FROM feed_event_registrations
                WHERE post_id = %s AND status != 'Cancelled'
            """, (post_id,))
            join_count = cursor.fetchone()[0]

            cursor.execute("""
                SELECT id, team_name, captain_name, contact, companions, status
                FROM feed_event_registrations
                WHERE post_id = %s AND member_id = %s
            """, (post_id, session["member_id"]))
            my_reg_row = cursor.fetchone()
            my_registration = None
            if my_reg_row:
                cursor.execute("""
                    SELECT member_name FROM feed_event_registration_members
                    WHERE registration_id = %s ORDER BY id ASC
                """, (my_reg_row[0],))
                my_registration = {
                    "id": my_reg_row[0],
                    "team_name": my_reg_row[1],
                    "captain_name": my_reg_row[2],
                    "contact": my_reg_row[3],
                    "companions": my_reg_row[4],
                    "status": my_reg_row[5],
                    "members": [m[0] for m in cursor.fetchall()],
                }

            max_slots = post[13]
            slots_remaining = (max_slots - join_count) if max_slots else None

            event_info = {
                "type": event_type,
                "icon": EVENT_TYPE_ICONS.get(event_type, "📌"),
                "team_mode": team_mode,
                "title": post[10],
                "date": post[11],
                "location": post[12],
                "time": post[15] if len(post) > 15 else None,
                "max_slots": max_slots,
                "slots_remaining": slots_remaining,
                "registration_closed": post[14],
                "join_count": join_count,
                "join_label": "teams joined" if team_mode else "confirmed",
                "my_registration": my_registration,
            }

        posts_with_meta.append({
            "id": post[0],
            "member_id": post[1],
            "full_name": post[2],
            "content": post[3],
            "photo_path": post[4],
            "is_pinned": post[5],
            "post_date": post[6],
            "post_time": post[7],
            "like_count": like_count,
            "liked_by_me": liked_by_me,
            "comments": comments,
            "is_event": is_event,
            "event": event_info,
        })

    return_db(conn)

    return render_template(
        "feed.html",
        posts=posts_with_meta,
        member_name=session["member_name"],
        member_id=session["member_id"]
    )


# ── DELETE POST (sariling post lang) ───────────────────────
@app.route("/feed/delete/<int:post_id>")
def feed_delete_post(post_id):

    if not session.get("member_logged_in"):
        return redirect("/member_login")

    conn = get_db()
    cursor = conn.cursor()

    # Siguraduhing sariling post lang ng member ang puwedeng i-delete
    cursor.execute(
        "DELETE FROM feed_posts WHERE id = %s AND member_id = %s",
        (post_id, session["member_id"])
    )

    conn.commit()
    return_db(conn)

    return redirect("/feed")


# ── EDIT POST (sariling post lang ang puwedeng baguhin) ─────
@app.route("/feed/edit/<int:post_id>", methods=["POST"])
def feed_edit_post(post_id):

    if not session.get("member_logged_in"):
        return redirect("/member_login")

    new_content = request.form.get("content", "").strip()

    conn = get_db()
    cursor = conn.cursor()

    # I-verify muna na ang post ay pag-aari ng member na ito
    cursor.execute(
        "SELECT member_id FROM feed_posts WHERE id = %s",
        (post_id,)
    )
    row = cursor.fetchone()

    if not row or row[0] != session["member_id"]:
        # Hindi sariling post — huwag payagang mag-edit
        return_db(conn)
        return redirect("/feed")

    # I-update lang ang content (sariling post)
    now = datetime.now()
    cursor.execute("""
        UPDATE feed_posts
        SET content = %s,
            post_date = %s,
            post_time = %s
        WHERE id = %s AND member_id = %s
    """, (
        new_content,
        now.strftime("%B %d, %Y"),
        now.strftime("%I:%M %p"),
        post_id,
        session["member_id"],
    ))

    conn.commit()
    return_db(conn)

    return redirect("/feed")


# ── LIKE / UNLIKE POST ──────────────────────────────────────
@app.route("/feed/like/<int:post_id>")
def feed_like_post(post_id):

    if not session.get("member_logged_in"):
        return redirect("/member_login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT id FROM feed_likes WHERE post_id = %s AND member_id = %s",
        (post_id, session["member_id"])
    )
    existing = cursor.fetchone()

    if existing:
        cursor.execute("DELETE FROM feed_likes WHERE id = %s", (existing[False],))
    else:
        cursor.execute(
            "INSERT INTO feed_likes (post_id, member_id) VALUES (%s, %s)",
            (post_id, session["member_id"])
        )

    conn.commit()
    return_db(conn)

    return redirect("/feed")


# ── ADD COMMENT ──────────────────────────────────────────────
@app.route("/feed/comment/<int:post_id>", methods=["POST"])
def feed_add_comment(post_id):

    if not session.get("member_logged_in"):
        return redirect("/member_login")

    comment_text = request.form.get("comment_text", "").strip()

    if comment_text:
        conn = get_db()
        cursor = conn.cursor()

        now = datetime.now()

        cursor.execute("""
        INSERT INTO feed_comments (post_id, member_id, full_name, comment_text, comment_date, comment_time)
        VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            post_id,
            session["member_id"],
            session["member_name"],
            comment_text,
            now.strftime("%B %d, %Y"),
            now.strftime("%I:%M %p")
        ))

        conn.commit()
        return_db(conn)

    return redirect("/feed")


# ── PIN / UNPIN POST (ADMIN/STAFF LANG, gamit ang main session) ──
@app.route("/feed/pin/<int:post_id>")
def feed_pin_post(post_id):

    # Gamit ang ADMIN session (yung "username" sa session, hindi member portal session)
    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT is_pinned FROM feed_posts WHERE id = %s", (post_id,))
    row = cursor.fetchone()

    if row:
        new_value = False if row[False] == True else True
        cursor.execute("UPDATE feed_posts SET is_pinned = %s WHERE id = %s", (new_value, post_id))
        conn.commit()

    return_db(conn)

    # Ibalik sa pinanggalingang page (admin feed view o member feed)
    return redirect(request.referrer or "/feed")


# ── JOIN EVENT (member, naka-login) ────────────────────────
@app.route("/feed/event/<int:post_id>/join", methods=["POST"])
def feed_event_join(post_id):

    if not session.get("member_logged_in"):
        return redirect("/member_login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT event_type, max_slots, registration_closed FROM feed_posts
        WHERE id = %s AND is_event = TRUE
    """, (post_id,))
    event_row = cursor.fetchone()

    if not event_row:
        return_db(conn)
        return redirect("/feed")

    event_type, max_slots, closed = event_row

    if closed:
        return_db(conn)
        return redirect("/feed")

    # Isang registration lang bawat member kada event
    cursor.execute("""
        SELECT id FROM feed_event_registrations WHERE post_id = %s AND member_id = %s
    """, (post_id, session["member_id"]))
    if cursor.fetchone():
        return_db(conn)
        return redirect("/feed")

    now = datetime.now()
    team_mode = _event_is_team_mode(event_type)

    if team_mode:
        team_name = request.form.get("team_name", "").strip() or "Unnamed Team"
        captain_name = request.form.get("captain_name", "").strip() or session["member_name"]
        contact = request.form.get("contact", "").strip()
        member_names = [
            m.strip() for m in request.form.getlist("members[]") if m.strip()
        ]
        companions = len(member_names)
    else:
        team_name = None
        captain_name = request.form.get("rsvp_name", "").strip() or session["member_name"]
        contact = request.form.get("contact", "").strip()
        companions = int(request.form.get("companions", 1) or 1)
        member_names = []

    notes = request.form.get("notes", "").strip()

    # I-check kung puno na (kung may max_slots) — kung puno, Waitlisted
    status = "Confirmed"
    if max_slots:
        cursor.execute("""
            SELECT COUNT(*) FROM feed_event_registrations
            WHERE post_id = %s AND status = 'Confirmed'
        """, (post_id,))
        current_count = cursor.fetchone()[0]
        if current_count >= max_slots:
            status = "Waitlisted"

    cursor.execute("""
        INSERT INTO feed_event_registrations
        (post_id, member_id, is_guest, team_name, captain_name, contact,
         companions, notes, status, registered_date, registered_time)
        VALUES (%s, %s, FALSE, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (
        post_id, session["member_id"], team_name, captain_name, contact,
        companions, notes, status,
        now.strftime("%B %d, %Y"), now.strftime("%I:%M %p")
    ))
    reg_id = cursor.fetchone()[0]

    for m_name in member_names:
        cursor.execute("""
            INSERT INTO feed_event_registration_members (registration_id, member_name)
            VALUES (%s, %s)
        """, (reg_id, m_name))

    conn.commit()
    return_db(conn)

    return redirect("/feed")


# ── CANCEL MY EVENT REGISTRATION ───────────────────────────
@app.route("/feed/event/<int:post_id>/cancel")
def feed_event_cancel(post_id):

    if not session.get("member_logged_in"):
        return redirect("/member_login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        DELETE FROM feed_event_registrations WHERE post_id = %s AND member_id = %s
    """, (post_id, session["member_id"]))

    conn.commit()
    return_db(conn)

    return redirect("/feed")


# ── GUEST JOIN (walang account/login — mula sa login page widget) ──
@app.route("/guest_event/<int:post_id>/join", methods=["POST"])
def guest_event_join(post_id):

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT event_type, max_slots, registration_closed FROM feed_posts
        WHERE id = %s AND is_event = TRUE
    """, (post_id,))
    event_row = cursor.fetchone()

    if not event_row:
        return_db(conn)
        return redirect("/member_login")

    event_type, max_slots, closed = event_row

    if closed:
        return_db(conn)
        return redirect("/member_login")

    now = datetime.now()
    team_mode = _event_is_team_mode(event_type)

    if team_mode:
        team_name = request.form.get("team_name", "").strip() or "Unnamed Team"
        captain_name = request.form.get("captain_name", "").strip() or "Guest"
        contact = request.form.get("contact", "").strip()
        member_names = [
            m.strip() for m in request.form.getlist("members[]") if m.strip()
        ]
        companions = len(member_names)
    else:
        team_name = None
        captain_name = request.form.get("rsvp_name", "").strip() or "Guest"
        contact = request.form.get("contact", "").strip()
        companions = int(request.form.get("companions", 1) or 1)
        member_names = []

    notes = request.form.get("notes", "").strip()

    status = "Confirmed"
    if max_slots:
        cursor.execute("""
            SELECT COUNT(*) FROM feed_event_registrations
            WHERE post_id = %s AND status = 'Confirmed'
        """, (post_id,))
        current_count = cursor.fetchone()[0]
        if current_count >= max_slots:
            status = "Waitlisted"

    cursor.execute("""
        INSERT INTO feed_event_registrations
        (post_id, member_id, is_guest, team_name, captain_name, contact,
         companions, notes, status, registered_date, registered_time)
        VALUES (%s, NULL, TRUE, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (
        post_id, team_name, captain_name, contact,
        companions, notes, status,
        now.strftime("%B %d, %Y"), now.strftime("%I:%M %p")
    ))
    reg_id = cursor.fetchone()[0]

    for m_name in member_names:
        cursor.execute("""
            INSERT INTO feed_event_registration_members (registration_id, member_name)
            VALUES (%s, %s)
        """, (reg_id, m_name))

    conn.commit()
    return_db(conn)

    return render_template("guest_event_success.html", team_name=team_name, captain_name=captain_name)


# ── MEMBER-FACING EVENTS VIEW (calendar + standings, read-only) ──
@app.route("/member_events")
def member_events():
    if not session.get("member_logged_in"):
        return redirect("/member_login")

    from datetime import date as _date
    conn = get_db()
    cursor = conn.cursor()

    # Kunin lahat ng events
    cursor.execute("""
        SELECT id, event_type, event_title, content, event_date,
               event_location, max_slots, registration_closed, event_time
        FROM feed_posts
        WHERE is_event = TRUE
        ORDER BY event_date ASC
    """)
    rows = cursor.fetchall()

    today_str = _date.today().isoformat()
    upcoming = []
    for r in rows:
        post_id = r[0]
        event_type = r[1]
        team_mode = _event_is_team_mode(event_type)
        regs = _get_feed_event_registrations(cursor, post_id)
        confirmed = [x for x in regs if x["status"] == "Confirmed"]
        max_slots = r[6]
        slots_remaining = (max_slots - len(confirmed)) if max_slots else None

        ev = {
            "id": post_id, "type": event_type,
            "icon": EVENT_TYPE_ICONS.get(event_type, "📌"),
            "title": r[2] or event_type, "description": r[3],
            "date": r[4], "location": r[5],
            "max_slots": max_slots, "slots_remaining": slots_remaining,
            "registration_closed": r[7], "team_mode": team_mode,
            "time": r[8] if len(r) > 8 else None,
        }
        # I-parse ang petsa para sa date badge (month abbr + day)
        ev["mo_abbr"] = ""
        ev["day_num"] = ""
        try:
            if r[4]:
                _months = ["", "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                           "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
                _mm = int(str(r[4])[5:7])
                _dd = int(str(r[4])[8:10])
                ev["mo_abbr"] = _months[_mm] if 1 <= _mm <= 12 else ""
                ev["day_num"] = _dd
        except Exception:
            pass
        # Paparating lang (petsa >= ngayon, o walang petsa)
        try:
            if not r[4] or str(r[4])[:10] >= today_str:
                upcoming.append(ev)
        except Exception:
            upcoming.append(ev)

    # ── Results & Standings: hanapin ang team events na may matches ──
    tournaments = []
    try:
        cursor.execute("""
            SELECT DISTINCT post_id FROM event_matches
            ORDER BY post_id DESC
        """)
        match_events = [row[0] for row in cursor.fetchall()]

        for pid in match_events:
            # Kunin ang event details
            cursor.execute("""
                SELECT event_type, event_title FROM feed_posts WHERE id = %s
            """, (pid,))
            ev_row = cursor.fetchone()
            if not ev_row:
                continue

            # Kunin ang matches
            cursor.execute("""
                SELECT round_name, round_order, team_a, team_b,
                       score_a, score_b, winner, status
                FROM event_matches WHERE post_id = %s
                ORDER BY round_order DESC, id DESC
            """, (pid,))
            m_rows = cursor.fetchall()

            matches = [{
                "round_name": m[0], "round_order": m[1],
                "team_a": m[2], "team_b": m[3],
                "score_a": m[4], "score_b": m[5],
                "winner": m[6], "status": m[7],
            } for m in m_rows]

            # Standings
            teams = set()
            for m in matches:
                if m["team_a"]: teams.add(m["team_a"])
                if m["team_b"]: teams.add(m["team_b"])
            standings = _compute_standings(matches, list(teams))

            # Champion (winner ng highest round na completed)
            champion = None
            for m in matches:
                if m["status"] == "Completed" and m["winner"]:
                    champion = m["winner"]
                    break

            # Completed results lang
            results = [m for m in matches if m["status"] == "Completed"]

            tournaments.append({
                "id": pid, "type": ev_row[0],
                "icon": EVENT_TYPE_ICONS.get(ev_row[0], "🏆"),
                "title": ev_row[1] or ev_row[0],
                "standings": standings, "results": results,
                "champion": champion,
            })
    except Exception:
        conn.rollback()  # baka wala pang event_matches table

    return_db(conn)

    return render_template("member_events.html",
        upcoming=upcoming, tournaments=tournaments,
        member_name=session.get("member_name", ""))


# ── MEMBER'S OWN PROFILE VIEW (sa loob ng portal) ──────────
@app.route("/member_portal_profile")
def member_portal_profile():

    if not session.get("member_logged_in"):
        return redirect("/member_login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id, member_id, full_name, contact, address, registration_fee, member_since, email, birthday, date_registered, status, proof_of_payment FROM members WHERE member_id = %s", (session["member_id"],))
    member = cursor.fetchone()

    cursor.execute("""
    SELECT payment_date, payment_type, amount
    FROM payments WHERE member_id = %s ORDER BY id DESC
    """, (session["member_id"],))
    payments = cursor.fetchall()

    cursor.execute("""
    SELECT COALESCE(SUM(amount), 0) FROM payments WHERE member_id = %s
    """, (session["member_id"],))
    total_payment = cursor.fetchone()[0]

    cursor.execute("""
    SELECT photo_path FROM member_photos WHERE member_id = %s ORDER BY id DESC LIMIT 1
    """, (session["member_id"],))
    photo_row = cursor.fetchone()
    photo_path = photo_row[0] if photo_row else None

    return_db(conn)

    return render_template(
        "member_portal_profile.html",
        member=member,
        payments=payments,
        total_payment=total_payment,
        photo_path=photo_path
    )


@app.route("/member_portal_edit", methods=["GET", "POST"])
def member_portal_edit():

    if not session.get("member_logged_in"):
        return redirect("/member_login")

    member_id = session["member_id"]
    conn = get_db()
    cursor = conn.cursor()

    if request.method == "POST":

        contact  = request.form.get("contact", "").strip()
        address  = request.form.get("address", "").strip()
        email    = request.form.get("email", "").strip()
        birthday = request.form.get("birthday", "").strip()
        photo    = request.files.get("photo")

        # I-update ang member info — hindi pwedeng palitan
        # ang full_name at member_id (admin lang ang pwede)
        cursor.execute("""
        UPDATE members
        SET contact  = %s,
            address  = %s,
            email    = %s,
            birthday = %s
        WHERE member_id = %s
        """, (contact, address, email, birthday, member_id))

        # I-upload ang bagong photo kung may nilagay
        if photo and photo.filename:
            photo_url = upload_photo(photo, folder="fcci_member_photos")
            if photo_url:
                cursor.execute("""
                SELECT id FROM member_photos WHERE member_id = %s
                """, (member_id,))
                existing = cursor.fetchone()
                if existing:
                    cursor.execute("""
                    UPDATE member_photos SET photo_path = %s
                    WHERE member_id = %s
                    """, (photo_url, member_id))
                else:
                    cursor.execute("""
                    INSERT INTO member_photos (member_id, photo_path)
                    VALUES (%s, %s)
                    """, (member_id, photo_url))

        conn.commit()
        return_db(conn)
        return redirect("/member_portal_profile")

    # GET — ipakita ang edit form
    cursor.execute("""
    SELECT id, member_id, full_name, contact, address,
           registration_fee, member_since, email, birthday,
           date_registered, status, proof_of_payment
    FROM members WHERE member_id = %s
    """, (member_id,))
    member = cursor.fetchone()

    cursor.execute("""
    SELECT photo_path FROM member_photos
    WHERE member_id = %s ORDER BY id DESC LIMIT 1
    """, (member_id,))
    photo_row = cursor.fetchone()
    photo_path = photo_row[0] if photo_row else None
    return_db(conn)

    return render_template(
        "member_portal_edit.html",
        member=member,
        photo_path=photo_path
    )


@app.route("/admin_feed")
def admin_feed():

    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT id, member_id, full_name, content, photo_path, is_pinned, post_date, post_time,
           is_event, event_type, event_title, event_date, event_location, event_time, max_slots
    FROM feed_posts
    ORDER BY is_pinned DESC, id DESC
    """)
    posts = cursor.fetchall()

    return_db(conn)

    return render_template(
        "admin_feed.html", posts=posts, username=session["username"],
        event_type_icons=EVENT_TYPE_ICONS
    )



@app.route("/logout")
def logout():

    session.clear()

    return redirect("/login")



@app.route("/admin_feed/post", methods=["POST"])
def admin_feed_post():
    if "username" not in session:
        return redirect("/login")
    content_text = request.form.get("content", "").strip()
    photo_file = request.files.get("photo")
    photo_filename = None
    if photo_file and photo_file.filename != "":
        photo_filename = upload_photo(photo_file, folder="fcci_feed")

    is_event = request.form.get("is_event") == "on"
    event_type = request.form.get("event_type", "").strip() if is_event else None
    event_title = request.form.get("event_title", "").strip() if is_event else None
    event_date = request.form.get("event_date", "").strip() if is_event else None
    event_time = request.form.get("event_time", "").strip() if is_event else None
    event_location = request.form.get("event_location", "").strip() if is_event else None
    max_slots_raw = request.form.get("max_slots", "").strip() if is_event else ""
    max_slots = int(max_slots_raw) if max_slots_raw.isdigit() else None

    if content_text or photo_filename or is_event:
        conn = get_db()
        cursor = conn.cursor()
        now = datetime.now()
        cursor.execute("""
        INSERT INTO feed_posts
        (member_id, full_name, content, photo_path, is_pinned, post_date, post_time,
         is_event, event_type, event_title, event_date, event_time, event_location, max_slots, registration_closed)
        VALUES (%s, %s, %s, %s, FALSE, %s, %s, %s, %s, %s, %s, %s, %s, %s, FALSE)
        """, (
            "ADMIN",
            session["username"],
            content_text,
            photo_filename,
            now.strftime("%B %d, %Y"),
            now.strftime("%I:%M %p"),
            is_event,
            event_type,
            event_title,
            event_date,
            event_time,
            event_location,
            max_slots,
        ))
        conn.commit()
        return_db(conn)
    return redirect("/admin_feed")


@app.route("/admin_feed/delete/<int:post_id>")
def admin_feed_delete(post_id):
    if "username" not in session:
        return redirect("/login")
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM feed_posts WHERE id = %s", (post_id,))
    conn.commit()
    return_db(conn)
    return redirect("/admin_feed")


# ── EDIT EVENT (admin) — baguhin ang event details ─────────
@app.route("/admin_feed/edit_event/<int:post_id>", methods=["POST"])
def admin_feed_edit_event(post_id):
    if "username" not in session:
        return redirect("/login")

    event_title = request.form.get("event_title", "").strip()
    event_type = request.form.get("event_type", "").strip()
    event_date = request.form.get("event_date", "").strip()
    event_time = request.form.get("event_time", "").strip()
    event_location = request.form.get("event_location", "").strip()
    content = request.form.get("content", "").strip()
    max_slots_raw = request.form.get("max_slots", "").strip()
    max_slots = int(max_slots_raw) if max_slots_raw.isdigit() else None

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            UPDATE feed_posts
            SET event_title = %s,
                event_type = %s,
                event_date = %s,
                event_time = %s,
                event_location = %s,
                content = %s,
                max_slots = %s
            WHERE id = %s AND is_event = TRUE
        """, (event_title, event_type, event_date, event_time,
              event_location, content, max_slots, post_id))
        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.error(f"[EDIT EVENT] Error: {e}")
    return_db(conn)
    return redirect("/admin_feed")


# ════════════════════════════════════════════════════════════
#  EVENT MANAGEMENT (admin) — tabs per event, manage/print/export
# ════════════════════════════════════════════════════════════

@app.route("/admin_events")
def admin_events():
    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, event_type, event_title, content, event_date, event_location,
               max_slots, registration_closed
        FROM feed_posts
        WHERE is_event = TRUE
        ORDER BY id DESC
    """)
    rows = cursor.fetchall()

    events = []
    for r in rows:
        post_id = r[0]
        event_type = r[1]
        team_mode = _event_is_team_mode(event_type)
        regs = _get_feed_event_registrations(cursor, post_id)

        confirmed = [x for x in regs if x["status"] == "Confirmed"]
        waitlisted = [x for x in regs if x["status"] == "Waitlisted"]

        if team_mode:
            total_people = sum(len(x["members"]) for x in regs)
        else:
            total_people = sum((x["companions"] or 1) for x in regs)

        max_slots = r[6]
        events.append({
            "id": post_id,
            "type": event_type,
            "icon": EVENT_TYPE_ICONS.get(event_type, "📌"),
            "title": r[2] or event_type,
            "description": r[3],
            "date": r[4],
            "location": r[5],
            "max_slots": max_slots,
            "slots_remaining": (max_slots - len(confirmed)) if max_slots else None,
            "registration_closed": r[7],
            "team_mode": team_mode,
            "registrations": regs,
            "confirmed_count": len(confirmed),
            "waitlisted_count": len(waitlisted),
            "total_people": total_people,
        })

    return_db(conn)

    return render_template("admin_events.html", events=events, username=session["username"])


@app.route("/admin_events/<int:post_id>/toggle_close")
def admin_events_toggle_close(post_id):
    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT registration_closed FROM feed_posts WHERE id = %s", (post_id,))
    row = cursor.fetchone()
    if row:
        cursor.execute(
            "UPDATE feed_posts SET registration_closed = %s WHERE id = %s",
            (not row[0], post_id)
        )
        conn.commit()
    return_db(conn)

    return redirect("/admin_events")


@app.route("/admin_events/<int:post_id>/remove/<int:reg_id>")
def admin_events_remove_registration(post_id, reg_id):
    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM feed_event_registrations WHERE id = %s AND post_id = %s", (reg_id, post_id))
    conn.commit()
    return_db(conn)

    return redirect("/admin_events")


def _get_single_event(cursor, post_id):
    cursor.execute("""
        SELECT id, event_type, event_title, content, event_date, event_location, max_slots
        FROM feed_posts WHERE id = %s AND is_event = TRUE
    """, (post_id,))
    r = cursor.fetchone()
    if not r:
        return None
    event_type = r[1]
    team_mode = _event_is_team_mode(event_type)
    regs = _get_feed_event_registrations(cursor, post_id)
    confirmed = [x for x in regs if x["status"] == "Confirmed"]
    waitlisted = [x for x in regs if x["status"] == "Waitlisted"]
    if team_mode:
        total_people = sum(len(x["members"]) for x in regs)
    else:
        total_people = sum((x["companions"] or 1) for x in regs)

    return {
        "id": r[0],
        "type": event_type,
        "icon": EVENT_TYPE_ICONS.get(event_type, "📌"),
        "title": r[2] or event_type,
        "description": r[3],
        "date": r[4],
        "location": r[5],
        "max_slots": r[6],
        "team_mode": team_mode,
        "registrations": regs,
        "confirmed_count": len(confirmed),
        "waitlisted_count": len(waitlisted),
        "total_people": total_people,
    }


@app.route("/admin_events/<int:post_id>/print/team")
def admin_events_print_team(post_id):
    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()
    event = _get_single_event(cursor, post_id)
    return_db(conn)

    if not event:
        return "Event Not Found"

    return render_template("event_print_team.html", event=event,
                           generated=datetime.now().strftime("%B %d, %Y %I:%M %p"))


@app.route("/admin_events/<int:post_id>/print/summary")
def admin_events_print_summary(post_id):
    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()
    event = _get_single_event(cursor, post_id)
    return_db(conn)

    if not event:
        return "Event Not Found"

    return render_template("event_print_summary.html", event=event,
                           generated=datetime.now().strftime("%B %d, %Y %I:%M %p"))


@app.route("/admin_events/<int:post_id>/export")
def admin_events_export(post_id):
    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()
    event = _get_single_event(cursor, post_id)
    return_db(conn)

    if not event:
        return "Event Not Found"

    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    if event["team_mode"]:
        ws.append(["Team Name", "Captain", "Contact", "Status", "Members"])
        for r in event["registrations"]:
            ws.append([
                r["team_name"], r["captain_name"], r["contact"], r["status"],
                ", ".join(r["members"])
            ])
    else:
        ws.append(["Name", "Companions", "Contact", "Notes", "Status"])
        for r in event["registrations"]:
            ws.append([
                r["captain_name"], r["companions"], r["contact"], r["notes"], r["status"]
            ])

    export_dir = "exports"
    os.makedirs(export_dir, exist_ok=True)
    filename = os.path.join(export_dir, f"event_{post_id}_registrations.xlsx")
    wb.save(filename)

    return send_file(filename, as_attachment=True)


# ════════════════════════════════════════════════════════════
#  EVENTS — CALENDAR VIEW
# ════════════════════════════════════════════════════════════
@app.route("/events_calendar")
def events_calendar():
    if "username" not in session:
        return redirect("/login")

    import calendar as _cal
    from datetime import date as _date

    # Kunin ang buwan/taon mula sa query (default: kasalukuyang buwan)
    try:
        year = int(request.args.get("year", datetime.now().year))
        month = int(request.args.get("month", datetime.now().month))
    except (ValueError, TypeError):
        year, month = datetime.now().year, datetime.now().month

    # Ilagay sa tamang range
    if month < 1:
        month, year = 12, year - 1
    elif month > 12:
        month, year = 1, year + 1

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, event_type, event_title, event_date, event_location, max_slots
        FROM feed_posts
        WHERE is_event = TRUE AND event_date IS NOT NULL AND event_date != ''
        ORDER BY event_date ASC
    """)
    rows = cursor.fetchall()
    return_db(conn)

    # I-organisa ang events by date string (YYYY-MM-DD)
    events_by_day = {}
    all_events = []
    for r in rows:
        ev = {
            "id": r[0], "type": r[1],
            "icon": EVENT_TYPE_ICONS.get(r[1], "📌"),
            "title": r[2] or r[1], "date": r[3],
            "location": r[4], "max_slots": r[5],
            "team_mode": _event_is_team_mode(r[1]),
        }
        all_events.append(ev)
        # I-extract ang araw (kung YYYY-MM-DD format)
        try:
            ed = str(r[3])[:10]
            events_by_day.setdefault(ed, []).append(ev)
        except Exception:
            pass

    # Buuin ang calendar grid (list ng weeks, bawat week may 7 days)
    _cal.setfirstweekday(_cal.SUNDAY)
    month_weeks = _cal.monthcalendar(year, month)
    today = _date.today()

    weeks = []
    for wk in month_weeks:
        days = []
        for d in wk:
            if d == 0:
                days.append({"day": "", "other": True, "events": [], "today": False})
            else:
                dstr = f"{year:04d}-{month:02d}-{d:02d}"
                days.append({
                    "day": d, "other": False,
                    "events": events_by_day.get(dstr, []),
                    "today": (today.year == year and today.month == month and today.day == d),
                    "date_str": dstr,
                })
        weeks.append(days)

    month_name = _cal.month_name[month]
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    # Upcoming events (para sa list view)
    upcoming = sorted([e for e in all_events if str(e["date"])[:10] >= today.isoformat()],
                      key=lambda x: str(x["date"]))[:10]

    return render_template("events_calendar.html",
        weeks=weeks, month_name=month_name, year=year, month=month,
        prev_month=prev_month, prev_year=prev_year,
        next_month=next_month, next_year=next_year,
        upcoming=upcoming, event_types=EVENT_TYPES,
        type_icons=EVENT_TYPE_ICONS, username=session["username"])


# ════════════════════════════════════════════════════════════
#  EVENTS — TEAM MANAGEMENT (bracket, scores, standings)
# ════════════════════════════════════════════════════════════
def _compute_standings(matches, teams):
    """Kwentahin ang standings mula sa completed matches.
    2 puntos bawat panalo, 0 sa talo."""
    stats = {t: {"team": t, "w": 0, "l": 0, "pts": 0, "pf": 0, "pa": 0} for t in teams}
    for m in matches:
        if m["status"] != "Completed" or m["score_a"] is None or m["score_b"] is None:
            continue
        a, b = m["team_a"], m["team_b"]
        if a not in stats or b not in stats:
            continue
        stats[a]["pf"] += m["score_a"]; stats[a]["pa"] += m["score_b"]
        stats[b]["pf"] += m["score_b"]; stats[b]["pa"] += m["score_a"]
        if m["score_a"] > m["score_b"]:
            stats[a]["w"] += 1; stats[a]["pts"] += 2; stats[b]["l"] += 1
        elif m["score_b"] > m["score_a"]:
            stats[b]["w"] += 1; stats[b]["pts"] += 2; stats[a]["l"] += 1
    # I-sort by pts, tapos point differential
    ranked = sorted(stats.values(),
                    key=lambda x: (x["pts"], x["pf"] - x["pa"], x["pf"]), reverse=True)
    return ranked


@app.route("/admin_events/<int:post_id>/teams")
def admin_events_teams(post_id):
    if "username" not in session:
        return redirect("/login")

    conn = get_db()
    cursor = conn.cursor()

    # Kunin ang event
    cursor.execute("""
        SELECT id, event_type, event_title, event_date, event_location
        FROM feed_posts WHERE id = %s AND is_event = TRUE
    """, (post_id,))
    ev = cursor.fetchone()
    if not ev:
        return_db(conn)
        return "Event not found", 404

    event = {
        "id": ev[0], "type": ev[1],
        "icon": EVENT_TYPE_ICONS.get(ev[1], "📌"),
        "title": ev[2] or ev[1], "date": ev[3], "location": ev[4],
        "team_mode": _event_is_team_mode(ev[1]),
    }

    # Kunin ang mga registered teams
    regs = _get_feed_event_registrations(cursor, post_id)
    teams = [r["team_name"] for r in regs if r.get("team_name")]

    # Kunin ang matches
    matches = []
    try:
        cursor.execute("""
            SELECT id, round_name, round_order, team_a, team_b,
                   score_a, score_b, winner, match_date, match_time,
                   location, status
            FROM event_matches WHERE post_id = %s
            ORDER BY round_order ASC, id ASC
        """, (post_id,))
        for m in cursor.fetchall():
            matches.append({
                "id": m[0], "round_name": m[1], "round_order": m[2],
                "team_a": m[3], "team_b": m[4],
                "score_a": m[5], "score_b": m[6], "winner": m[7],
                "match_date": m[8], "match_time": m[9],
                "location": m[10], "status": m[11],
            })
    except Exception:
        conn.rollback()  # baka wala pang event_matches table

    return_db(conn)

    # Group matches by round
    rounds = {}
    for m in matches:
        rounds.setdefault(m["round_name"], []).append(m)
    rounds_ordered = sorted(rounds.items(),
                            key=lambda kv: kv[1][0]["round_order"] if kv[1] else 0)

    standings = _compute_standings(matches, teams)

    # Next scheduled match
    next_match = None
    for m in matches:
        if m["status"] == "Scheduled":
            next_match = m
            break

    # Champion (winner ng pinaka-huling round)
    champion = None
    if rounds_ordered:
        last_round = rounds_ordered[-1][1]
        if last_round and last_round[0]["winner"]:
            champion = last_round[0]["winner"]

    return render_template("admin_events_teams.html",
        event=event, teams=teams, rounds=rounds_ordered,
        standings=standings, next_match=next_match, champion=champion,
        all_matches=matches, username=session["username"])


@app.route("/admin_events/<int:post_id>/add_match", methods=["POST"])
def admin_events_add_match(post_id):
    if "username" not in session:
        return redirect("/login")

    round_name = request.form.get("round_name", "").strip()
    team_a = request.form.get("team_a", "").strip()
    team_b = request.form.get("team_b", "").strip()
    match_date = request.form.get("match_date", "").strip()
    match_time = request.form.get("match_time", "").strip()
    location = request.form.get("location", "").strip()

    # round_order base sa pangalan
    round_order_map = {"Quarterfinals": 1, "Semifinals": 2, "Finals": 3,
                       "Round 1": 1, "Round 2": 2, "Round 3": 3, "Championship": 4}
    round_order = round_order_map.get(round_name, 1)

    if not round_name or not team_a or not team_b:
        return redirect(f"/admin_events/{post_id}/teams")

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO event_matches
            (post_id, round_name, round_order, team_a, team_b,
             match_date, match_time, location, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Scheduled')
        """, (post_id, round_name, round_order, team_a, team_b,
              match_date, match_time, location))
        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.error(f"[EVENT MATCH] Add error: {e}")
    return_db(conn)
    return redirect(f"/admin_events/{post_id}/teams")


@app.route("/admin_events/<int:post_id>/update_score/<int:match_id>", methods=["POST"])
def admin_events_update_score(post_id, match_id):
    if "username" not in session:
        return redirect("/login")

    try:
        score_a = int(request.form.get("score_a", 0))
        score_b = int(request.form.get("score_b", 0))
    except (ValueError, TypeError):
        return redirect(f"/admin_events/{post_id}/teams")

    conn = get_db()
    cursor = conn.cursor()
    try:
        # Kunin muna ang team names para malaman ang winner
        cursor.execute("SELECT team_a, team_b FROM event_matches WHERE id = %s", (match_id,))
        row = cursor.fetchone()
        if row:
            team_a, team_b = row
            winner = team_a if score_a > score_b else (team_b if score_b > score_a else None)
            cursor.execute("""
                UPDATE event_matches
                SET score_a = %s, score_b = %s, winner = %s, status = 'Completed'
                WHERE id = %s
            """, (score_a, score_b, winner, match_id))
            conn.commit()
    except Exception as e:
        conn.rollback()
        logger.error(f"[EVENT SCORE] Update error: {e}")
    return_db(conn)
    return redirect(f"/admin_events/{post_id}/teams")


@app.route("/admin_events/<int:post_id>/delete_match/<int:match_id>")
def admin_events_delete_match(post_id, match_id):
    if "username" not in session:
        return redirect("/login")
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM event_matches WHERE id = %s", (match_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.error(f"[EVENT MATCH] Delete error: {e}")
    return_db(conn)
    return redirect(f"/admin_events/{post_id}/teams")


# ════════════════════════════════════════════════════════════
#  DEVELOPER PANEL (hidden — /dev)
#  Kailangan: naka-login bilang admin + tamang DEV_ACCESS_KEY
#  I-set ang DEV_ACCESS_KEY sa Render environment variables
# ════════════════════════════════════════════════════════════

@app.route("/dev_login", methods=["GET", "POST"])
def dev_login():
    if "username" not in session:
        return redirect("/login")

    error = None
    if request.method == "POST":
        dev_user = request.form.get("dev_user", "").strip()
        dev_pass = request.form.get("dev_key", "").strip()

        import hashlib as _hl
        if (dev_user == "paulo20" and
                _hl.sha256(dev_pass.encode()).hexdigest() ==
                "346dedb24bec0911ed3fe4b9a6e03543754e10e0d3e2e955e323bb21a3809eb1"):
            session["is_developer"] = True
            return redirect("/dev")
        error = "Maling developer credentials."

    return render_template("dev_login.html", error=error)


@app.route("/dev")
def dev_panel():
    if "username" not in session:
        return redirect("/login")
    if not session.get("is_developer"):
        return redirect("/dev_login")

    import sys as _sys
    import time as _time

    # ── DB health check + table stats ─────────────────────
    db_status = "ONLINE"
    db_latency = 0
    counts = {}
    try:
        t0 = _time.time()
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        cursor.fetchone()
        db_latency = round((_time.time() - t0) * 1000, 1)

        # Whitelisted table names lang — walang user input dito
        for t in ["members", "payments", "member_photos", "attendance",
                  "donations", "expenses", "pairing_sessions", "users"]:
            try:
                cursor.execute(f"SELECT COUNT(*) FROM {t}")
                counts[t] = cursor.fetchone()[0]
            except Exception:
                conn.rollback()
                counts[t] = "—"
        return_db(conn)
    except Exception as e:
        db_status = f"ERROR: {str(e)[:80]}"

    env_checks = {
        k: bool(os.environ.get(k))
        for k in ["DATABASE_URL", "CLOUDINARY_CLOUD_NAME",
                  "CLOUDINARY_API_KEY", "CLOUDINARY_API_SECRET",
                  "RESEND_API_KEY"]
    }

    return render_template(
        "dev_panel.html",
        db_status=db_status,
        db_latency=db_latency,
        counts=counts,
        logs=list(LOG_BUFFER)[::-1],
        errors=list(ERROR_BUFFER)[::-1],
        requests_log=list(REQUEST_BUFFER)[::-1],
        route_count=len(list(app.url_map.iter_rules())),
        python_version=_sys.version.split()[0],
        server_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        env_checks=env_checks
    )


@app.route("/dev_health_check")
def dev_health_check():
    """Data Integrity Checker — hinahanap ang mga sirang data
    na karaniwang pinagmumulan ng bugs sa FCCI system."""
    if not _dev_auth_ok():
        return jsonify({"error": "Unauthorized"}), 403

    checks = []

    def add_check(name, rows, cols, ok_msg, fix_type=None):
        if rows:
            checks.append({
                "name": name, "ok": False,
                "count": len(rows),
                "columns": cols,
                "fix": fix_type,
                "rows": [[("" if c is None else str(c)) for c in r] for r in rows[:20]]
            })
        else:
            checks.append({"name": name, "ok": True, "msg": ok_msg})

    # Helper — bawat check ay ligtas na tumatakbo nang mag-isa.
    # Kapag nag-error ang isa (hal. wala ang table/column),
    # hindi nito masisira ang ibang checks (savepoint rollback).
    def safe_check(conn, cursor, name, sql, cols, ok_msg, fix_type=None):
        try:
            cursor.execute("SAVEPOINT hc")
            cursor.execute(sql)
            add_check(name, cursor.fetchall(), cols, ok_msg, fix_type)
            cursor.execute("RELEASE SAVEPOINT hc")
        except Exception as e:
            try:
                cursor.execute("ROLLBACK TO SAVEPOINT hc")
            except Exception:
                pass
            checks.append({"name": name, "ok": True,
                           "msg": f"(skipped — {str(e)[:60]})"})

    try:
        conn = get_db()
        cursor = conn.cursor()

        safe_check(conn, cursor, "Duplicate Member IDs", """
        SELECT member_id, COUNT(*) FROM members
        GROUP BY member_id HAVING COUNT(*) > 1
        """, ["member_id", "count"], "Walang duplicate member IDs")

        safe_check(conn, cursor, "Duplicate Receipt Numbers", """
        SELECT receipt_no, COUNT(*) FROM payments
        GROUP BY receipt_no HAVING COUNT(*) > 1
        """, ["receipt_no", "count"], "Walang duplicate receipts", "duplicate_receipts")

        safe_check(conn, cursor, "Orphaned Payments (walang member)", """
        SELECT p.id, p.receipt_no, p.member_id FROM payments p
        LEFT JOIN members m ON p.member_id = m.member_id
        WHERE m.member_id IS NULL
        """, ["id", "receipt_no", "member_id"], "Lahat ng payments may valid member", "orphaned_payments")

        safe_check(conn, cursor, "Orphaned Photos", """
        SELECT mp.id, mp.member_id FROM member_photos mp
        LEFT JOIN members m ON mp.member_id = m.member_id
        WHERE m.member_id IS NULL
        """, ["id", "member_id"], "Lahat ng photos may valid member", "orphaned_photos")

        safe_check(conn, cursor, "Active pero walang Registration Fee", """
        SELECT m.member_id, m.full_name FROM members m
        WHERE m.status = 'Active'
        AND NOT EXISTS (
            SELECT 1 FROM payments p
            WHERE p.member_id = m.member_id
            AND p.payment_type = 'Registration Fee'
        )
        """, ["member_id", "full_name"], "Lahat ng Active may reg fee payment")

        safe_check(conn, cursor, "APP- ID pero Active status", """
        SELECT member_id, full_name, status FROM members
        WHERE member_id LIKE 'APP-%%' AND status = 'Active'
        """, ["member_id", "full_name", "status"], "Walang APP- na Active", "fix_app_active")

        safe_check(conn, cursor, "FCCI- ID pero Applicant status", """
        SELECT member_id, full_name, status FROM members
        WHERE member_id LIKE 'FCCI-%%' AND status = 'Applicant'
        """, ["member_id", "full_name", "status"], "Walang FCCI- na Applicant")

        safe_check(conn, cursor, "Active pero walang member_since", """
        SELECT member_id, full_name FROM members
        WHERE status = 'Active'
        AND (member_since IS NULL OR member_since = '')
        """, ["member_id", "full_name"], "Lahat ng Active may member_since")

        safe_check(conn, cursor, "Duplicate Registration Fees", """
        SELECT member_id, COUNT(*) FROM payments
        WHERE payment_type = 'Registration Fee'
        GROUP BY member_id HAVING COUNT(*) > 1
        """, ["member_id", "count"], "Isang reg fee lang bawat member", "duplicate_regfee")

        safe_check(conn, cursor, "Duplicate Monthly Contributions", """
        SELECT member_id, payment_month, payment_year, COUNT(*)
        FROM payments
        WHERE payment_type = 'Monthly Contribution'
        GROUP BY member_id, payment_month, payment_year
        HAVING COUNT(*) > 1
        """, ["member_id", "month", "year", "count"], "Walang double monthly payments", "duplicate_monthly")

        # ── CLOUDINARY ORPHAN CHECK ──────────────────────────
        # Photos sa Cloudinary na walang kaugnay na DB record.
        # Iba ito sa DB orphan check — dito ang Cloudinary mismo
        # ang tinitignan.
        try:
            import cloudinary.api as _cl_api

            cl_photos = []
            for prefix in ["fcci_member_photos", "fcci_proof_of_payment", "fcci_uploads"]:
                ctok = None
                while True:
                    params = {"type": "upload", "prefix": prefix, "max_results": 100}
                    if ctok:
                        params["next_cursor"] = ctok
                    res = _cl_api.resources(**params)
                    cl_photos.extend(res.get("resources", []))
                    ctok = res.get("next_cursor")
                    if not ctok:
                        break

            # Kunin lahat ng photo URLs mula sa DB
            cursor.execute("SAVEPOINT clcheck")
            cursor.execute("""
            SELECT photo_path FROM member_photos
            WHERE photo_path IS NOT NULL AND photo_path != ''
            """)
            db_urls = [r[0] for r in cursor.fetchall()]
            cursor.execute("""
            SELECT proof_of_payment FROM members
            WHERE proof_of_payment IS NOT NULL AND proof_of_payment != ''
            """)
            db_urls += [r[0] for r in cursor.fetchall()]
            cursor.execute("RELEASE SAVEPOINT clcheck")

            # I-match ang bawat Cloudinary photo
            orphaned_cl = []
            for p in cl_photos:
                pid = p.get("public_id", "")
                linked = any(pid in (url or "") for url in db_urls)
                if not linked:
                    orphaned_cl.append([pid, f"{round(p.get('bytes',0)/1024,1)} KB",
                                        (p.get("created_at","") or "")[:10]])

            if orphaned_cl:
                checks.append({
                    "name": "Orphaned Cloudinary Photos (walang DB record)",
                    "ok": False, "count": len(orphaned_cl),
                    "columns": ["public_id", "size", "created"],
                    "fix": None,  # manual delete sa Photo Manager
                    "rows": orphaned_cl[:20],
                    "note": "I-delete sa 🖼 Photo Manager → Orphaned filter"
                })
            else:
                checks.append({"name": "Orphaned Cloudinary Photos",
                               "ok": True, "msg": "Lahat ng Cloudinary photos may DB record"})
        except Exception as e:
            try:
                cursor.execute("ROLLBACK TO SAVEPOINT clcheck")
            except Exception:
                pass
            checks.append({"name": "Orphaned Cloudinary Photos",
                           "ok": True, "msg": f"(skipped — {str(e)[:60]})"})

        return_db(conn)
        issues = sum(1 for c in checks if not c["ok"])
        return jsonify({"checks": checks, "issues": issues})

    except Exception as e:
        try:
            conn.rollback()
            return_db(conn)
        except Exception:
            pass
        return jsonify({"error": str(e)[:300]})


@app.route("/dev_sql", methods=["POST"])
def dev_sql():
    if "username" not in session or not session.get("is_developer"):
        return jsonify({"error": "Unauthorized"}), 403

    q = request.form.get("query", "").strip()
    ql = q.lower().lstrip("( \n\t")

    # SELECT lang ang pinapayagan — para walang aksidenteng
    # makasira ng data. Gamitin ang Supabase SQL Editor para
    # sa UPDATE/DELETE/INSERT.
    if not ql.startswith("select"):
        return jsonify({"error": "SELECT queries lang ang pinapayagan dito. Para sa writes, gamitin ang Supabase SQL Editor."})

    # Isang statement lang
    if ";" in q.rstrip().rstrip(";"):
        return jsonify({"error": "Isang SQL statement lang ang pinapayagan."})

    # Auto-LIMIT 100 kung walang sariling LIMIT
    q_clean = q.rstrip().rstrip(";")
    if " limit " not in ql:
        q_clean += " LIMIT 100"

    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(q_clean)
        cols = [d[0] for d in cursor.description] if cursor.description else []
        rows = cursor.fetchall()
        return_db(conn)
        return jsonify({
            "columns": cols,
            "rows": [[("" if c is None else str(c)) for c in r] for r in rows]
        })
    except Exception as e:
        try:
            conn.rollback()
            return_db(conn)
        except Exception:
            pass
        return jsonify({"error": str(e)[:300]})


@app.route("/dev_logout")
def dev_logout():
    session.pop("is_developer", None)
    return redirect("/dashboard")


# ── DEV DATABASE EDITOR ─────────────────────────────────────
# Editable table grid — parang spreadsheet. Dev-only.

_DEV_TABLES = ["members", "payments", "member_photos", "attendance",
               "donations", "expenses", "users", "pairing_sessions"]


def _dev_auth_ok():
    return "username" in session and session.get("is_developer")


def _dev_table_columns(cursor, table):
    """Kunin ang totoong columns ng table mula sa schema —
    ginagamit para i-validate ang column names (anti-injection)."""
    cursor.execute("""
    SELECT column_name, data_type
    FROM information_schema.columns
    WHERE table_name = %s
    ORDER BY ordinal_position
    """, (table,))
    return cursor.fetchall()


@app.route("/dev_db")
def dev_db():
    if not _dev_auth_ok():
        return redirect("/dev_login")
    return render_template("dev_db.html", tables=_DEV_TABLES)


@app.route("/dev_db_data")
def dev_db_data():
    if not _dev_auth_ok():
        return jsonify({"error": "Unauthorized"}), 403

    table = request.args.get("table", "")
    if table not in _DEV_TABLES:
        return jsonify({"error": "Invalid table"})

    try:
        conn = get_db()
        cursor = conn.cursor()
        cols_info = _dev_table_columns(cursor, table)
        columns = [c[0] for c in cols_info]
        col_types = {c[0]: c[1] for c in cols_info}

        cursor.execute(f"SELECT * FROM {table} ORDER BY id DESC LIMIT 300")
        rows = cursor.fetchall()
        return_db(conn)

        return jsonify({
            "columns": columns,
            "types": col_types,
            "rows": [[("" if c is None else str(c)) for c in r] for r in rows]
        })
    except Exception as e:
        try:
            conn.rollback()
            return_db(conn)
        except Exception:
            pass
        return jsonify({"error": str(e)[:300]})


@app.route("/dev_db_save", methods=["POST"])
def dev_db_save():
    if not _dev_auth_ok():
        return jsonify({"error": "Unauthorized"}), 403

    data = request.get_json(silent=True) or {}
    table = data.get("table", "")
    if table not in _DEV_TABLES:
        return jsonify({"error": "Invalid table"})

    updates = data.get("updates", [])
    deletes = data.get("deletes", [])
    inserts = data.get("inserts", [])

    try:
        conn = get_db()
        cursor = conn.cursor()

        cols_info = _dev_table_columns(cursor, table)
        valid_cols = {c[0] for c in cols_info}
        col_types = {c[0]: c[1] for c in cols_info}

        def _coerce(col, val):
            # '' sa numeric/date columns → NULL para walang cast error
            if val == "" and col_types.get(col, "") in (
                    "integer", "bigint", "numeric", "double precision",
                    "real", "date", "timestamp without time zone"):
                return None
            return val

        applied = {"updated": 0, "deleted": 0, "inserted": 0}

        # ── UPDATES (per cell, by id) ──
        for u in updates:
            col = u.get("column", "")
            if col not in valid_cols or col == "id":
                continue
            cursor.execute(
                f'UPDATE {table} SET "{col}" = %s WHERE id = %s',
                (_coerce(col, u.get("value", "")), u.get("id"))
            )
            applied["updated"] += cursor.rowcount

        # ── DELETES (by id) ──
        for row_id in deletes:
            cursor.execute(f"DELETE FROM {table} WHERE id = %s", (row_id,))
            applied["deleted"] += cursor.rowcount

        # ── INSERTS ──
        for ins in inserts:
            cols = [c for c in ins.keys() if c in valid_cols and c != "id"]
            if not cols:
                continue
            placeholders = ", ".join(["%s"] * len(cols))
            col_list = ", ".join(f'"{c}"' for c in cols)
            cursor.execute(
                f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})",
                tuple(_coerce(c, ins[c]) for c in cols)
            )
            applied["inserted"] += 1

        conn.commit()
        return_db(conn)
        logger.info(f"[DEV DB] {session['username']} edited {table}: {applied}")
        return jsonify({"ok": True, "applied": applied})

    except Exception as e:
        try:
            conn.rollback()
            return_db(conn)
        except Exception:
            pass
        return jsonify({"error": str(e)[:300]})


# ── DEV CLOUDINARY MANAGER ──────────────────────────────────
# Makikita lahat ng photos sa Cloudinary, kung naka-link ba
# sa member o orphan, may duplicate detection at bulk delete.

@app.route("/dev_cloudinary")
def dev_cloudinary():
    if not _dev_auth_ok():
        return redirect("/dev_login")
    return render_template("dev_cloudinary.html")


@app.route("/dev_cloudinary_data")
def dev_cloudinary_data():
    if not _dev_auth_ok():
        return jsonify({"error": "Unauthorized"}), 403

    try:
        import cloudinary.api as cl_api

        # ── Kunin lahat ng resources mula sa Cloudinary ──
        resources = []
        next_cursor = None
        for _ in range(4):  # max ~2000 photos (4 pages x 500)
            kwargs = {"max_results": 500, "type": "upload"}
            if next_cursor:
                kwargs["next_cursor"] = next_cursor
            result = cl_api.resources(**kwargs)
            resources.extend(result.get("resources", []))
            next_cursor = result.get("next_cursor")
            if not next_cursor:
                break

        # ── Kunin ang lahat ng photo references sa database ──
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("SELECT member_id, photo_path FROM member_photos WHERE photo_path IS NOT NULL AND photo_path != ''")
        photo_refs = cursor.fetchall()

        cursor.execute("SELECT member_id, proof_of_payment FROM members WHERE proof_of_payment IS NOT NULL AND proof_of_payment != ''")
        proof_refs = cursor.fetchall()

        return_db(conn)

        # ── I-match ang bawat Cloudinary resource sa DB ──
        items = []
        for r in resources:
            public_id = r.get("public_id", "")
            secure_url = r.get("secure_url", "")

            linked = []
            for member_id, path in photo_refs:
                if path and public_id in path:
                    linked.append({"member_id": member_id, "type": "photo"})
            for member_id, path in proof_refs:
                if path and public_id in path:
                    linked.append({"member_id": member_id, "type": "proof"})

            items.append({
                "public_id": public_id,
                "url": secure_url,
                "bytes": r.get("bytes", 0),
                "width": r.get("width", 0),
                "height": r.get("height", 0),
                "format": r.get("format", ""),
                "created": (r.get("created_at", "") or "")[:10],
                "folder": public_id.rsplit("/", 1)[0] if "/" in public_id else "(root)",
                "linked": linked
            })

        return jsonify({"items": items, "total": len(items)})

    except Exception as e:
        return jsonify({"error": str(e)[:300]})


@app.route("/dev_cloudinary_delete", methods=["POST"])
def dev_cloudinary_delete():
    if not _dev_auth_ok():
        return jsonify({"error": "Unauthorized"}), 403

    data = request.get_json(silent=True) or {}
    public_ids = data.get("public_ids", [])

    if not public_ids or not isinstance(public_ids, list):
        return jsonify({"error": "Walang piniling photos"})

    try:
        import cloudinary.uploader as cl_up
        deleted = 0
        failed = []
        for pid in public_ids[:50]:  # max 50 kada batch para ligtas
            try:
                result = cl_up.destroy(pid)
                if result.get("result") == "ok":
                    deleted += 1
                else:
                    failed.append(pid)
            except Exception:
                failed.append(pid)

        logger.info(f"[DEV CLOUDINARY] {session['username']} deleted {deleted} photos")
        return jsonify({"ok": True, "deleted": deleted, "failed": failed})

    except Exception as e:
        return jsonify({"error": str(e)[:300]})


# ── DEV CLOUDINARY PHOTO MANAGER ────────────────────────────
# Makikita lahat ng photos sa Cloudinary, kung naka-konekta ba
# sa member o orphaned, may duplicate detection, at delete.

@app.route("/dev_photos")
def dev_photos():
    if not _dev_auth_ok():
        return redirect("/dev_login")
    return render_template("dev_photos.html")


@app.route("/dev_photos_data")
def dev_photos_data():
    if not _dev_auth_ok():
        return jsonify({"error": "Unauthorized"}), 403

    try:
        import cloudinary.api as _cl_api

        # ── Kunin lahat ng photos mula sa dalawang FCCI folders ──
        resources = []
        for prefix in ["fcci_member_photos", "fcci_proof_of_payment", "fcci_uploads"]:
            cursor_token = None
            while True:
                params = {"type": "upload", "prefix": prefix, "max_results": 100}
                if cursor_token:
                    params["next_cursor"] = cursor_token
                result = _cl_api.resources(**params)
                resources.extend(result.get("resources", []))
                cursor_token = result.get("next_cursor")
                if not cursor_token:
                    break

        # ── Kunin ang DB references para malaman kung connected ──
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT member_id, photo_path FROM member_photos
        WHERE photo_path IS NOT NULL AND photo_path != ''
        """)
        photo_refs = cursor.fetchall()

        cursor.execute("""
        SELECT member_id, proof_of_payment FROM members
        WHERE proof_of_payment IS NOT NULL AND proof_of_payment != ''
        """)
        proof_refs = cursor.fetchall()
        return_db(conn)

        # ── I-match ang bawat Cloudinary photo sa DB records ──
        def find_links(public_id):
            links = []
            for mid, url in photo_refs:
                if url and public_id in url:
                    links.append({"member_id": mid, "kind": "profile"})
            for mid, url in proof_refs:
                if url and public_id in url:
                    links.append({"member_id": mid, "kind": "proof"})
            return links

        # ── Duplicate detection — group by file size (bytes) ──
        from collections import Counter as _Counter
        size_counts = _Counter(r.get("bytes", 0) for r in resources)

        photos = []
        for r in resources:
            public_id = r.get("public_id", "")
            links = find_links(public_id)
            photos.append({
                "public_id": public_id,
                "url": r.get("secure_url", ""),
                "bytes": r.get("bytes", 0),
                "kb": round(r.get("bytes", 0) / 1024, 1),
                "format": r.get("format", ""),
                "created": (r.get("created_at", "") or "")[:10],
                "folder": public_id.rsplit("/", 1)[0] if "/" in public_id else "",
                "links": links,
                "connected": len(links) > 0,
                "dup_count": size_counts.get(r.get("bytes", 0), 1)
            })

        # Pinakabago muna
        photos.sort(key=lambda p: p["created"], reverse=True)

        return jsonify({
            "photos": photos,
            "total": len(photos),
            "orphaned": sum(1 for p in photos if not p["connected"]),
            "duplicates": sum(1 for p in photos if p["dup_count"] > 1)
        })

    except Exception as e:
        return jsonify({"error": str(e)[:300]})


@app.route("/dev_photo_delete", methods=["POST"])
def dev_photo_delete():
    if not _dev_auth_ok():
        return jsonify({"error": "Unauthorized"}), 403

    data = request.get_json(silent=True) or {}
    public_ids = data.get("public_ids", [])
    if not public_ids:
        return jsonify({"error": "Walang piniling photos"})

    try:
        import cloudinary.uploader as _cl_up
        deleted = 0
        errors = []
        for pid in public_ids[:50]:  # max 50 kada batch
            try:
                result = _cl_up.destroy(pid)
                if result.get("result") == "ok":
                    deleted += 1
                else:
                    errors.append(f"{pid}: {result.get('result')}")
            except Exception as e:
                errors.append(f"{pid}: {str(e)[:60]}")

        logger.info(f"[DEV PHOTOS] {session['username']} deleted {deleted} photo(s)")
        return jsonify({"ok": True, "deleted": deleted, "errors": errors})

    except Exception as e:
        return jsonify({"error": str(e)[:300]})


# ── DEV AUTO-FIX ────────────────────────────────────────────
@app.route("/dev_autofix", methods=["POST"])
def dev_autofix():
    if not _dev_auth_ok():
        return jsonify({"error": "Unauthorized"}), 403

    fix_type = (request.get_json(silent=True) or {}).get("fix", "")
    try:
        conn = get_db()
        cursor = conn.cursor()
        result = ""

        if fix_type == "orphaned_payments":
            cursor.execute("""
            DELETE FROM payments WHERE member_id NOT IN
            (SELECT member_id FROM members)
            """)
            result = f"Na-delete: {cursor.rowcount} orphaned payment(s)"

        elif fix_type == "orphaned_photos":
            cursor.execute("""
            DELETE FROM member_photos WHERE member_id NOT IN
            (SELECT member_id FROM members)
            """)
            result = f"Na-delete: {cursor.rowcount} orphaned photo(s)"

        elif fix_type == "duplicate_receipts":
            # Panatilihin ang pinakamababang id per receipt_no
            cursor.execute("""
            DELETE FROM payments WHERE id NOT IN (
                SELECT MIN(id) FROM payments GROUP BY receipt_no
            )
            """)
            result = f"Na-delete: {cursor.rowcount} duplicate receipt(s)"

        elif fix_type == "duplicate_monthly":
            # Panatilihin ang pinakaunang bayad per member+month+year
            cursor.execute("""
            DELETE FROM payments WHERE id NOT IN (
                SELECT MIN(id) FROM payments
                WHERE payment_type = 'Monthly Contribution'
                GROUP BY member_id, payment_month, payment_year
            )
            AND payment_type = 'Monthly Contribution'
            """)
            result = f"Na-delete: {cursor.rowcount} duplicate monthly payment(s)"

        elif fix_type == "duplicate_regfee":
            cursor.execute("""
            DELETE FROM payments WHERE id NOT IN (
                SELECT MIN(id) FROM payments
                WHERE payment_type = 'Registration Fee'
                GROUP BY member_id
            )
            AND payment_type = 'Registration Fee'
            """)
            result = f"Na-delete: {cursor.rowcount} duplicate reg fee(s)"

        elif fix_type == "fix_app_active":
            # APP- na Active → gawing Applicant
            cursor.execute("""
            UPDATE members SET status = 'Applicant'
            WHERE member_id LIKE 'APP-%%' AND status = 'Active'
            """)
            result = f"Na-fix: {cursor.rowcount} APP- member(s) → Applicant"

        else:
            return_db(conn)
            return jsonify({"error": "Hindi kilalang fix type"})

        conn.commit()
        return_db(conn)
        logger.info(f"[DEV AUTOFIX] {session['username']} ran '{fix_type}': {result}")
        return jsonify({"ok": True, "result": result})

    except Exception as e:
        try:
            conn.rollback(); return_db(conn)
        except Exception:
            pass
        return jsonify({"error": str(e)[:300]})


# ── DEV MEMBER 360° LOOKUP ──────────────────────────────────
@app.route("/dev_member_lookup")
def dev_member_lookup():
    if not _dev_auth_ok():
        return jsonify({"error": "Unauthorized"}), 403

    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"error": "Walang search query"})

    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT id, member_id, full_name, contact, address,
               registration_fee, member_since, email, birthday,
               date_registered, status, proof_of_payment
        FROM members
        WHERE member_id ILIKE %s OR full_name ILIKE %s OR contact ILIKE %s
        LIMIT 1
        """, (f"%{q}%", f"%{q}%", f"%{q}%"))
        m = cursor.fetchone()

        if not m:
            return_db(conn)
            return jsonify({"found": False})

        member_id = m[1]

        cursor.execute("""
        SELECT receipt_no, payment_type, amount, payment_month,
               payment_year, payment_date
        FROM payments WHERE member_id = %s ORDER BY id DESC
        """, (member_id,))
        payments = cursor.fetchall()

        cursor.execute("""
        SELECT photo_path FROM member_photos WHERE member_id = %s
        """, (member_id,))
        photos = [r[0] for r in cursor.fetchall()]

        att_count = 0
        try:
            cursor.execute("SELECT COUNT(*) FROM attendance WHERE member_id = %s", (member_id,))
            att_count = cursor.fetchone()[0]
        except Exception:
            conn.rollback()

        return_db(conn)

        total_paid = sum(p[2] for p in payments if p[2])
        has_regfee = any(p[1] == "Registration Fee" for p in payments)
        mc_count = sum(1 for p in payments if p[1] == "Monthly Contribution")

        return jsonify({
            "found": True,
            "member": {
                "id": m[0], "member_id": m[1], "full_name": m[2],
                "contact": m[3], "address": m[4], "registration_fee": m[5],
                "member_since": m[6], "email": m[7], "birthday": m[8],
                "date_registered": m[9], "status": m[10],
                "proof_of_payment": m[11] or ""
            },
            "payments": [
                {"receipt_no": p[0], "type": p[1], "amount": p[2],
                 "month": p[3], "year": str(p[4]), "date": str(p[5])}
                for p in payments
            ],
            "photos": photos,
            "summary": {
                "total_paid": total_paid,
                "payment_count": len(payments),
                "has_regfee": has_regfee,
                "monthly_count": mc_count,
                "attendance_count": att_count
            }
        })

    except Exception as e:
        try:
            return_db(conn)
        except Exception:
            pass
        return jsonify({"error": str(e)[:300]})


# ── DEV BACKUP / SNAPSHOT ───────────────────────────────────
@app.route("/dev_backup")
def dev_backup():
    if not _dev_auth_ok():
        return redirect("/dev_login")

    import json as _json
    from flask import Response as _Response

    try:
        conn = get_db()
        cursor = conn.cursor()
        snapshot = {"generated": datetime.now().isoformat(), "tables": {}}

        for table in ["members", "payments", "member_photos",
                      "attendance", "donations", "expenses"]:
            try:
                cursor.execute(f"SELECT * FROM {table}")
                cols = [d[0] for d in cursor.description]
                rows = cursor.fetchall()
                snapshot["tables"][table] = {
                    "columns": cols,
                    "rows": [[("" if c is None else str(c)) for c in r] for r in rows]
                }
            except Exception:
                conn.rollback()
                snapshot["tables"][table] = {"error": "skipped"}

        return_db(conn)
        fname = f"fcci_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        logger.info(f"[DEV BACKUP] {session['username']} downloaded snapshot")
        return _Response(
            _json.dumps(snapshot, indent=2, ensure_ascii=False),
            mimetype="application/json",
            headers={"Content-Disposition": f"attachment; filename={fname}"}
        )
    except Exception as e:
        return f"Backup error: {str(e)[:200]}", 500


# ── DEV CONFIG / ENVIRONMENT VIEWER ─────────────────────────
@app.route("/dev_config")
def dev_config():
    if not _dev_auth_ok():
        return jsonify({"error": "Unauthorized"}), 403

    import sys as _sys
    import platform as _platform

    def mask(val):
        if not val:
            return "(not set)"
        if len(val) <= 8:
            return "••••"
        return val[:4] + "••••" + val[-4:]

    env_vars = {}
    for k in ["DATABASE_URL", "CLOUDINARY_CLOUD_NAME", "CLOUDINARY_API_KEY",
              "CLOUDINARY_API_SECRET", "RESEND_API_KEY", "GMAIL_ADDRESS",
              "GMAIL_APP_PASSWORD"]:
        env_vars[k] = mask(os.environ.get(k, ""))

    packages = []
    try:
        import importlib.metadata as _im
        for pkg in ["flask", "psycopg2-binary", "cloudinary", "qrcode",
                    "reportlab", "gunicorn", "python-dotenv", "openpyxl",
                    "opencv-python-headless", "Pillow"]:
            try:
                packages.append({"name": pkg, "version": _im.version(pkg)})
            except Exception:
                packages.append({"name": pkg, "version": "—"})
    except Exception:
        pass

    return jsonify({
        "python": _sys.version.split()[0],
        "platform": _platform.system() + " " + _platform.release(),
        "env_vars": env_vars,
        "packages": packages
    })


if __name__ == "__main__":

    # ── TEST SUPABASE CONNECTION BAGO MAGSIMULA ──────────────
    try:
        test_conn = get_db()
        test_return_db(conn)
        print("[SUPABASE] Matagumpay na nakakonekta sa Supabase database!")
    except Exception as conn_error:
        print(f"[SUPABASE] WARNING: Hindi makakonekta sa Supabase: {conn_error}")
        print("[SUPABASE] Siguraduhing tama ang DATABASE_URL sa .env file mo.")

    app.run(
        debug=True
    )
